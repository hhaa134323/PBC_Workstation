"""v7: 生成测试数据脚本。

审计员反馈 #8：造一套完整的测试数据，让审计员拿来创建项目测试。

生成内容：
1. PBC 清单 Excel（10 项，覆盖 5 个一级分类、3 个实体）
2. 10 个测试文件（PDF/Excel/TXT，文件名带编号前缀便于 filename-match）
3. 1 个穿行测试文件夹（含 3 个文件，验证整目录归档）

用法：
    python scripts/generate_test_data.py
    # 输出到 data/test_data_package/
    # 用户下载这个包，解压后用 PBC 模板导入清单 + 客户文件夹指向解压目录
"""
from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# 输出目录
OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data" / "test_data_package"


# 10 项 PBC 清单数据（覆盖 5 个一级分类、3 个实体）
# 字段：item_id, category, subject, description, file_format, priority,
#        raised_at, expected_by, overdue_days, status_raw, remark, entity,
#        confidence, file_path, required_period
PBC_ITEMS = [
    {
        "item_id": "历-1", "category": "历史沿革", "subject": "公司治理",
        "description": "股权架构图", "file_format": "PDF",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-15",
        "entity": "集团合并", "required_period": "2024年度",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "历-2", "category": "历史沿革", "subject": "公司治理",
        "description": "公司章程及修订决议", "file_format": "PDF",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-15",
        "entity": "集团合并", "required_period": "2024年度",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "财-1", "category": "财务报表", "subject": "合并报表",
        "description": "三年一期合并资产负债表", "file_format": "Excel",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-20",
        "entity": "集团合并", "required_period": "2023年度/2024年度/2025年度/2026年一季度",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "财-2", "category": "财务报表", "subject": "个别报表",
        "description": "ABC子公司单体利润表", "file_format": "Excel",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-20",
        "entity": "ABC子公司", "required_period": "2023年度/2024年度/2025年度",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "存-1", "category": "存货", "subject": "存货盘点",
        "description": "存货盘点表", "file_format": "Excel",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-10",
        "entity": "ABC子公司", "required_period": "2025年12月31日",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "存-4", "category": "存货", "subject": "存货盘点",
        "description": "存货盘点计划", "file_format": "PDF",
        "priority": "中", "raised_at": "2026-07-01", "expected_by": "2026-07-10",
        "entity": "ABC子公司", "required_period": "2025年12月31日",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "货-1", "category": "货币资金", "subject": "银行存款",
        "description": "银行流水", "file_format": "Excel",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-15",
        "entity": "ABC子公司", "required_period": "2023年度/2024年度/2025年度",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "货-2", "category": "货币资金", "subject": "银行存款",
        "description": "银行对账单", "file_format": "PDF",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-15",
        "entity": "DEF子公司", "required_period": "2025年12月31日",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "销-1", "category": "收入", "subject": "主营业务收入",
        "description": "销售合同台账", "file_format": "Excel",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-25",
        "entity": "ABC子公司", "required_period": "2023年度/2024年度/2025年度",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
    {
        "item_id": "综-1", "category": "综合性资料", "subject": "年度审计",
        "description": "2024年度审计报告", "file_format": "PDF",
        "priority": "高", "raised_at": "2026-07-01", "expected_by": "2026-07-20",
        "entity": "集团合并", "required_period": "2024年度",
        "status_raw": "未提供", "remark": "", "overdue_days": 0,
    },
]


# 10 个测试文件（文件名带编号前缀，便于 filename-match）
# (file_name, content_type, content_text)
TEST_FILES = [
    ("历-1_股权架构图.pdf", "pdf", "ABC 集团股权架构图：母公司 ABC 集团持有 ABC子公司 100%、DEF子公司 80%。期间：2024年度。"),
    ("历-2_公司章程.pdf", "pdf", "ABC 集团公司章程（2024年修订版）：注册资本、股东会、董事会、监事会。"),
    ("财-1_合并资产负债表.xlsx", "excel", "合并资产负债表 2023-2025：资产、负债、所有者权益。"),
    ("财-2_ABC子公司利润表.xlsx", "excel", "ABC子公司单体利润表 2023-2025：营业收入、营业成本、净利润。"),
    ("存-1_存货盘点表.xlsx", "excel", "存货盘点表 2025年12月31日：原材料、在产品、产成品。"),
    ("存-4_存货盘点计划.pdf", "pdf", "存货盘点计划：盘点时间 2025-12-31，盘点范围、盘点人员。"),
    ("货-1_银行流水.xlsx", "excel", "ABC子公司银行流水 2023-2025：日期、摘要、收入、支出、余额。"),
    ("货-2_银行对账单.pdf", "pdf", "DEF子公司银行对账单 2025-12-31：账户余额、交易明细。"),
    ("销-1_销售合同台账.xlsx", "excel", "销售合同台账 2023-2025：合同编号、客户、金额、签订日期。"),
    ("综-1_2024年度审计报告.pdf", "pdf", "ABC 集团 2024年度审计报告：审计意见、关键审计事项、财务报表。"),
]

# 穿行测试文件夹（含 3 个文件，验证整目录归档）
WALKTHROUGH_FOLDER = "穿行测试_销售收款控制"
WALKTHROUGH_FILES = [
    ("B0206_系统截图.pdf", "pdf", "销售订单系统截图：订单创建、审批、发货流程。"),
    ("B0207_纸质签字.pdf", "pdf", "销售合同纸质签字扫描件：销售经理、财务经理签字。"),
    ("B0208_银行回单.pdf", "pdf", "银行回单：客户付款回单 3 张。"),
]


def _make_pdf(text: str) -> bytes:
    """生成一个最小 PDF 文件（含 text）。"""
    # 简化版：用 reportlab 不一定装，生成纯文本伪 PDF
    # 实际 PBC 应用 pdfplumber 解析，能从内容流提取 Text
    # 这里用最简 PDF 结构
    content = f"BT /F1 12 Tf 72 720 Td ({text}) Tj ET"
    pdf = (
        b"%PDF-1.4\n"
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj\n"
        b"4 0 obj<</Length " + str(len(content)).encode() + b">>stream\n" +
        content.encode() + b"\nendstream\nendobj\n"
        b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helbert>>endobj\n"
        b"xref\n0 6\n0000000000 65535 f \n"
        b"trailer<</Size 6/Root 1 0 R>>\nstartxref\n0\n%%EOF\n"
    )
    return pdf


def _make_excel(content_text: str) -> bytes:
    """生成一个最小 Excel 文件（含 content_text）。"""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(1, 1, content_text)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_txt(text: str) -> bytes:
    return text.encode("utf-8")


def _make_file(file_name: str, content_type: str, content_text: str) -> bytes:
    if content_type == "pdf":
        return _make_pdf(content_text)
    if content_type == "excel":
        return _make_excel(content_text)
    return _make_txt(content_text)


def generate_pbc_template(output_path: Path) -> None:
    """生成 PBC 清单 Excel（15 列表头 + 必填标注，复用 db._create_empty_pbc_xlsx 风格）。"""
    from openpyxl import Workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "PBC清单"
    headers = [
        ("资料编号", True), ("一级分类", True), ("相关科目", False),
        ("问题/需求描述", True), ("格式", False), ("优先级", False),
        ("提出时间", False), ("期望提供日期", True), ("需求期间", True),
        ("逾期天数", False), ("资料提供情况", False), ("备注", False),
        ("实体归属", True), ("置信度", False), ("文件路径", False),
    ]
    required_fill = PatternFill(start_color="FCEBEB", end_color="FCEBEB", fill_type="solid")
    optional_fill = PatternFill(start_color="F1EFE8", end_color="F1EFE8", fill_type="solid")
    for i, (h, required) in enumerate(headers, 1):
        cell = ws.cell(1, i, ("* " if required else "") + h)
        cell.font = Font(bold=True, color="A32D2D" if required else "5F5E5A")
        cell.fill = required_fill if required else optional_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = [12, 14, 16, 40, 8, 8, 14, 14, 30, 10, 16, 24, 14, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    dv = DataValidation(type="list", formula1='"未提供,已提供，审核中,已提供,不适用"', allow_blank=True)
    dv.add("J2:J1000")
    ws.add_data_validation(dv)
    ws.freeze_panes = "B2"

    # 填 10 行测试数据
    keys = ["item_id", "category", "subject", "description", "file_format", "priority",
            "raised_at", "expected_by", "required_period", "overdue_days", "status_raw",
            "remark", "entity", "confidence", "file_path"]
    for ri, item in enumerate(PBC_ITEMS, 2):
        for ci, k in enumerate(keys, 1):
            ws.cell(ri, ci, item.get(k, ""))

    wb.save(str(output_path))
    wb.close()


def generate_test_files(client_folder: Path) -> None:
    """生成 10 个测试文件 + 1 个穿行测试文件夹（含 3 文件）。"""
    client_folder.mkdir(parents=True, exist_ok=True)
    for fname, ctype, ctext in TEST_FILES:
        data = _make_file(fname, ctype, ctext)
        (client_folder / fname).write_bytes(data)
    # 穿行测试文件夹
    wt_folder = client_folder / WALKTHROUGH_FOLDER
    wt_folder.mkdir(parents=True, exist_ok=True)
    for fname, ctype, ctext in WALKTHROUGH_FILES:
        data = _make_file(fname, ctype, ctext)
        (wt_folder / fname).write_bytes(data)


def main() -> None:
    print(f"生成测试数据包到: {OUTPUT_DIR}")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 1. PBC 清单 Excel
    pbc_path = OUTPUT_DIR / "01_PBC_List_测试.xlsx"
    generate_pbc_template(pbc_path)
    print(f"  PBC 清单: {pbc_path}（{len(PBC_ITEMS)} 项）")

    # 2. 客户共享文件夹（含 10 文件 + 1 穿行测试文件夹）
    client_folder = OUTPUT_DIR / "客户共享文件夹"
    generate_test_files(client_folder)
    print(f"  客户共享文件夹: {client_folder}（{len(TEST_FILES)} 文件 + 1 穿行测试文件夹含 {len(WALKTHROUGH_FILES)} 文件）")

    # 3. README
    readme = OUTPUT_DIR / "README_测试数据说明.txt"
    readme.write_text(
        "PBC 智能管理工作站 · 测试数据包\n"
        "========================================\n\n"
        "使用步骤：\n"
        "1. 双击 PBC-Agent.exe 启动工作站\n"
        "2. 创建新项目\n"
        "3. 项目向导第 2 步：上传 01_PBC_List_测试.xlsx 导入清单\n"
        "4. 项目向导第 3 步：配置客户共享文件夹路径指向 本目录/客户共享文件夹/\n"
        "5. 进入项目，点「扫描新文件」看 AI 自动分类归档\n"
        "6. 扫描完成后查看「文件流向图」+「归档目录」\n\n"
        "测试覆盖：\n"
        f"- {len(PBC_ITEMS)} 项 PBC 清单（5 个一级分类、3 个实体）\n"
        f"- {len(TEST_FILES)} 个客户文件（文件名带编号前缀，验证 filename-match 快路径）\n"
        f"- 1 个穿行测试文件夹（{WALKTHROUGH_FOLDER}，含 {len(WALKTHROUGH_FILES)} 文件，验证整目录归档）\n\n"
        "关键验证点：\n"
        "- AI 文件名匹配：历-1_股权架构图.pdf 应匹配到 item_id=历-1\n"
        "- AI 内容分类：综-1_2024年度审计报告.pdf 应匹配到 item_id=综-1\n"
        "- 整目录归档：穿行测试文件夹应整目录归档到归档根目录下\n"
        "- 期间检查：财-1 需求期间 2023/2024/2025，文件内容应覆盖\n"
        "- 归档路径：归档到 PBC归档/历史沿革/历-1_股权架构图_2024_v1.pdf\n\n"
        f"生成时间：{__import__('datetime').datetime.now().isoformat(timespec='seconds')}\n",
        encoding="utf-8",
    )
    print(f"  说明文档: {readme}")

    print("\n测试数据包生成完成。")
    print(f"用户下载路径：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()
