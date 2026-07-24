"""生成混合交付形态测试文件——五种形态同时存在 + 特殊情况。

与之前两版测试数据的区别：
- 一个客户文件夹里同时包含5种交付形态（子目录结构各异）
- 验证 Layer 0（结构感知）能否识别不同形态并提取线索
- 覆盖特殊情况 S1（穿行测试整包）S3（一项拆多文件）S5（清单外文件）

目录结构：
  客户共享文件夹_混合形态/
    ├── 历-1_股权架构图.pdf              ← 形态1: 文件名带编号
    ├── 销-1_销售合同台账.xlsx
    ├── 财务报表/                        ← 形态2: 按类别分文件夹
    │   ├── 合并资产负债表.xlsx
    │   └── 子公司利润表.xlsx
    ├── 存货盘点/
    │   └── 盘点表.xlsx
    ├── 盘点计划.pdf                     ← 形态3: 扁平有意义
    ├── DEF银行对账单.pdf
    ├── ABC集团章程_修订版.pdf
    ├── 2023年度/                        ← 形态4: 按年度分文件夹
    │   └── 银行流水.xlsx
    ├── 2024年度/
    │   └── 银行流水.xlsx
    ├── 2025年度/
    │   └── 银行流水.xlsx
    ├── 新建文档(1).pdf                  ← 形态5: 混乱命名（内容=审计报告→综-1）
    ├── 扫描件001.pdf                    ← 形态5: 混乱命名（内容=员工花名册→清单外S5）
    └── 穿行测试_销售收款控制/           ← S1: 穿行测试整包
        ├── 系统截图_订单审批.pdf
        ├── 合同签字件.pdf
        └── 银行回单.pdf
"""
from __future__ import annotations

from pathlib import Path
import io

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data" / "test_data_package" / "客户共享文件夹_混合形态"


def _make_pdf(text: str) -> bytes:
    content = f"BT /F1 12 Tf 72 720 Td ({text}) Tj ET"
    pdf = (
        b"%PDF-1.4\n"
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj\n"
        b"4 0 obj<</Length " + str(len(content)).encode() + b">>stream\n" +
        content.encode("latin-1", errors="replace") + b"\nendstream\nendobj\n"
        b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\n"
        b"xref\n0 6\n0000000000 65535 f \n"
        b"trailer<</Size 6/Root 1 0 R>>\nstartxref\n0\n%%EOF\n"
    )
    return pdf


def _make_excel(content_text: str) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row_idx, line in enumerate(content_text.split("\n"), 1):
        cells = line.split("\t")
        for col_idx, cell_text in enumerate(cells, 1):
            ws.cell(row_idx, col_idx, cell_text)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_file(path: Path, content: str, ctype: str = "pdf") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if ctype == "pdf":
        path.write_bytes(_make_pdf(content))
    elif ctype == "excel":
        path.write_bytes(_make_excel(content))
    else:
        path.write_text(content, encoding="utf-8")


# ── PBC 清单项（20项，其中11项有对应客户文件，9项未提供→模拟缺料）
# 新增 doc_name（资料名称）字段，description 改为更详细的需求说明
PBC_ITEMS = [
    # ── 历史沿革 ──
    {"item_id": "历-1", "category": "历史沿革", "doc_name": "股权架构图", "description": "ABC集团股权架构图，含母子公司持股比例", "entity": "集团合并", "required_period": "2024年度"},
    {"item_id": "历-2", "category": "历史沿革", "doc_name": "公司章程", "description": "ABC集团公司章程及修订决议，含注册资本、股东会、董事会、监事会", "entity": "集团合并", "required_period": "2024年度"},
    {"item_id": "历-3", "category": "历史沿革", "doc_name": "营业执照", "description": "ABC集团及子公司营业执照副本", "entity": "集团合并", "required_period": "2024年度"},
    {"item_id": "历-4", "category": "历史沿革", "doc_name": "高管名单", "description": "ABC集团高管名单及简历，含董事长、总经理、财务总监", "entity": "集团合并", "required_period": "2025年度"},
    # ── 财务报表 ──
    {"item_id": "财-1", "category": "财务报表", "doc_name": "合并资产负债表", "description": "三年一期合并资产负债表（2023-2025年度+2026一季度）", "entity": "集团合并", "required_period": "2023年度/2024年度/2025年度/2026年一季度"},
    {"item_id": "财-2", "category": "财务报表", "doc_name": "子公司利润表", "description": "ABC子公司单体利润表（2023-2025年度）", "entity": "ABC子公司", "required_period": "2023年度/2024年度/2025年度"},
    {"item_id": "财-3", "category": "财务报表", "doc_name": "现金流量表", "description": "三年一期合并现金流量表", "entity": "集团合并", "required_period": "2023年度/2024年度/2025年度/2026年一季度"},
    # ── 存货 ──
    {"item_id": "存-1", "category": "存货", "doc_name": "存货盘点表", "description": "存货盘点表（2025年12月31日，含原材料/在产品/产成品）", "entity": "ABC子公司", "required_period": "2025年12月31日"},
    {"item_id": "存-4", "category": "存货", "doc_name": "存货盘点计划", "description": "存货盘点计划，含盘点范围/时间/人员/方法", "entity": "ABC子公司", "required_period": "2025年12月31日"},
    {"item_id": "存-5", "category": "存货", "doc_name": "存货跌价准备计算表", "description": "存货跌价准备计算表，含可变现净值评估", "entity": "ABC子公司", "required_period": "2025年度"},
    # ── 货币资金 ──
    {"item_id": "货-1", "category": "货币资金", "doc_name": "银行流水", "description": "ABC子公司银行流水明细（2023-2025年度）", "entity": "ABC子公司", "required_period": "2023年度/2024年度/2025年度"},
    {"item_id": "货-2", "category": "货币资金", "doc_name": "银行对账单", "description": "DEF子公司银行对账单（2025年12月31日）", "entity": "DEF子公司", "required_period": "2025年12月31日"},
    {"item_id": "货-3", "category": "货币资金", "doc_name": "银行存款余额调节表", "description": "各银行账户存款余额调节表", "entity": "集团合并", "required_period": "2025年12月31日"},
    # ── 收入 ──
    {"item_id": "销-1", "category": "收入", "doc_name": "销售合同台账", "description": "销售合同台账（2023-2025年度，含合同金额/客户/签订日期）", "entity": "ABC子公司", "required_period": "2023年度/2024年度/2025年度"},
    {"item_id": "销-2", "category": "收入", "doc_name": "收入明细账", "description": "主营业务收入明细账，按月份和产品分类", "entity": "ABC子公司", "required_period": "2023年度/2024年度/2025年度"},
    # ── 综合性资料 ──
    {"item_id": "综-1", "category": "综合性资料", "doc_name": "2024年度审计报告", "description": "ABC集团2024年度审计报告，含审计意见和关键审计事项", "entity": "集团合并", "required_period": "2024年度"},
    {"item_id": "综-2", "category": "综合性资料", "doc_name": "管理建议书", "description": "前任审计师管理建议书及整改情况", "entity": "集团合并", "required_period": "2024年度"},
    # ── 穿行测试 ──
    {"item_id": "穿-1", "category": "穿行测试", "doc_name": "销售收款穿行测试", "description": "销售收款内部控制穿行测试资料包（系统截图+签字件+回单）", "entity": "ABC子公司", "required_period": "2025年度"},
    {"item_id": "穿-2", "category": "穿行测试", "doc_name": "采购付款穿行测试", "description": "采购付款内部控制穿行测试资料包（系统截图+审批单+回单）", "entity": "ABC子公司", "required_period": "2025年度"},
]

# 标记哪些 item 有对应客户文件（其余=未提供，模拟缺料）
HAS_FILE = {"历-1", "历-2", "财-1", "财-2", "存-1", "存-4", "货-1", "货-2", "销-1", "综-1", "穿-1"}


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"生成混合形态测试文件到: {OUTPUT_DIR}\n")

    # ════════════════════════════════════════════
    # 形态1: 完全规范（文件名带编号前缀）
    # ════════════════════════════════════════════
    print("── 形态1: 完全规范（文件名带编号前缀）──")
    write_file(OUTPUT_DIR / "历-1_股权架构图.pdf",
        "ABC集团股权架构图\n母公司：ABC集团股份有限公司\n"
        "子公司1：ABC制造有限公司（持股100%）\n"
        "子公司2：ABC商贸有限公司（持股100%）\n"
        "子公司3：ABC科技有限公司（持股100%）\n"
        "子公司4：DEF实业有限公司（持股80%）\n"
        "时间：2024年度")
    print("  历-1_股权架构图.pdf        → 历-1 (filename-match)")

    write_file(OUTPUT_DIR / "销-1_销售合同台账.xlsx",
        "销售合同台账\n单位：ABC制造有限公司\n期间：2023年度-2025年度\n"
        "合同编号\t客户名称\t合同金额\t签订日期\n"
        "SC-2023-001\t深圳XX电子\t520,000\t2023-02-15\n"
        "SC-2024-001\t深圳XX电子\t580,000\t2024-01-10\n"
        "SC-2025-001\t深圳XX电子\t630,000\t2025-02-14", "excel")
    print("  销-1_销售合同台账.xlsx      → 销-1 (filename-match)")

    # ════════════════════════════════════════════
    # 形态2: 按类别分文件夹
    # ════════════════════════════════════════════
    print("\n── 形态2: 按类别分文件夹 ──")
    write_file(OUTPUT_DIR / "财务报表" / "合并资产负债表.xlsx",
        "ABC集团合并资产负债表\n单位：元\n"
        "科目\\年度\t2023-12-31\t2024-12-31\t2025-12-31\t2026-03-31\n"
        "货币资金\t1,200,000\t1,500,000\t1,800,000\t1,900,000\n"
        "应收账款\t800,000\t950,000\t1,100,000\t1,050,000\n"
        "存货\t600,000\t700,000\t750,000\t780,000\n"
        "固定资产\t2,000,000\t2,200,000\t2,400,000\t2,380,000\n"
        "资产总计\t4,900,000\t5,630,000\t6,310,000\t6,365,000\n"
        "负债合计\t2,300,000\t2,550,000\t2,700,000\t2,790,000\n"
        "所有者权益合计\t2,600,000\t3,080,000\t3,610,000\t3,575,000", "excel")
    print("  /财务报表/合并资产负债表.xlsx → 财-1 (文件夹名=类别→缩小候选→关键词)")

    write_file(OUTPUT_DIR / "财务报表" / "子公司利润表.xlsx",
        "ABC制造有限公司单体利润表\n单位：元\n"
        "项目\\年度\t2023年度\t2024年度\t2025年度\n"
        "营业收入\t3,000,000\t3,500,000\t4,000,000\n"
        "营业成本\t2,100,000\t2,450,000\t2,800,000\n"
        "营业利润\t370,000\t480,000\t590,000\n"
        "净利润\t279,750\t361,500\t442,875", "excel")
    print("  /财务报表/子公司利润表.xlsx   → 财-2 (文件夹名=类别→缩小候选→关键词)")

    write_file(OUTPUT_DIR / "存货盘点" / "盘点表.xlsx",
        "存货盘点表\n盘点日期：2025年12月31日\n盘点单位：ABC制造有限公司\n"
        "序号\t存货名称\t账存数量\t实存数量\t差异\n"
        "1\t原材料-钢材\t120\t118\t-2\n"
        "2\t原材料-铝材\t50\t50\t0\n"
        "3\t在产品-半成品A\t300\t298\t-2\n"
        "4\t产成品-成品X\t80\t80\t0", "excel")
    print("  /存货盘点/盘点表.xlsx        → 存-1 (文件夹名=类别→缩小候选→关键词)")

    # ════════════════════════════════════════════
    # 形态3: 扁平但有意义（文件名有意义但无编号前缀）
    # ════════════════════════════════════════════
    print("\n── 形态3: 扁平但有意义 ──")
    write_file(OUTPUT_DIR / "盘点计划.pdf",
        "存货盘点计划\n编制单位：ABC制造有限公司\n"
        "盘点基准日：2025年12月31日\n"
        "一、盘点范围：原材料仓库、在产品仓库、产成品仓库\n"
        "二、盘点时间：2025年12月31日 08:00-18:00\n"
        "三、盘点人员：总负责 李四，原材料组 王五，在产品组 赵六\n"
        "四、盘点方法：全面盘点，双签确认\n"
        "五、审计监盘：安永审计团队现场监盘")
    print("  盘点计划.pdf                  → 存-4 (关键词: 盘点计划)")

    write_file(OUTPUT_DIR / "DEF银行对账单.pdf",
        "银行对账单\n户名：DEF实业有限公司\n"
        "开户行：中国银行深圳科技园支行\n"
        "账号：4450 9876 5432 1098\n"
        "对账期间：2025年12月1日 至 2025年12月31日\n"
        "期初余额\t456,000.00\n"
        "期末余额\t619,000.00\n"
        "银行盖章：中国银行深圳科技园支行")
    print("  DEF银行对账单.pdf            → 货-2 (关键词: 银行对账单/DEF)")

    write_file(OUTPUT_DIR / "ABC集团章程_修订版.pdf",
        "ABC集团股份有限公司章程（2024年修订版）\n"
        "第一章 总则\n第一条 公司名称：ABC集团股份有限公司\n"
        "第二条 注册资本：人民币5亿元\n"
        "第二章 股东会\n第六条 股东会由全体股东组成\n"
        "第三章 董事会\n第十二条 董事会由7名董事组成\n"
        "第四章 监事会\n第十八条 监事会由3名监事组成\n"
        "本次修订经2024年3月15日股东会决议通过")
    print("  ABC集团章程_修订版.pdf        → 历-2 (关键词: 公司章程/章程)")

    # ════════════════════════════════════════════
    # 形态4: 按年度分文件夹（一个清单项拆成多个文件 → S3）
    # ════════════════════════════════════════════
    print("\n── 形态4: 按年度分文件夹（S3: 货-1拆3份）──")
    for year, amt in [("2023", "3,850,000"), ("2024", "4,200,000"), ("2025", "4,580,000")]:
        write_file(OUTPUT_DIR / f"{year}年度" / "银行流水.xlsx",
            f"银行流水明细\n账户：ABC制造有限公司\n"
            f"开户行：招商银行深圳南山支行\n"
            f"年度：{year}年度\n\n"
            f"日期\t摘要\t借方\t贷方\t余额\n"
            f"{year}-01-05\t客户货款\t200,000\t\t856,000\n"
            f"{year}-01-10\t供应商款\t\t150,000\t706,000\n"
            f"{year}-01-15\t工资\t\t320,000\t386,000\n"
            f"...\n年度合计：借方 {amt}", "excel")
        print(f"  /{year}年度/银行流水.xlsx         → 货-1 (文件夹名=年份→期间预匹配)")

    # ════════════════════════════════════════════
    # 形态5: 混乱命名（完全无意义文件名，只能靠 LLM 读内容）
    # ════════════════════════════════════════════
    print("\n── 形态5: 混乱命名 ──")
    write_file(OUTPUT_DIR / "新建文档(1).pdf",
        "ABC集团股份有限公司\n2024年度审计报告\n\n"
        "一、审计意见\n我们审计了ABC集团2024年12月31日的合并资产负债表及2024年度合并利润表。\n"
        "认为财务报表公允反映了ABC集团的财务状况和经营成果。\n\n"
        "二、关键审计事项\n1. 收入确认：2024年度合并营业收入3,500万元\n"
        "2. 存货减值：2024年末存货账面价值70万元\n\n"
        "事务所：XX会计师事务所  2025年3月20日")
    print("  新建文档(1).pdf              → 综-1 (LLM兜底, 文件名无线索)")

    write_file(OUTPUT_DIR / "扫描件001.pdf",
        "员工花名册\nABC制造有限公司\n截至2025年12月31日\n\n"
        "序号\t姓名\t部门\t职位\t入职日期\n"
        "1\t张三\t财务部\t财务经理\t2018-03-01\n"
        "2\t李四\t财务部\t财务总监\t2015-06-15\n"
        "3\t王五\t仓储部\t仓库主管\t2019-01-10\n"
        "4\t赵六\t仓储部\t仓管员\t2020-07-01\n"
        "5\t钱七\t销售部\t销售经理\t2017-09-15\n"
        "6\t孙八\t生产部\t生产主管\t2016-02-20")
    print("  扫描件001.pdf                → S5: 清单外文件 (LLM兜底→低confidence→未分类)")

    # ════════════════════════════════════════════
    # S1: 穿行测试文件夹整包提交
    # ════════════════════════════════════════════
    print("\n── S1: 穿行测试整包 ──")
    write_file(OUTPUT_DIR / "穿行测试_销售收款控制" / "系统截图_订单审批.pdf",
        "ERP系统截图 - 销售订单审批流程\n"
        "截图1：销售订单创建（业务员张三，2025-06-15）\n"
        "  订单号：SO-2025-003  客户：上海AA实业  金额：750,000\n"
        "截图2：销售经理审批（王经理，2025-06-15）\n"
        "  审批意见：同意\n"
        "截图3：财务审核（李会计，2025-06-15）\n"
        "  信用额度检查：通过  审核结果：同意发货")
    print("  /穿行测试_销售收款控制/系统截图_订单审批.pdf → 穿-1 整目录归档")

    write_file(OUTPUT_DIR / "穿行测试_销售收款控制" / "合同签字件.pdf",
        "销售合同纸质签字扫描件\n"
        "合同编号：SC-2025-003\n甲方：ABC制造有限公司  乙方：上海AA实业有限公司\n"
        "合同金额：人民币750,000元\n签订日期：2025年6月22日\n\n"
        "甲方签字：王经理  甲方盖章：ABC制造有限公司合同专用章\n"
        "乙方签字：陈总  乙方盖章：上海AA实业有限公司合同专用章")
    print("  /穿行测试_销售收款控制/合同签字件.pdf       → 穿-1 整目录归档")

    write_file(OUTPUT_DIR / "穿行测试_销售收款控制" / "银行回单.pdf",
        "银行收款回单\n收款人：ABC制造有限公司\n"
        "付款人：上海AA实业有限公司\n付款金额：750,000.00元\n"
        "到账日期：2025年8月15日\n摘要：货款 SC-2025-003\n"
        "银行盖章：招商银行深圳南山支行")
    print("  /穿行测试_销售收款控制/银行回单.pdf         → 穿-1 整目录归档")

    # ── 生成对应 PBC 清单 ──
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    pbc_path = OUTPUT_DIR.parent / "01_PBC_List_混合形态.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "PBC清单"
    headers = [
        ("一级分类", True), ("二级分类", True), ("相关科目", False),
        ("资料名称", True), ("问题/需求描述", True), ("报告期间", True),
        ("格式", False), ("优先级", False),
        ("提出时间", False), ("期望提供日期", True),
        ("逾期天数", False), ("资料提供情况", False), ("备注", False),
        ("实体归属", True), ("置信度", False), ("文件路径", False),
    ]
    required_fill = PatternFill(start_color="FCEBEB", end_color="FCEBEB", fill_type="solid")
    optional_fill = PatternFill(start_color="F1EFE8", end_color="F1EFE8", fill_type="solid")
    for i, (h, required) in enumerate(headers, 1):
        cell = ws.cell(1, i, ("* " if required else "") + h)
        cell.font = Font(bold=True, color="A32D2D" if required else "5F5E5A")
        cell.fill = required_fill if required else optional_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    widths = [14, 12, 16, 20, 40, 30, 8, 8, 14, 14, 10, 16, 24, 14, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    dv = DataValidation(type="list", formula1='"未提供,已提供，审核中,已提供,不适用"', allow_blank=True)
    dv.add("L2:L1000")
    ws.add_data_validation(dv)
    ws.freeze_panes = "C2"

    # 所有项的额外字段（subject/file_format/priority/raised_at/expected_by/overdue_days）
    # 今天=2026-07-24，期望提供日期设在 6/15~7/20，未提供的项算逾期天数
    # 有文件的项 overdue_days=0（已提供），未提供的项 overdue_days=实际天数
    extra_fields = {
        # 有文件（已提供）——overdue_days=0
        "历-1": {"subject": "公司治理", "file_format": "PDF", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-15", "overdue_days": 0},
        "历-2": {"subject": "公司治理", "file_format": "PDF", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-15", "overdue_days": 0},
        "财-1": {"subject": "合并报表", "file_format": "Excel", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-10", "overdue_days": 0},
        "财-2": {"subject": "个别报表", "file_format": "Excel", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-10", "overdue_days": 0},
        "存-1": {"subject": "存货盘点", "file_format": "Excel", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-05", "overdue_days": 0},
        "存-4": {"subject": "存货盘点", "file_format": "PDF", "priority": "中", "raised_at": "2026-06-01", "expected_by": "2026-07-05", "overdue_days": 0},
        "货-1": {"subject": "银行存款", "file_format": "Excel", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-15", "overdue_days": 0},
        "货-2": {"subject": "银行存款", "file_format": "PDF", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-15", "overdue_days": 0},
        "销-1": {"subject": "主营业务收入", "file_format": "Excel", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-20", "overdue_days": 0},
        "综-1": {"subject": "年度审计", "file_format": "PDF", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-10", "overdue_days": 0},
        "穿-1": {"subject": "内部控制", "file_format": "PDF", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-20", "overdue_days": 0},
        # 未提供（超期）——overdue_days 按 2026-07-24 算
        "历-3": {"subject": "公司治理", "file_format": "PDF", "priority": "中", "raised_at": "2026-06-01", "expected_by": "2026-07-10", "overdue_days": 14},   # 超14天
        "历-4": {"subject": "公司治理", "file_format": "PDF", "priority": "中", "raised_at": "2026-06-01", "expected_by": "2026-07-10", "overdue_days": 14},   # 超14天
        "财-3": {"subject": "合并报表", "file_format": "Excel", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-06-20", "overdue_days": 34},   # 超34天→high
        "存-5": {"subject": "存货减值", "file_format": "Excel", "priority": "中", "raised_at": "2026-06-01", "expected_by": "2026-07-05", "overdue_days": 19},   # 超19天→medium
        "货-3": {"subject": "银行存款", "file_format": "Excel", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-06-15", "overdue_days": 39},   # 超39天→high
        "销-2": {"subject": "主营业务收入", "file_format": "Excel", "priority": "高", "raised_at": "2026-06-01", "expected_by": "2026-07-05", "overdue_days": 19},   # 超19天→medium
        "综-2": {"subject": "年度审计", "file_format": "PDF", "priority": "中", "raised_at": "2026-06-01", "expected_by": "2026-06-25", "overdue_days": 29},   # 超29天→medium
        "穿-2": {"subject": "内部控制", "file_format": "PDF", "priority": "中", "raised_at": "2026-06-01", "expected_by": "2026-07-15", "overdue_days": 9},   # 超9天→low
    }
    # 列顺序：一级分类, 二级分类, 相关科目, 资料名称, 问题/需求描述, 报告期间,
    #         格式, 优先级, 提出时间, 期望提供日期, 逾期天数, 资料提供情况,
    #         备注, 实体归属, 置信度, 文件路径
    keys = ["category", "item_id", "subject", "doc_name", "description", "required_period",
            "file_format", "priority", "raised_at", "expected_by", "overdue_days", "status_raw",
            "remark", "entity", "confidence", "file_path"]
    for ri, item in enumerate(PBC_ITEMS, 2):
        extra = extra_fields.get(item["item_id"], {})
        row = {**item, **extra, "status_raw": "未提供",
               "remark": "" if item["item_id"] in HAS_FILE else "客户未提供（超期）",
               "confidence": "", "file_path": ""}
        for ci, k in enumerate(keys, 1):
            ws.cell(ri, ci, row.get(k, ""))

    wb.save(str(pbc_path))
    wb.close()
    print(f"\n── PBC清单: {pbc_path}（{len(PBC_ITEMS)}项，{len(HAS_FILE)}项有文件覆盖，{len(PBC_ITEMS)-len(HAS_FILE)}项未提供）──")

    # ── README ──
    readme = OUTPUT_DIR.parent / "README_混合形态测试.txt"
    readme.write_text(
        "PBC 智能管理工作站 · 混合形态测试数据包\n"
        "========================================\n\n"
        "本数据包在一个客户文件夹内混合了5种交付形态 + 3种特殊情况，\n"
        "用于验证 Layer 0（结构感知）+ 四层漏斗分类架构。\n\n"
        "使用步骤：\n"
        "1. 启动 PBC 工作站\n"
        "2. 创建新项目\n"
        "3. 导入 01_PBC_List_混合形态.xlsx（20项，含9项未提供模拟缺料）\n"
        "4. 客户文件夹指向 本目录/客户共享文件夹_混合形态/\n"
        "5. 扫描新文件\n\n"
        "目录结构：\n"
        "  客户共享文件夹_混合形态/\n"
        "    ├── 历-1_股权架构图.pdf          [形态1] filename-match\n"
        "    ├── 销-1_销售合同台账.xlsx        [形态1] filename-match\n"
        "    ├── 财务报表/                     [形态2] 文件夹名=类别\n"
        "    │   ├── 合并资产负债表.xlsx\n"
        "    │   └── 子公司利润表.xlsx\n"
        "    ├── 存货盘点/\n"
        "    │   └── 盘点表.xlsx\n"
        "    ├── 盘点计划.pdf                  [形态3] 扁平有意义\n"
        "    ├── DEF银行对账单.pdf\n"
        "    ├── ABC集团章程_修订版.pdf\n"
        "    ├── 2023年度/                    [形态4] 文件夹名=年份\n"
        "    │   └── 银行流水.xlsx            [S3] 货-1拆3份\n"
        "    ├── 2024年度/\n"
        "    │   └── 银行流水.xlsx\n"
        "    ├── 2025年度/\n"
        "    │   └── 银行流水.xlsx\n"
        "    ├── 新建文档(1).pdf               [形态5] LLM兜底→综-1\n"
        "    ├── 扫描件001.pdf                 [S5] 清单外→未分类\n"
        "    └── 穿行测试_销售收款控制/        [S1] 整目录归档\n"
        "        ├── 系统截图_订单审批.pdf\n"
        "        ├── 合同签字件.pdf\n"
        "        └── 银行回单.pdf\n\n"
        "验证目标：\n"
        "  Layer 0: 感知到5种形态混合存在\n"
        "  Layer 1: 形态1的2个文件秒级命中\n"
        "  Layer 2: 形态2/3/4的文件百毫秒命中（用文件夹名/关键词线索）\n"
        "  Layer 3: 形态5的2个文件走LLM（1个命中综-1, 1个未分类S5）\n"
        "  S1: 穿行测试文件夹整目录归档\n"
        "  S3: 货-1的3个年度文件全部匹配同一item_id\n"
        "  S5: 扫描件001.pdf 低confidence→归档未分类+toast\n",
        encoding="utf-8",
    )
    print(f"\n── 说明文档: {readme} ──")

    print(f"\n{'='*60}")
    print(f"共 16 个文件 + 1 个穿行测试文件夹(3文件) = 19 文件")
    print(f"PBC清单 20 项（11项有文件覆盖，9项未提供模拟缺料）")
    print(f"覆盖 5 种交付形态 + 3 种特殊情况")


if __name__ == "__main__":
    main()
