"""把 v7 的 PBC 清单（15 列）迁移到 v7.5 的 16 列结构。

v7 → v7.5 字段映射：
  v7 第1列 资料编号(历-1)       → v7.5 第2列 二级分类(item_id)
  v7 第2列 一级分类(历史沿革)    → v7.5 第1列 一级分类
  v7 第3列 相关科目             → v7.5 第3列 相关科目
  v7 第4列 问题/需求描述(股权架构图) → v7.5 第4列 资料名称 + 第5列 问题/需求描述（拆分）
  v7 第15列 需求期间(2024年度)   → v7.5 第6列 报告期间
  v7 第5-14列                   → v7.5 第7-16列（顺序基本一致）

用法：
  python scripts/migrate_pbc_v7_to_v75.py <input.xlsx> [output.xlsx]
"""
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule


# v7.5 16 列结构
V75_HEADERS = [
    "* 一级分类", "* 二级分类", "相关科目", "* 资料名称",
    "* 问题/需求描述", "* 报告期间", "格式", "优先级",
    "提出时间", "* 期望提供日期", "逾期天数", "资料提供情况",
    "备注", "* 实体归属", "置信度", "文件路径",
]

REQUIRED_COLS = {1, 2, 4, 5, 6, 10, 14}  # 带 * 的列索引


def migrate(input_path: Path, output_path: Path = None):
    """迁移 v7 PBC 清单到 v7.5 16 列结构。"""
    if output_path is None:
        output_path = input_path  # 原地覆盖

    wb_src = load_workbook(str(input_path), data_only=True)
    ws_src = wb_src.active

    # 读 v7 数据
    src_rows = []
    for r in range(2, ws_src.max_row + 1):
        # 跳过完全空行
        if not any(ws_src.cell(r, c).value for c in range(1, 16)):
            continue
        row = {
            "item_id": ws_src.cell(r, 1).value,        # 资料编号 → 二级分类
            "category": ws_src.cell(r, 2).value,        # 一级分类
            "subject": ws_src.cell(r, 3).value,          # 相关科目
            "description": ws_src.cell(r, 4).value,     # 问题/需求描述（v7.5 拆成 doc_name + description）
            "file_format": ws_src.cell(r, 5).value,     # 格式
            "priority": ws_src.cell(r, 6).value,        # 优先级
            "raised_at": ws_src.cell(r, 7).value,      # 提出时间
            "expected_by": ws_src.cell(r, 8).value,     # 期望提供日期
            "overdue_days": ws_src.cell(r, 9).value or 0,
            "status_raw": ws_src.cell(r, 10).value or "未提供",
            "remark": ws_src.cell(r, 11).value,
            "entity": ws_src.cell(r, 12).value,
            "confidence": ws_src.cell(r, 13).value,
            "file_path": ws_src.cell(r, 14).value,
            "required_period": ws_src.cell(r, 15).value,  # 需求期间 → 报告期间
        }
        src_rows.append(row)

    # 写 v7.5 16 列结构
    wb_dst = Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "PBC清单"

    # 表头
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E2E38")
    required_fill = PatternFill("solid", fgColor="C0392B")  # 必填红底
    thin_border = Border(
        left=Side(style="thin", color="C4C4CD"),
        right=Side(style="thin", color="C4C4CD"),
        top=Side(style="thin", color="C4C4CD"),
        bottom=Side(style="thin", color="C4C4CD"),
    )
    for c, header in enumerate(V75_HEADERS, start=1):
        cell = ws_dst.cell(1, c, header)
        cell.font = bold_font
        cell.fill = required_fill if c in REQUIRED_COLS else header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 数据行
    for r_idx, row in enumerate(src_rows, start=2):
        # v7.5 第4列"资料名称" 从 v7 第4列"问题/需求描述"取（取简短名）
        # v7.5 第5列"问题/需求描述" 写详细说明（v7 没有详细说明，复制过来 + 加"请提供"前缀）
        doc_name = row.get("description") or ""
        detailed_desc = f"请提供{doc_name}相关资料" if doc_name else ""

        values = [
            row.get("category"),       # 1 一级分类
            row.get("item_id"),        # 2 二级分类（原资料编号）
            row.get("subject"),        # 3 相关科目
            doc_name,                  # 4 资料名称
            detailed_desc,             # 5 问题/需求描述
            row.get("required_period"),  # 6 报告期间
            row.get("file_format"),    # 7 格式
            row.get("priority"),       # 8 优先级
            row.get("raised_at"),      # 9 提出时间
            row.get("expected_by"),    # 10 期望提供日期
            row.get("overdue_days"),   # 11 逾期天数
            row.get("status_raw"),     # 12 资料提供情况
            row.get("remark"),         # 13 备注
            row.get("entity"),         # 14 实体归属
            row.get("confidence"),     # 15 置信度
            row.get("file_path"),      # 16 文件路径
        ]
        for c, v in enumerate(values, start=1):
            cell = ws_dst.cell(r_idx, c, v)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 列宽
    col_widths = [14, 10, 12, 18, 30, 12, 10, 8, 12, 14, 10, 14, 16, 14, 8, 30]
    for i, w in enumerate(col_widths, start=1):
        ws_dst.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w

    # 冻结首行
    ws_dst.freeze_panes = "A2"

    # 状态机数据验证（第 12 列 资料提供情况）
    dv_status = DataValidation(
        type="list",
        formula1='"未提供,已提供,审核中,不适用,已提供，审核中"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="无效状态",
        error="状态必须是：未提供/已提供/审核中/不适用/已提供，审核中",
    )
    dv_status.add(f"L2:L{ws_dst.max_row}")
    ws_dst.add_data_validation(dv_status)

    # 逾期天数条件格式（红色高亮 > 0）
    from openpyxl.formatting.rule import CellIsRule
    red_fill = PatternFill("solid", fgColor="FFCDD2")
    red_font = Font(color="C62828", bold=True)
    ws_dst.conditional_formatting.add(
        f"K2:K{ws_dst.max_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill, font=red_font),
    )

    wb_dst.save(str(output_path))
    print(f"迁移完成: {input_path} → {output_path}")
    print(f"  数据行: {len(src_rows)}")
    print(f"  列数: 16（v7.5 结构）")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python migrate_pbc_v7_to_v75.py <input.xlsx> [output.xlsx]")
        sys.exit(1)
    inp = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    if not inp.exists():
        print(f"输入文件不存在: {inp}")
        sys.exit(1)
    migrate(inp, out)
