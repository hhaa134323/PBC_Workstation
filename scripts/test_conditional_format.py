"""P0-3 验证脚本：write_pbc_list 增量更新后 Excel 条件格式规则数 ≥ 4。

思路：
1. 复制 mock_data/01_PBC_List.xlsx 到临时路径
2. 调用 write_pbc_list 更新某条 item_id 的 status_raw
3. 重新打开文件，检查 conditional_formatting 规则数 ≥ 4（红/绿/浅绿/灰 4 条）
4. 同时验证：写入前后规则数一致（不能丢失）

运行：
    python scripts/test_conditional_format.py
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

# 把项目根加到 sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.core.excel_io import write_pbc_list  # noqa: E402


def _count_cf_rules(xlsx_path: Path) -> tuple[int, list[str]]:
    """返回 (规则总数, 每条规则的 type 列表)。"""
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    n = 0
    rule_types: list[str] = []
    for cf in ws.conditional_formatting:
        for rule in cf.rules:
            n += 1
            rule_types.append(rule.type or "unknown")
    wb.close()
    return n, rule_types


def main() -> int:
    src = PROJECT_ROOT / "mock_data" / "01_PBC_List.xlsx"
    if not src.exists():
        print(f"FAIL: 源文件不存在: {src}")
        return 1

    with tempfile.TemporaryDirectory() as td:
        dst = Path(td) / "test_pbc_list.xlsx"
        shutil.copy2(str(src), str(dst))

        before, before_types = _count_cf_rules(dst)
        print(f"写入前规则数: {before}")
        print(f"写入前规则类型: {before_types}")
        assert before > 0, "源文件应至少有 1 条条件格式规则，实际为 0"

        # 找一个真实存在的 item_id 改它的 status_raw
        wb = load_workbook(str(dst))
        ws = wb.active
        sample_item_id = ws.cell(2, 1).value
        wb.close()
        assert sample_item_id, "找不到示例 item_id"
        print(f"测试 item_id: {sample_item_id}")

        # 调用 write_pbc_list 增量更新
        write_pbc_list(
            rows=[{"item_id": sample_item_id, "status_raw": "已提供，审核中"}],
            xlsx_path=str(dst),
        )

        after, after_types = _count_cf_rules(dst)
        print(f"写入后规则数: {after}")
        print(f"写入后规则类型: {after_types}")

        # P0-3 要求：≥ 4 条（红/绿/浅绿/灰）
        if after < 4:
            print(f"FAIL: 写入后条件格式规则数 {after} < 4，期望 ≥ 4")
            return 2

        if after != before:
            print(f"FAIL: 条件格式规则数变化: 写入前 {before} → 写入后 {after}")
            return 3

        print(f"PASS: write_pbc_list 保留条件格式 (before={before}, after={after}, ≥4)")
        return 0


if __name__ == "__main__":
    sys.exit(main())
