"""P0-3 单测：验证 write_pbc_list 增量更新后 Excel 条件格式规则数不变。

思路：
1. 复制 mock_data/01_PBC_List.xlsx 到临时路径
2. 调用 write_pbc_list 更新某条 item_id 的 status_raw
3. 重新打开文件，检查 conditional_formatting 规则数与原文件一致
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

# 把项目根加到 sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from app.core.excel_io import write_pbc_list  # noqa: E402


def _count_cf_rules(xlsx_path: Path) -> int:
    wb = load_workbook(str(xlsx_path))
    ws = wb.active
    n = 0
    for cf in ws.conditional_formatting:
        n += len(cf.rules)
    wb.close()
    return n


def test_write_pbc_list_preserves_conditional_formatting() -> None:
    src = PROJECT_ROOT / "mock_data" / "01_PBC_List.xlsx"
    assert src.exists(), f"源文件不存在: {src}"

    with tempfile.TemporaryDirectory() as td:
        dst = Path(td) / "test_pbc_list.xlsx"
        shutil.copy2(str(src), str(dst))

        # 写入前的规则数
        before = _count_cf_rules(dst)
        assert before > 0, "源文件应至少有 1 条条件格式规则，实际为 0"

        # 找一个真实存在的 item_id 改它的 status_raw
        wb = load_workbook(str(dst))
        ws = wb.active
        sample_item_id = ws.cell(2, 1).value
        wb.close()
        assert sample_item_id, "找不到示例 item_id"

        # 调用 write_pbc_list 增量更新
        write_pbc_list(
            rows=[{"item_id": sample_item_id, "status_raw": "已提供，审核中"}],
            xlsx_path=str(dst),
        )

        # 写入后的规则数
        after = _count_cf_rules(dst)
        assert after == before, (
            f"条件格式规则数变化: 写入前 {before} 条 → 写入后 {after} 条"
        )
        print(f"PASS: write_pbc_list 保留条件格式 (before={before}, after={after})")


if __name__ == "__main__":
    test_write_pbc_list_preserves_conditional_formatting()
