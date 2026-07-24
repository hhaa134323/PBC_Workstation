"""PBC 清单 CRUD 路由（M2 + 多项目支持）。

多项目支持：
- 新路由带 project_id：GET /api/pbc/{project_id}/list 等
- 旧路由（不带 project_id）走 demo 项目（兼容）

路由声明顺序很重要（FastAPI 按声明顺序匹配，更具体的模式先声明）：
  1. /list                              (1 seg, literal "list")
  2. /{project_id}/list                 (2 seg, literal "list")
  3. /{item_id}/status                  (2 seg, literal "status", legacy)
  4. /{project_id}/{item_id}/status     (3 seg, literal "status")
  5. /{project_id}/{item_id}            (2 seg, general)
  6. /{item_id}                         (1 seg, general, legacy)
"""
from __future__ import annotations

import logging
import os
from typing import Any, Optional

from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field

from app.config import get_config
from app.core.db import get_project
from app.core.excel_io import (
    get_item_by_id,
    read_pbc_list,
    update_item_status,
)

logger = logging.getLogger("pbc.routes_pbc")

router = APIRouter(prefix="/api/pbc", tags=["pbc"])

_DEFAULT_PROJECT = "demo"


def _xlsx_path(project_id: Optional[str] = None) -> str:
    """兼容旧调用：不传 project_id 时用全局 config.pbc_list_path。"""
    if project_id:
        proj = get_project(project_id)
        if proj is None:
            raise FileNotFoundError(f"项目不存在: {project_id}")
        return proj.get("pbc_list_path") or ""
    return str(get_config().pbc_list_path)


def _filter_items(
    items: list[dict[str, Any]],
    status: Optional[str],
    entity: Optional[str],
    category: Optional[str],
    risk: Optional[str],
    overdue_min: Optional[int],
) -> list[dict[str, Any]]:
    """按 query 参数过滤。"""
    out = items
    if status:
        out = [it for it in out if it.get("status_normalized") == status]
    if entity:
        out = [it for it in out if it.get("entity") == entity]
    if category:
        out = [it for it in out if it.get("category") == category]
    if risk:
        out = [it for it in out if it.get("risk_level") == risk]
    if overdue_min is not None:
        out = [
            it for it in out
            if (it.get("overdue_days") or 0) >= overdue_min
        ]
    return out


class StatusUpdateBody(BaseModel):
    status: str = Field(..., description="新状态（如 已提供 / 已提供，审核中 / 未提供 / 不适用 / 待定）")
    changed_by: str = Field(..., description="操作人（如 Senior / Staff）")
    note: str = Field("", description="备注")


# ----------------------------------------------------------------------
# PBC 清单导入（Demo 录制用：从 0 到 1）
# ----------------------------------------------------------------------
@router.post("/{project_id}/import")
async def import_pbc_list(project_id: str, file: UploadFile = File(...)):
    """导入 PBC 清单 Excel 文件，替换该项目的 01_PBC_List.xlsx。

    v7.2: 加必填字段校验。必填：一级分类 / 二级分类 / 资料名称 /
    问题/需求描述 / 报告期间 / 期望提供日期 / 实体归属。缺字段返回友好错误（不覆盖现有清单）。
    """
    project = get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail=f"项目不存在: {project_id}")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    import io
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析 Excel: {e}")

    ws = wb.active
    if ws.max_row < 2:
        raise HTTPException(status_code=400, detail="Excel 至少需要 1 行表头 + 1 行数据")

    # v7: 必填字段校验
    header_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            h = str(v).strip().lstrip("* ").strip()
            header_map[h] = c

    required_fields = ["一级分类", "二级分类", "资料名称", "问题/需求描述", "报告期间", "期望提供日期", "实体归属"]
    missing_cols = [f for f in required_fields if f not in header_map]
    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail=f"Excel 缺少必填列: {', '.join(missing_cols)}。请下载最新模板。",
        )

    errors: list[str] = []
    valid_rows = 0
    for r in range(2, ws.max_row + 1):
        if all(ws.cell(r, c).value in (None, "") for c in range(1, ws.max_column + 1)):
            continue
        valid_rows += 1
        for field in required_fields:
            col = header_map[field]
            v = ws.cell(r, col).value
            if v is None or str(v).strip() == "":
                item_id = ws.cell(r, header_map["二级分类"]).value or f"第{r}行"
                errors.append(f"{item_id} 缺必填字段「{field}」")
    if errors:
        raise HTTPException(
            status_code=400,
            detail=f"必填校验失败（共 {len(errors)} 处）: {'; '.join(errors[:10])}",
        )
    if valid_rows == 0:
        raise HTTPException(status_code=400, detail="Excel 没有有效数据行")

    from pathlib import Path
    pbc_path = Path(project["pbc_list_path"])
    pbc_path.parent.mkdir(parents=True, exist_ok=True)

    # 安全写入：用非 data_only 模式重新打开，直接写
    # （shutil.move 在 Windows 沙箱环境下有 safe-delete 限制，不用临时文件）
    wb_save = openpyxl.load_workbook(io.BytesIO(content))
    wb_save.save(str(pbc_path))

    return {
        "ok": True,
        "project_id": project_id,
        "imported_rows": valid_rows,
        "sheet_name": ws.title,
        "message": f"已导入 {valid_rows} 项 PBC 清单（必填校验通过）",
    }


# ----------------------------------------------------------------------
# v7: 下载 PBC 导入模板
# ----------------------------------------------------------------------
@router.get("/{project_id}/download-template")
async def download_pbc_template(project_id: str):
    """下载 PBC 导入模板（v7: 15 列表头 + 必填标注 + 数据验证）。

    生成带必填红色星标 + 数据验证 + 单元格注释的标准模板，
    审计员填好后再上传 /import 接口。
    """
    from fastapi.responses import StreamingResponse
    import io
    import tempfile
    from pathlib import Path
    from app.core.db import _create_empty_pbc_xlsx

    project = get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail=f"项目不存在: {project_id}")

    # 生成临时模板文件
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
    _create_empty_pbc_xlsx(tmp_path)

    def _file_iter():
        with open(tmp_path, "rb") as f:
            while True:
                chunk = f.read(65536)
                if not chunk:
                    break
                yield chunk
        try:
            tmp_path.unlink()
        except Exception:
            pass

    return StreamingResponse(
        _file_iter(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''PBC_Template_{project_id}.xlsx"
        },
    )


@router.get("/{project_id}/export")
async def export_pbc_list(project_id: str):
    """导出当前项目的 PBC 清单（含已回写的状态/文件路径/置信度）。

    直接读取项目的 01_PBC_List.xlsx 返回文件流，
    审计员可下载查看最新状态。
    """
    from fastapi.responses import FileResponse
    from pathlib import Path

    project = get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail=f"项目不存在: {project_id}")

    pbc_path = project.get("pbc_list_path") or ""
    if not pbc_path:
        raise HTTPException(status_code=404, detail="该项目未配置 PBC 清单路径")

    p = Path(pbc_path)
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"PBC 清单文件不存在: {pbc_path}")

    # 用项目名+日期做文件名
    proj_name = project.get("name") or project_id
    from datetime import datetime
    date_str = datetime.now().strftime("%Y%m%d")
    download_name = f"PBC_List_{proj_name}_{date_str}.xlsx"

    return FileResponse(
        path=str(p),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_name,
    )
# 路由顺序：字面值路由优先于通配路由
# ----------------------------------------------------------------------
@router.get("/list")
async def list_items_legacy(
    status: Optional[str] = Query(None, description="按 status_normalized 过滤"),
    entity: Optional[str] = Query(None, description="按 entity 过滤"),
    category: Optional[str] = Query(None, description="按 category 过滤"),
    risk: Optional[str] = Query(None, description="按 risk_level 过滤"),
    overdue_min: Optional[int] = Query(None, description="按 overdue_days >= N 过滤"),
) -> dict:
    """列出 PBC 项（兼容旧调用，走 demo 项目）。"""
    try:
        items = read_pbc_list(project_id=_DEFAULT_PROJECT)
    except FileNotFoundError as e:
        # 兜底：demo 项目不存在时回退到全局 config
        try:
            items = read_pbc_list()
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail=str(e))
        except Exception as e:
            logger.exception("read_pbc_list failed")
            raise HTTPException(status_code=500, detail=f"读取清单失败: {e}")
    except Exception as e:
        logger.exception("read_pbc_list failed")
        raise HTTPException(status_code=500, detail=f"读取清单失败: {e}")

    items = _filter_items(items, status, entity, category, risk, overdue_min)
    return {
        "count": len(items),
        "items": items,
        "filters": {
            "status": status, "entity": entity, "category": category,
            "risk": risk, "overdue_min": overdue_min,
        },
        "deprecated_hint": f"请改用 /api/pbc/{_DEFAULT_PROJECT}/list",
    }


# ----------------------------------------------------------------------
# 多项目路由（推荐使用）
# ----------------------------------------------------------------------
@router.get("/{project_id}/list")
async def list_items_by_project(
    project_id: str,
    status: Optional[str] = Query(None, description="按 status_normalized 过滤"),
    entity: Optional[str] = Query(None, description="按 entity 过滤"),
    category: Optional[str] = Query(None, description="按 category 过滤"),
    risk: Optional[str] = Query(None, description="按 risk_level 过滤"),
    overdue_min: Optional[int] = Query(None, description="按 overdue_days >= N 过滤"),
) -> dict:
    """列出某项目的 PBC 项，支持过滤。"""
    try:
        items = read_pbc_list(project_id=project_id)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception("read_pbc_list failed")
        raise HTTPException(status_code=500, detail=f"读取清单失败: {e}")

    items = _filter_items(items, status, entity, category, risk, overdue_min)
    return {
        "project_id": project_id,
        "count": len(items),
        "items": items,
        "filters": {
            "status": status, "entity": entity, "category": category,
            "risk": risk, "overdue_min": overdue_min,
        },
    }


# ----------------------------------------------------------------------
# 兼容旧路由：PUT /{item_id}/status （走 demo 项目）
# 必须在 /{project_id}/{item_id} 之前声明（字面值 "status" 更具体）
# ----------------------------------------------------------------------
@router.put("/{item_id}/status")
async def update_status_legacy(item_id: str, body: StatusUpdateBody) -> dict:
    """更新 PBC 项状态（兼容旧调用，走 demo 项目）。"""
    try:
        ok, msg, updated = update_item_status(
            item_id=item_id,
            new_status=body.status,
            changed_by=body.changed_by,
            note=body.note,
            project_id=_DEFAULT_PROJECT,
        )
    except FileNotFoundError as e:
        # 兜底：用全局 config
        try:
            ok, msg, updated = update_item_status(
                item_id=item_id,
                new_status=body.status,
                changed_by=body.changed_by,
                note=body.note,
                xlsx_path=str(get_config().pbc_list_path),
            )
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail=str(e))
        except Exception as e:
            logger.exception("update_item_status failed")
            raise HTTPException(status_code=500, detail=f"更新失败: {e}")
    except Exception as e:
        logger.exception("update_item_status failed")
        raise HTTPException(status_code=500, detail=f"更新失败: {e}")

    if not ok:
        raise HTTPException(status_code=400, detail=msg)

    return {
        "ok": True,
        "message": msg,
        "item": updated,
        "deprecated_hint": f"请改用 /api/pbc/{_DEFAULT_PROJECT}/{item_id}/status",
    }


# ----------------------------------------------------------------------
# 多项目路由：PUT /{project_id}/{item_id}/status
# ----------------------------------------------------------------------
@router.put("/{project_id}/{item_id}/status")
async def update_status_by_project(
    project_id: str, item_id: str, body: StatusUpdateBody
) -> dict:
    """更新某项目的 PBC 项状态（含状态机校验 + Excel 写入 + SQLite state_changes 记录）。"""
    try:
        ok, msg, updated = update_item_status(
            item_id=item_id,
            new_status=body.status,
            changed_by=body.changed_by,
            note=body.note,
            project_id=project_id,
        )
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception("update_item_status failed")
        raise HTTPException(status_code=500, detail=f"更新失败: {e}")

    if not ok:
        raise HTTPException(status_code=400, detail=msg)

    return {
        "ok": True,
        "project_id": project_id,
        "message": msg,
        "item": updated,
    }


# ----------------------------------------------------------------------
# 多项目路由：GET /{project_id}/{item_id}
# 必须在 /{item_id} 之前声明（2 seg 比 1 seg 更具体）
# ----------------------------------------------------------------------
@router.get("/{project_id}/{item_id}")
async def get_item_by_project(project_id: str, item_id: str) -> dict:
    """获取某项目的单个 PBC 项详情。

    注意：item_id 为 "list" 时由 /{project_id}/list 处理（已先声明）；
    item_id 为 "status" 时不会进这里（PUT 路由）。
    """
    if item_id == "list":
        # 不应该到这里，/list 路由已先匹配
        raise HTTPException(status_code=404, detail="无效的 item_id: list")
    try:
        item = get_item_by_id(item_id, project_id=project_id)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception("get_item_by_id failed")
        raise HTTPException(status_code=500, detail=f"查询失败: {e}")

    if item is None:
        raise HTTPException(status_code=404, detail=f"未找到资料项: {item_id}")
    return {"project_id": project_id, "item": item}


# ----------------------------------------------------------------------
# 兼容旧路由：GET /{item_id} （1 seg, 走 demo 项目）
# 必须最后声明（最通用的模式）
# ----------------------------------------------------------------------
@router.get("/{item_id}")
async def get_item_legacy(item_id: str) -> dict:
    """获取单个 PBC 项详情（兼容旧调用，走 demo 项目）。"""
    if item_id == "list":
        # 由 /list 路由处理
        raise HTTPException(status_code=404, detail="无效的 item_id: list")
    try:
        item = get_item_by_id(item_id, project_id=_DEFAULT_PROJECT)
    except FileNotFoundError as e:
        # 兜底：用全局 config
        try:
            item = get_item_by_id(item_id, xlsx_path=str(get_config().pbc_list_path))
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail=str(e))
        except Exception as e:
            logger.exception("get_item_by_id failed")
            raise HTTPException(status_code=500, detail=f"查询失败: {e}")
    except Exception as e:
        logger.exception("get_item_by_id failed")
        raise HTTPException(status_code=500, detail=f"查询失败: {e}")

    if item is None:
        raise HTTPException(status_code=404, detail=f"未找到资料项: {item_id}")
    return {"item": item, "deprecated_hint": f"请改用 /api/pbc/{_DEFAULT_PROJECT}/{item_id}"}
