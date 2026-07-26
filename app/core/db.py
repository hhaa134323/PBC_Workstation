"""SQLite 初始化 + 业务表（M1 骨架 + M4 扩展 + 多项目支持）。

M1：state_changes / ai_history / sessions
M4：tasks（任务状态持久化） / file_archive（已归档文件索引）
多项目：projects 表 + 现有表全部加 project_id 字段

用标准库 sqlite3，不依赖 sqlalchemy。

P0-9: 启用 WAL 模式避免 SQLite 锁竞争；所有写入操作包 retry 3 次（50ms 间隔）。
"""
from __future__ import annotations

import json
import logging
import re
import sqlite3
import threading
import time
import uuid
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator, Optional

from app.config import get_config

logger = logging.getLogger("pbc.db")

# 线程局部连接（sqlite 默认禁止跨线程共享）
_tls = threading.local()

# P0-9: 写入 retry 配置
_WRITE_RETRY_TIMES = 3
_WRITE_RETRY_INTERVAL = 0.05  # 50ms


_SCHEMA = """
CREATE TABLE IF NOT EXISTS state_changes (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id      TEXT,
    old_status   TEXT,
    new_status   TEXT,
    changed_by   TEXT,
    changed_at   TEXT DEFAULT (datetime('now','localtime')),
    note         TEXT
);

CREATE TABLE IF NOT EXISTS ai_history (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id      TEXT,
    action       TEXT,
    prompt       TEXT,
    response     TEXT,
    confidence   REAL,
    model        TEXT,
    created_at   TEXT DEFAULT (datetime('now','localtime'))
);

CREATE TABLE IF NOT EXISTS sessions (
    id            TEXT PRIMARY KEY,
    role          TEXT,
    username      TEXT,
    started_at    TEXT DEFAULT (datetime('now','localtime')),
    last_active_at TEXT
);

-- M4: 任务状态持久化（替代 routes_files 内存 dict）
CREATE TABLE IF NOT EXISTS tasks (
    task_id        TEXT PRIMARY KEY,
    status         TEXT NOT NULL,
    progress       INTEGER DEFAULT 0,
    total          INTEGER,
    done_count     INTEGER,
    source         TEXT,
    folder         TEXT,
    received_json  TEXT,
    results_json   TEXT,
    started_at     TEXT NOT NULL,
    finished_at    TEXT,
    error          TEXT
);
CREATE INDEX IF NOT EXISTS idx_tasks_status  ON tasks(status);
CREATE INDEX IF NOT EXISTS idx_tasks_started ON tasks(started_at DESC);

-- M4: 文件归档索引
CREATE TABLE IF NOT EXISTS file_archive (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id       TEXT NOT NULL,
    original_path TEXT NOT NULL,
    archived_path TEXT NOT NULL,
    sha256        TEXT,
    file_size     INTEGER,
    entity        TEXT,
    archived_at   TEXT NOT NULL,
    archived_by   TEXT NOT NULL,
    UNIQUE(archived_path)
);
CREATE INDEX IF NOT EXISTS idx_archive_item ON file_archive(item_id);
CREATE INDEX IF NOT EXISTS idx_archive_sha  ON file_archive(sha256);

-- 多项目: 项目元数据表
CREATE TABLE IF NOT EXISTS projects (
    project_id     TEXT PRIMARY KEY,
    name           TEXT NOT NULL,
    client_name    TEXT,
    folder_path    TEXT NOT NULL,
    pbc_list_path  TEXT NOT NULL,
    client_folder  TEXT,
    archive_root   TEXT,
    created_at     TEXT NOT NULL,
    updated_at     TEXT,
    is_active      INTEGER DEFAULT 1,
    note           TEXT
);
CREATE INDEX IF NOT EXISTS idx_projects_active ON projects(is_active);

-- v7.6: 文件变更日志表（持久化操作记录，审计留痕）
CREATE TABLE IF NOT EXISTS file_change_log (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id    TEXT,
    file_name     TEXT,
    sha256        TEXT,
    change_type   TEXT NOT NULL,   -- added/modified/deleted/archived/reclassified/approved/missing
    item_id       TEXT,
    changed_by    TEXT,            -- watchdog / ai-auto / manual
    changed_at    TEXT NOT NULL,
    detail        TEXT
);
CREATE INDEX IF NOT EXISTS idx_change_log_project ON file_change_log(project_id);
CREATE INDEX IF NOT EXISTS idx_change_log_type ON file_change_log(change_type);
CREATE INDEX IF NOT EXISTS idx_change_log_at ON file_change_log(changed_at DESC);

-- v7.7: 待确认队列（AI预分析结果，等人确认后才归档）
CREATE TABLE IF NOT EXISTS pending_confirm (
    id                INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id        TEXT,
    file_path         TEXT NOT NULL,      -- 客户文件夹里的文件路径
    file_name         TEXT,
    sha256            TEXT,
    suggested_item_id TEXT,               -- AI建议的item_id
    confidence        REAL,
    decision          TEXT,               -- auto/suggest/llm
    conflict_signal   TEXT,              -- JSON序列化
    advisory_notes    TEXT,              -- JSON序列化
    created_at        TEXT NOT NULL,
    confirmed         INTEGER DEFAULT 0   -- 0未确认 1已确认 2已跳过 3已归档
);
CREATE INDEX IF NOT EXISTS idx_pending_confirm_project ON pending_confirm(project_id);
CREATE INDEX IF NOT EXISTS idx_pending_confirm_status ON pending_confirm(confirmed);
"""


# 多项目: ALTER TABLE 给现有表加 project_id 字段（幂等；已存在会报错，try/except 忽略）
# v7: file_archive 加 is_directory 字段（整目录归档用）
_ALTER_PROJECT_ID = [
    ("state_changes", "project_id", "CREATE INDEX IF NOT EXISTS idx_state_changes_project ON state_changes(project_id)"),
    ("ai_history",    "project_id", "CREATE INDEX IF NOT EXISTS idx_ai_history_project ON ai_history(project_id)"),
    ("sessions",      "project_id", None),
    ("tasks",         "project_id", "CREATE INDEX IF NOT EXISTS idx_tasks_project ON tasks(project_id)"),
    ("file_archive",  "project_id", "CREATE INDEX IF NOT EXISTS idx_archive_project ON file_archive(project_id)"),
]

# v7: file_archive 额外加 is_directory + version 字段（独立 ALTER，幂等）
_ALTER_FILE_ARCHIVE_EXTRAS = [
    ("is_directory", "INTEGER DEFAULT 0", None),
    ("version",      "TEXT",              None),
    ("doc_type",     "TEXT",              None),  # basic/walkthrough/unknown
    ("category",     "TEXT",              "CREATE INDEX IF NOT EXISTS idx_archive_category ON file_archive(category)"),
]


def init_db(db_path: Path | None = None) -> None:
    """初始化 SQLite：父目录不存在则创建，再建表 + ALTER 加 project_id。

    P0-9: 同时启用 WAL 模式（journal_mode=WAL）避免锁竞争。
    """
    cfg = get_config()
    path = db_path or cfg.db_path
    path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(path)) as conn:
        # P0-9: 启用 WAL 模式（更小的写锁竞争，读不阻塞写）
        try:
            conn.execute("PRAGMA journal_mode=WAL;")
            # 提升并发友好度
            conn.execute("PRAGMA synchronous=NORMAL;")
            conn.execute("PRAGMA busy_timeout=3000;")  # 等待锁最多 3 秒
        except Exception as e:
            logger.warning("PRAGMA WAL 设置失败: %r", e)
        conn.executescript(_SCHEMA)
        # 幂等地给现有表加 project_id 字段
        for table, col, idx_sql in _ALTER_PROJECT_ID:
            try:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} TEXT")
            except sqlite3.OperationalError:
                # 字段已存在，忽略
                pass
            if idx_sql:
                try:
                    conn.execute(idx_sql)
                except sqlite3.OperationalError:
                    pass
        # v7: file_archive 额外字段
        for col, col_type, idx_sql in _ALTER_FILE_ARCHIVE_EXTRAS:
            try:
                conn.execute(f"ALTER TABLE file_archive ADD COLUMN {col} {col_type}")
            except sqlite3.OperationalError:
                pass
            if idx_sql:
                try:
                    conn.execute(idx_sql)
                except sqlite3.OperationalError:
                    pass
        conn.commit()


def _get_conn() -> sqlite3.Connection:
    if not hasattr(_tls, "conn") or _tls.conn is None:
        cfg = get_config()
        _tls.conn = sqlite3.connect(str(cfg.db_path), check_same_thread=False)
        _tls.conn.row_factory = sqlite3.Row
        # P0-9: 每个连接也设 WAL + busy_timeout
        try:
            _tls.conn.execute("PRAGMA journal_mode=WAL;")
            _tls.conn.execute("PRAGMA synchronous=NORMAL;")
            _tls.conn.execute("PRAGMA busy_timeout=3000;")
        except Exception as e:
            logger.warning("连接级 PRAGMA 设置失败: %r", e)
    return _tls.conn


@contextmanager
def get_conn() -> Iterator[sqlite3.Connection]:
    """提供连接上下文管理器（同一线程复用）。

    P0-9: 失败时 rollback + raise（调用方需自己包 retry）。
    """
    conn = _get_conn()
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise


def execute_with_retry(
    sql: str,
    params: tuple = (),
    *,
    times: int = _WRITE_RETRY_TIMES,
    interval: float = _WRITE_RETRY_INTERVAL,
) -> sqlite3.Cursor:
    """P0-9: 包 retry 的写入操作（INSERT/UPDATE/DELETE）。

    遇到 SQLITE_BUSY / SQLITE_LOCKED 时等待 interval 后重试，最多 times 次。
    其他异常直接抛出（业务错误不重试）。
    """
    last_exc: Optional[Exception] = None
    for attempt in range(1, times + 1):
        try:
            with get_conn() as conn:
                cur = conn.execute(sql, params)
            return cur
        except sqlite3.OperationalError as e:
            msg = str(e).lower()
            if "database is locked" in msg or "database table is locked" in msg or "busy" in msg:
                last_exc = e
                logger.warning(
                    "SQLite 写入被锁（attempt %d/%d）: %r",
                    attempt, times, e,
                )
                if attempt < times:
                    time.sleep(interval)
                continue
            raise
        except Exception as e:
            raise
    # 重试耗尽
    if last_exc is not None:
        raise last_exc
    raise RuntimeError("execute_with_retry 不明失败")


# ----------------------------------------------------------------------
# 多项目: slug 生成 + projects 表 CRUD
# ----------------------------------------------------------------------

def slugify(name: str) -> str:
    """把项目名转成 slug（如 "华银康医疗 IPO" → "huayinkang-yiliao-ipo"）。

    规则：
      - 中文逐字转拼音（懒加载，无 pypinyin 时退回原字符）
      - 英文小写
      - 非字母数字字符替换为 "-"
      - 折叠多个 "-"
      - 去掉首尾 "-"
    """
    if not name:
        return f"proj-{uuid.uuid4().hex[:8]}"
    s = str(name).strip().lower()

    # 尝试中文转拼音
    try:
        from pypinyin import lazy_pinyin, Style  # type: ignore
        s = " ".join(lazy_pinyin(s, style=Style.NORMAL))
    except Exception:
        # 无 pypinyin 就保留原字符
        pass

    # 非字母数字 → "-"
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-{2,}", "-", s).strip("-")
    return s or f"proj-{uuid.uuid4().hex[:8]}"


def _ensure_unique_slug(base_slug: str) -> str:
    """确保 slug 唯一：若已存在则加 -2 / -3 后缀。"""
    with get_conn() as conn:
        cur = conn.execute("SELECT project_id FROM projects WHERE project_id=?", (base_slug,))
        if cur.fetchone() is None:
            return base_slug
        i = 2
        while True:
            candidate = f"{base_slug}-{i}"
            cur = conn.execute("SELECT project_id FROM projects WHERE project_id=?", (candidate,))
            if cur.fetchone() is None:
                return candidate
            i += 1


def _create_empty_pbc_xlsx(xlsx_path: Path) -> None:
    """在 xlsx_path 创建一张空 PBC 清单（v7: 15 列表头，含「需求期间」）。

    必填字段用浅红底 + 红色星标 + 数据验证；选填字段灰色。
    必填：资料编号 / 一级分类 / 问题/需求描述 / 期望提供日期 / 实体归属 / 需求期间
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "PBC清单"
    headers = [
        ("一级分类", True, "必填。归档文件夹按此分类建，如 历史沿革 / 货币资金 / 穿行测试"),
        ("二级分类", True, "必填。该分类下的编号项，如 历-1 / 存-4。编号前缀对应一级分类缩写"),
        ("相关科目", False, "选填。相关会计科目，如 银行存款 / 应收账款"),
        ("资料名称", True, "必填。资料简短名称，如 股权架构图 / 银行流水 / 存货盘点表"),
        ("问题/需求描述", True, "必填。该资料的详细需求说明，归档命名会用"),
        ("报告期间", True, "必填。该资料需覆盖的期间，如 2023年度/2024年度/2025年度/2026年一季度"),
        ("格式", False, "选填。期望格式，如 PDF / Excel / 扫描件"),
        ("优先级", False, "选填。高 / 中 / 低"),
        ("提出时间", False, "选填。日期，如 2026-07-01"),
        ("期望提供日期", True, "必填。日期，超 5 工作日触发风险雷达"),
        ("逾期天数", False, "选填。自动计算，留空即可"),
        ("资料提供情况", False, "选填。状态机自动管理，留空默认 未提供"),
        ("备注", False, "选填。异常情况记录"),
        ("实体归属", True, "必填。公司级 vs 集团级，如 ABC子公司 / 集团合并"),
        ("置信度", False, "选填。AI 自动回填，留空"),
        ("文件路径", False, "选填。AI 归档后自动回填，留空"),
    ]
    required_fill = PatternFill(start_color="FCEBEB", end_color="FCEBEB", fill_type="solid")
    optional_fill = PatternFill(start_color="F1EFE8", end_color="F1EFE8", fill_type="solid")
    for i, (h, required, comment_text) in enumerate(headers, 1):
        cell = ws.cell(1, i, ("* " if required else "") + h)
        cell.font = Font(bold=True, color="A32D2D" if required else "5F5E5A")
        cell.fill = required_fill if required else optional_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.comment = Comment(comment_text, "PBC工作站")
    widths = [14, 12, 16, 20, 40, 30, 8, 8, 14, 14, 10, 16, 24, 14, 10, 40]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    dv = DataValidation(
        type="list",
        formula1='"未提供,已提供，审核中,已提供,不适用"',
        allow_blank=True,
    )
    dv.add("L2:L1000")
    ws.add_data_validation(dv)
    ws.freeze_panes = "C2"
    wb.save(str(xlsx_path))
    wb.close()


def create_project(
    name: str,
    client_name: str = "",
    base_dir: Path | None = None,
    project_id: str | None = None,
    note: str = "",
) -> dict:
    """创建新项目：
    1. 生成 project_id（用 slug，如 "demo" / "huayinkang-ipo"）
    2. 在 base_dir 下创建项目目录（projects/{project_id}/）
    3. 在项目目录下创建空的 01_PBC_List.xlsx（14 列表头）
    4. 在项目目录下创建空的 客户共享文件夹/ 和 archives/
    5. 写入 projects 表
    6. 返回项目元数据
    """
    from app.config import PROJECTS_DIR
    base = Path(base_dir) if base_dir else PROJECTS_DIR
    base.mkdir(parents=True, exist_ok=True)

    # 生成 project_id
    if project_id:
        pid = slugify(project_id)
    else:
        pid = slugify(name)
    pid = _ensure_unique_slug(pid)

    folder_path = base / f"project_{pid}"
    folder_path.mkdir(parents=True, exist_ok=True)

    pbc_list_path = folder_path / "01_PBC_List.xlsx"
    if not pbc_list_path.exists():
        _create_empty_pbc_xlsx(pbc_list_path)

    client_folder = folder_path / "客户共享文件夹"
    client_folder.mkdir(parents=True, exist_ok=True)

    archive_root = folder_path / "archives"
    archive_root.mkdir(parents=True, exist_ok=True)

    now = datetime.now().isoformat(timespec="seconds")
    record = {
        "project_id": pid,
        "name": name,
        "client_name": client_name or "",
        "folder_path": str(folder_path),
        "pbc_list_path": str(pbc_list_path),
        "client_folder": str(client_folder),
        "archive_root": str(archive_root),
        "created_at": now,
        "updated_at": now,
        "is_active": 1,
        "note": note or "",
    }

    # P0-9: 写入包 retry
    execute_with_retry(
        """INSERT INTO projects
           (project_id, name, client_name, folder_path, pbc_list_path,
            client_folder, archive_root, created_at, updated_at, is_active, note)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            pid, name, client_name or "", str(folder_path), str(pbc_list_path),
            str(client_folder), str(archive_root), now, now, 1, note or "",
        ),
    )
    return record


def list_projects(active_only: bool = True) -> list[dict]:
    """列出所有项目。"""
    with get_conn() as conn:
        if active_only:
            cur = conn.execute(
                "SELECT * FROM projects WHERE is_active=1 ORDER BY created_at ASC"
            )
        else:
            cur = conn.execute(
                "SELECT * FROM projects ORDER BY is_active DESC, created_at ASC"
            )
        rows = cur.fetchall()
    return [dict(r) for r in rows]


def get_project(project_id: str) -> Optional[dict]:
    """获取单个项目元数据；不存在返回 None。"""
    if not project_id:
        return None
    with get_conn() as conn:
        cur = conn.execute("SELECT * FROM projects WHERE project_id=?", (project_id,))
        row = cur.fetchone()
    return dict(row) if row is not None else None


def update_project(project_id: str, **fields) -> dict:
    """更新项目元数据（如配置 client_folder 路径）。

    允许更新的字段：name / client_name / folder_path / pbc_list_path /
    client_folder / archive_root / is_active / note
    """
    if not project_id:
        raise ValueError("project_id 不能为空")

    allowed = {
        "name", "client_name", "folder_path", "pbc_list_path",
        "client_folder", "archive_root", "is_active", "note",
    }
    updates = {k: v for k, v in fields.items() if k in allowed and v is not None}
    if not updates:
        return get_project(project_id) or {"project_id": project_id, "error": "no fields to update"}

    updates["updated_at"] = datetime.now().isoformat(timespec="seconds")

    set_clause = ", ".join(f"{k}=?" for k in updates.keys())
    values = list(updates.values()) + [project_id]

    with get_conn() as conn:
        conn.execute(
            f"UPDATE projects SET {set_clause} WHERE project_id=?",
            values,
        )
    return get_project(project_id) or {"project_id": project_id, **updates}


def delete_project(project_id: str, soft: bool = True) -> dict:
    """删除项目。
    - soft=True：只标记 is_active=0，不真删文件
    - soft=False：从 SQLite 真删，但**不删磁盘文件**（防止误删用户数据）
    """
    if not project_id:
        raise ValueError("project_id 不能为空")

    existing = get_project(project_id)
    if existing is None:
        return {"ok": False, "error": f"项目不存在: {project_id}"}

    if soft:
        execute_with_retry(
            "UPDATE projects SET is_active=0, updated_at=? WHERE project_id=?",
            (datetime.now().isoformat(timespec="seconds"), project_id),
        )
        return {"ok": True, "project_id": project_id, "soft_deleted": True}
    else:
        execute_with_retry(
            "DELETE FROM projects WHERE project_id=?",
            (project_id,),
        )
        return {"ok": True, "project_id": project_id, "soft_deleted": False}


def get_or_create_default_project() -> dict:
    """启动时检查是否有项目，没有就创建"示例项目"（project_id="demo"）。

    "示例项目"的 PBC 清单从 mock_data/01_PBC_List.xlsx 复制一份作为初始数据，
    方便用户开箱即可看到带数据的演示。
    """
    existing = get_project("demo")
    if existing is not None:
        return existing

    from app.config import MOCK_DATA_DIR, PROJECTS_DIR
    record = create_project(
        name="示例项目",
        client_name="ABC 集团（演示）",
        base_dir=PROJECTS_DIR,
        project_id="demo",
        note="系统自动创建的示例项目，PBC 清单从 mock_data 复制",
    )

    # Demo 录制"从 0 到 1"：默认空清单（只表头），让用户自己导入 PBC 清单
    # 不再自动复制 6 项或 103 项，让展示逻辑更清晰（用户点"导入清单"才有数据）
    demo_pbc = Path(record["pbc_list_path"])
    if not demo_pbc.exists() or demo_pbc.stat().st_size == 0:
        # v7: 复用 _create_empty_pbc_xlsx（15 列表头 + 必填标注 + 数据验证）
        try:
            _create_empty_pbc_xlsx(demo_pbc)
            logger.info("demo 空清单已创建（v7: 15 列表头 + 必填标注）")
        except Exception as e:
            logger.warning("创建空清单失败: %r", e)
            # fallback：如果空清单创建失败，从 mock_data 复制
            mock_pbc = MOCK_DATA_DIR / "01_PBC_List.xlsx"
            if mock_pbc.exists():
                import shutil
                shutil.copy2(str(mock_pbc), str(demo_pbc))

    # 注：mock_data 的客户共享文件夹内容**不**自动复制到 demo 项目。
    # 原因：watchdog 启动时会扫描已有文件并触发 AI 调用，24 个并发会拖垮服务器。
    # 用户想演示 scan-folder 时，可手动调 POST /api/files/demo/scan-folder，
    # 或手动复制 mock_data/客户共享文件夹/* 到 projects/project_demo/客户共享文件夹/
    # 如果一定要启动时复制，设环境变量 PBC_DEMO_COPY_CLIENT=1。
    import os
    if os.environ.get("PBC_DEMO_COPY_CLIENT") == "1":
        mock_client = MOCK_DATA_DIR / "客户共享文件夹"
        demo_client = Path(record["client_folder"])
        if mock_client.exists():
            try:
                import shutil
                for p in mock_client.rglob("*"):
                    rel = p.relative_to(mock_client)
                    dest = demo_client / rel
                    if p.is_dir():
                        dest.mkdir(parents=True, exist_ok=True)
                    elif p.is_file():
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        shutil.copy2(str(p), str(dest))
            except Exception:
                pass

    # 更新 note
    return get_project("demo") or record


# ----------------------------------------------------------------------
# M4: tasks 表操作
# ----------------------------------------------------------------------
_TASK_FIELDS = (
    "task_id", "status", "progress", "total", "done_count",
    "source", "folder", "received_json", "results_json",
    "started_at", "finished_at", "error", "project_id",
)


def upsert_task(task_id: str, **fields: Any) -> None:
    """写入/更新任务（INSERT OR REPLACE 风格的增量更新）。

    支持任意字段子集；不存在则插入，存在则更新指定字段。
    首次插入需要至少提供 status + started_at。
    """
    if not task_id:
        raise ValueError("task_id 不能为空")

    # 取当前已存在的行（若 any）
    existing: dict[str, Any] = {}
    try:
        with get_conn() as conn:
            cur = conn.execute("SELECT * FROM tasks WHERE task_id=?", (task_id,))
            row = cur.fetchone()
            if row is not None:
                existing = dict(row)
    except Exception:
        existing = {}

    # 合并字段
    merged = {**existing, **fields, "task_id": task_id}

    # 复杂类型走 JSON
    for k in ("received_json", "results_json"):
        if k in merged and not isinstance(merged[k], (str, type(None))):
            try:
                merged[k] = json.dumps(merged[k], ensure_ascii=False)
            except Exception:
                merged[k] = None

    # 保证 started_at 存在（首次插入）
    if not merged.get("started_at"):
        merged["started_at"] = datetime.now().isoformat(timespec="seconds")

    cols = list(_TASK_FIELDS)
    values = [merged.get(c) for c in cols]

    # P0-9: 写入包 retry
    placeholders = ",".join("?" for _ in cols)
    col_list = ",".join(cols)
    execute_with_retry(
        f"INSERT OR REPLACE INTO tasks ({col_list}) VALUES ({placeholders})",
        tuple(values),
    )


def get_task(task_id: str) -> Optional[dict[str, Any]]:
    """查询单个任务；不存在返回 None。返回的 dict 把 received_json/results_json 解析回 list/dict。"""
    with get_conn() as conn:
        cur = conn.execute("SELECT * FROM tasks WHERE task_id=?", (task_id,))
        row = cur.fetchone()
    if row is None:
        return None
    return _task_row_to_dict(row)


def list_recent_tasks(limit: int = 20, project_id: Optional[str] = None) -> list[dict[str, Any]]:
    """列出最近 N 个任务（按 started_at DESC）。

    传 project_id：只返回该项目的任务。
    """
    limit = max(1, min(int(limit), 500))
    with get_conn() as conn:
        if project_id:
            cur = conn.execute(
                "SELECT * FROM tasks WHERE project_id=? ORDER BY started_at DESC LIMIT ?",
                (project_id, limit),
            )
        else:
            cur = conn.execute(
                "SELECT * FROM tasks ORDER BY started_at DESC LIMIT ?", (limit,)
            )
        rows = cur.fetchall()
    return [_task_row_to_dict(r) for r in rows]


def _task_row_to_dict(row: sqlite3.Row) -> dict[str, Any]:
    d = dict(row)
    for k in ("received_json", "results_json"):
        v = d.get(k)
        if isinstance(v, str) and v:
            try:
                d[k] = json.loads(v)
            except Exception:
                # 保留原始字符串
                pass
    return d


# ----------------------------------------------------------------------
# M4: file_archive 表操作
# ----------------------------------------------------------------------
def insert_archive(
    item_id: str,
    original_path: str,
    archived_path: str,
    sha256: Optional[str],
    file_size: Optional[int],
    entity: Optional[str],
    archived_by: str,
    project_id: Optional[str] = None,
    is_directory: int = 0,
    version: Optional[str] = None,
    doc_type: Optional[str] = None,
    category: Optional[str] = None,
) -> int:
    """插入一条归档记录，返回新 id。UNIQUE(archived_path) 冲突时忽略。

    v7: 加 is_directory / version / doc_type / category 字段。
    P0-9: 写入包 retry。
    """
    cur = execute_with_retry(
        """INSERT OR IGNORE INTO file_archive
           (item_id, original_path, archived_path, sha256, file_size,
            entity, archived_at, archived_by, project_id,
            is_directory, version, doc_type, category)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            item_id or "",
            original_path or "",
            archived_path or "",
            sha256 or "",
            file_size,
            entity or "",
            datetime.now().isoformat(timespec="seconds"),
            archived_by or "",
            project_id or "",
            is_directory,
            version,
            doc_type,
            category or "",
        ),
    )
    return int(cur.lastrowid or 0)


def get_archive_by_item(item_id: str, project_id: Optional[str] = None) -> list[dict[str, Any]]:
    """按 item_id 查归档文件列表。"""
    with get_conn() as conn:
        if project_id:
            cur = conn.execute(
                "SELECT * FROM file_archive WHERE item_id=? AND project_id=? ORDER BY archived_at DESC",
                (item_id, project_id),
            )
        else:
            cur = conn.execute(
                "SELECT * FROM file_archive WHERE item_id=? ORDER BY archived_at DESC",
                (item_id,),
            )
        return [dict(r) for r in cur.fetchall()]


def get_archive_by_sha(sha256: str, project_id: Optional[str] = None) -> Optional[dict[str, Any]]:
    """按 SHA-256 查归档（用于去重）；找不到返回 None。

    传 project_id：只在该项目内查（同 SHA 不同项目可独立归档）。
    """
    if not sha256:
        return None
    with get_conn() as conn:
        if project_id:
            cur = conn.execute(
                "SELECT * FROM file_archive WHERE sha256=? AND project_id=? LIMIT 1",
                (sha256, project_id),
            )
        else:
            cur = conn.execute(
                "SELECT * FROM file_archive WHERE sha256=? LIMIT 1", (sha256,)
            )
        row = cur.fetchone()
    return dict(row) if row is not None else None


def list_recent_archives(limit: int = 100, project_id: Optional[str] = None) -> list[dict[str, Any]]:
    """列出最近 N 条归档记录（GET /api/files/list 用）。"""
    limit = max(1, min(int(limit), 500))
    with get_conn() as conn:
        if project_id:
            cur = conn.execute(
                "SELECT * FROM file_archive WHERE project_id=? ORDER BY archived_at DESC, id DESC LIMIT ?",
                (project_id, limit),
            )
        else:
            cur = conn.execute(
                "SELECT * FROM file_archive ORDER BY archived_at DESC, id DESC LIMIT ?",
                (limit,),
            )
        return [dict(r) for r in cur.fetchall()]


def delete_archive_by_item(item_id: str, project_id: Optional[str] = None) -> int:
    """删除指定 item_id 的归档记录（改分类时用）。

    返回删除的行数。注意：只删 DB 记录，不删归档目录里的文件（文件由调用方处理）。
    """
    with get_conn() as conn:
        if project_id:
            cur = conn.execute(
                "DELETE FROM file_archive WHERE item_id=? AND project_id=?",
                (item_id, project_id),
            )
        else:
            cur = conn.execute(
                "DELETE FROM file_archive WHERE item_id=?",
                (item_id,),
            )
        conn.commit()
        return cur.rowcount


def delete_archive_by_path(archived_path: str, project_id: Optional[str] = None) -> int:
    """按归档路径删除单条归档记录（改分类单文件时用）。"""
    with get_conn() as conn:
        if project_id:
            cur = conn.execute(
                "DELETE FROM file_archive WHERE archived_path=? AND project_id=?",
                (archived_path, project_id),
            )
        else:
            cur = conn.execute(
                "DELETE FROM file_archive WHERE archived_path=?",
                (archived_path,),
            )
        conn.commit()
        return cur.rowcount


# ----------------------------------------------------------------------
# v7.6: 文件变更日志（持久化操作记录）
# ----------------------------------------------------------------------

def insert_change_log(
    project_id: Optional[str],
    file_name: str,
    change_type: str,
    item_id: Optional[str] = None,
    sha256: Optional[str] = None,
    changed_by: str = "manual",
    detail: str = "",
) -> int:
    """写入一条变更日志。返回新 id。

    change_type: added/modified/deleted/archived/reclassified/approved/missing
    changed_by: watchdog / ai-auto / manual
    """
    now = datetime.now().isoformat(timespec="seconds")
    with get_conn() as conn:
        cur = execute_with_retry(
            """INSERT INTO file_change_log
               (project_id, file_name, sha256, change_type, item_id, changed_by, changed_at, detail)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                project_id or "",
                file_name or "",
                sha256 or "",
                change_type,
                item_id or "",
                changed_by,
                now,
                detail,
            ),
        )
        conn.commit()
        return cur.lastrowid


def get_change_log(
    project_id: Optional[str] = None,
    change_type: Optional[str] = None,
    limit: int = 100,
) -> list[dict[str, Any]]:
    """查询变更日志（按时间倒序）。"""
    limit = max(1, min(int(limit), 500))
    with get_conn() as conn:
        sql = "SELECT * FROM file_change_log"
        params: list = []
        conditions: list[str] = []
        if project_id:
            conditions.append("project_id = ?")
            params.append(project_id)
        if change_type:
            conditions.append("change_type = ?")
            params.append(change_type)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY changed_at DESC, id DESC LIMIT ?"
        params.append(limit)
        cur = conn.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


# ----------------------------------------------------------------------
# v7.7: 待确认队列（pending_confirm CRUD）
# ----------------------------------------------------------------------

def insert_pending_confirm(
    project_id: Optional[str],
    file_path: str,
    file_name: str = "",
    sha256: str = "",
    suggested_item_id: str = "",
    confidence: float = 0.0,
    decision: str = "",
    conflict_signal: str = "",
    advisory_notes: str = "",
) -> int:
    """写入一条待确认记录。返回新 id。
    
    v7.7: 去重——同 project_id + file_path + confirmed=0 已存在则不重复插入。
    """
    now = datetime.now().isoformat(timespec="seconds")
    with get_conn() as conn:
        # 先检查是否已有未确认的相同记录
        existing = conn.execute(
            """SELECT id FROM pending_confirm 
               WHERE project_id=? AND file_path=? AND confirmed=0""",
            (project_id or "", file_path),
        ).fetchone()
        if existing:
            # v7.7: 文件被改过重跑AI——更新建议（不重新插入）
            execute_with_retry(
                """UPDATE pending_confirm 
                   SET suggested_item_id=?, confidence=?, decision=?, 
                       conflict_signal=?, advisory_notes=?, created_at=?
                   WHERE id=?""",
                (suggested_item_id, confidence, decision,
                 conflict_signal, advisory_notes, now, existing[0]),
            )
            conn.commit()
            return existing[0]
        cur = execute_with_retry(
            """INSERT INTO pending_confirm
               (project_id, file_path, file_name, sha256, suggested_item_id,
                confidence, decision, conflict_signal, advisory_notes, created_at, confirmed)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)""",
            (
                project_id or "", file_path, file_name, sha256,
                suggested_item_id, confidence, decision,
                conflict_signal, advisory_notes, now,
            ),
        )
        conn.commit()
        return cur.lastrowid


def get_pending_confirm_list(
    project_id: Optional[str] = None,
    confirmed: int = 0,
) -> list[dict[str, Any]]:
    """获取待确认列表。confirmed: 0未确认 1已确认 2已跳过 3已归档。"""
    with get_conn() as conn:
        sql = "SELECT * FROM pending_confirm"
        params: list = []
        conditions: list[str] = []
        if project_id:
            conditions.append("project_id = ?")
            params.append(project_id)
        conditions.append("confirmed = ?")
        params.append(confirmed)
        sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY created_at DESC, id DESC"
        cur = conn.execute(sql, params)
        return [dict(r) for r in cur.fetchall()]


def get_pending_confirm_by_id(confirm_id: int) -> Optional[dict[str, Any]]:
    """按 id 取单条待确认记录。"""
    with get_conn() as conn:
        cur = conn.execute("SELECT * FROM pending_confirm WHERE id = ?", (confirm_id,))
        row = cur.fetchone()
        return dict(row) if row else None


def update_pending_confirm_status(confirm_id: int, confirmed: int) -> bool:
    """更新待确认状态。confirmed: 0未确认 1已确认 2已跳过 3已归档。"""
    with get_conn() as conn:
        cur = conn.execute(
            "UPDATE pending_confirm SET confirmed = ? WHERE id = ?",
            (confirmed, confirm_id),
        )
        conn.commit()
        return cur.rowcount > 0


def update_pending_confirm_item(confirm_id: int, new_item_id: str) -> bool:
    """更新待确认的建议 item_id（改分类时用）。"""
    with get_conn() as conn:
        cur = conn.execute(
            "UPDATE pending_confirm SET suggested_item_id = ? WHERE id = ?",
            (new_item_id, confirm_id),
        )
        conn.commit()
        return cur.rowcount > 0


def get_pending_confirm_count(project_id: Optional[str] = None) -> int:
    """获取未确认数量（前端待确认 tab 角标用）。"""
    with get_conn() as conn:
        sql = "SELECT COUNT(*) FROM pending_confirm WHERE confirmed = 0"
        params: list = []
        if project_id:
            sql += " AND project_id = ?"
            params.append(project_id)
        cur = conn.execute(sql, params)
        return cur.fetchone()[0]


def get_pending_confirm_count_by_item(item_id: str, project_id: Optional[str] = None) -> int:
    """获取某 item_id 下未确认数量（判断是否全部确认完）。"""
    with get_conn() as conn:
        sql = "SELECT COUNT(*) FROM pending_confirm WHERE confirmed = 0 AND suggested_item_id = ?"
        params: list = [item_id]
        if project_id:
            sql += " AND project_id = ?"
            params.append(project_id)
        cur = conn.execute(sql, params)
        return cur.fetchone()[0]


# ----------------------------------------------------------------------
# M4: ai_history 回填 item_id
# ----------------------------------------------------------------------
def update_ai_history_item_id(history_id: int, actual_item_id: str) -> bool:
    """AI 完成识别后，把实际匹配到的 item_id 回填到 ai_history 行。

    Returns:
        True 表示更新成功，False 表示行不存在或未变更。
    """
    if not history_id:
        return False
    cur = execute_with_retry(
        "UPDATE ai_history SET item_id=? WHERE id=?",
        (actual_item_id or "", int(history_id)),
    )
    return cur.rowcount > 0


def get_latest_ai_history_id(action: str, after_id: int = 0) -> Optional[int]:
    """取最近一条 ai_history 的 id（用于回填）。"""
    with get_conn() as conn:
        cur = conn.execute(
            "SELECT id FROM ai_history WHERE action=? AND id>? ORDER BY id DESC LIMIT 1",
            (action, int(after_id)),
        )
        row = cur.fetchone()
    return int(row["id"]) if row is not None else None


if __name__ == "__main__":
    init_db()
    print("db initialized at:", get_config().db_path)
