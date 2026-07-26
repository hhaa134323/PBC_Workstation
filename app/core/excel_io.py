"""Excel 读写层（M2 实现 + 多项目支持）。

读取/更新 PBC 清单：
- read_pbc_list 读取 14 列清单，加 status_normalized / risk_level
- write_pbc_list 用 load_workbook 增量改（保留条件格式/数据验证）
- update_item_status 单条状态更新（含状态机校验 + SQLite state_changes 写入）

多项目支持：所有函数加 project_id 参数；传 project_id 时从 projects 表取该项目
PBC 清单路径；不传时用默认路径（兼容旧调用）。
"""
from __future__ import annotations

import datetime as _dt
import logging
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.config import get_config
from app.core.db import get_conn, get_project

logger = logging.getLogger("pbc.excel_io")

# 列顺序（中文表头 → 英文键），1-based 索引
# v7.2: 前列重构——一级分类→二级分类→相关科目→资料名称→问题/需求描述→报告期间
# 资料编号改名二级分类（编号前缀=一级分类缩写，如历-1=历史沿革下的第1项）
# 新增「资料名称」列（简短名称），原「问题/需求描述」改为更详细的需求说明
# 需求期间改名报告期间，挪到第6列
_COLUMN_MAP: list[tuple[int, str, str]] = [
    (1,  "一级分类",          "category"),
    (2,  "二级分类",          "item_id"),
    (3,  "相关科目",          "subject"),
    (4,  "资料名称",          "doc_name"),
    (5,  "问题/需求描述",     "description"),
    (6,  "报告期间",          "required_period"),
    (7,  "格式",              "file_format"),
    (8,  "优先级",            "priority"),
    (9,  "提出时间",          "raised_at"),
    (10, "期望提供日期",       "expected_by"),
    (11, "逾期天数",          "overdue_days"),
    (12, "资料提供情况",       "status_raw"),
    (13, "备注",              "remark"),
    (14, "实体归属",          "entity"),
    (15, "置信度",            "confidence"),
    (16, "文件路径",          "file_path"),
]

# 可更新字段（write_pbc_list 增量改只动这些列）
# v7.2: 列索引跟着新顺序调整
_UPDATABLE_FIELDS: dict[str, int] = {
    "status_raw": 12,
    "confidence": 15,
    "file_path":  16,
    "remark":     13,
    "required_period": 6,
}

# 标准状态枚举
STATUS_NOT_PROVIDED = "未提供"
STATUS_REVIEWING    = "已提供，审核中"
STATUS_PROVIDED     = "已提供"
STATUS_NA           = "不适用"
STATUS_PENDING      = "待定"

# 合法状态流转（from -> set(to)）
# P0-5: 已提供 → 已提供，审核中（撤销归档，Manager 权限）
_TRANSITIONS: dict[str, set[str]] = {
    STATUS_NOT_PROVIDED: {STATUS_REVIEWING, STATUS_PROVIDED, STATUS_NA},  # v7.7: 加直接到已提供（auto批量确认）
    STATUS_REVIEWING:    {STATUS_PROVIDED, STATUS_NOT_PROVIDED, STATUS_NA},
    STATUS_PROVIDED:     {STATUS_REVIEWING, STATUS_NA},
    STATUS_NA:           {STATUS_NA},
    STATUS_PENDING:      {STATUS_REVIEWING, STATUS_NA, STATUS_NOT_PROVIDED},
}


# ---------------------------------------------------------------------------
# 公共工具函数
# ---------------------------------------------------------------------------

def normalize_status(status_raw: str | None) -> str:
    """将原始状态文本归一化到 5 种标准状态。

    规则（思路 6.1 + 待定）：
      - 精确等于 "已提供"           → "已提供"
      - 含 "审核中"                  → "已提供，审核中"
      - 含 "未提供" / "待定" / "部分" / "纳税申报表"  → "未提供"
      - 含 "不适用"                  → "不适用"
      - 精确等于 "待定"              → "待定"
      - 其他                          → "未提供"（兜底）
    """
    if not status_raw or not isinstance(status_raw, str):
        return STATUS_PENDING
    s = status_raw.strip()
    if s == STATUS_PROVIDED:
        return STATUS_PROVIDED
    if "审核中" in s:
        return STATUS_REVIEWING
    if "不适用" in s:
        return STATUS_NA
    if s == STATUS_PENDING:
        return STATUS_PENDING
    if ("未提供" in s) or ("部分" in s) or ("纳税申报表" in s):
        return STATUS_NOT_PROVIDED
    if "待定" in s:
        # 含 "待定" 但不精确等于 "待定"（例如 "暂定..." 不归到这里）
        return STATUS_PENDING
    # 兜底
    return STATUS_NOT_PROVIDED


def compute_risk_level(overdue_days: int | float | None) -> str:
    """风险等级：high / medium / low / none。"""
    if overdue_days is None:
        return "none"
    try:
        d = float(overdue_days)
    except (TypeError, ValueError):
        return "none"
    if d > 100:
        return "high"
    if d > 30:
        return "medium"
    if d > 0:
        return "low"
    return "none"


def is_valid_transition(old: str, new: str) -> bool:
    """状态机流转合法性（思路 6.1）。

    允许：
      未提供 → 已提供，审核中
      已提供，审核中 → 已提供 / 未提供（Senior 退回）
      任何 → 不适用
    其他流转拒绝。
    """
    if old == new:
        return True
    allowed = _TRANSITIONS.get(old, set())
    return new in allowed


# ---------------------------------------------------------------------------
# 内部辅助
# ---------------------------------------------------------------------------

def _parse_date(value: Any) -> Optional[_dt.date]:
    """把 Excel 单元格值转 datetime.date，无法解析返回 None。"""
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # 尝试 ISO: 2025-12-08
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return _dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        # 含 "待定" 等非日期字符串
        return None
    return None


def _parse_overdue_days(value: Any) -> Optional[int]:
    """逾期天数：int 直接返回；str 尝试解析，失败 None。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # 取数字前缀（如 "186" / "暂定2026..." 这种取不到，返回 None）
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def _parse_confidence(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _row_to_item(row_idx: int, ws) -> dict[str, Any]:
    """把一个 worksheet row 转成 PBCItem dict。"""
    item: dict[str, Any] = {}
    for col_idx, _zh, key in _COLUMN_MAP:
        item[key] = ws.cell(row_idx, col_idx).value

    # 类型转换
    item["raised_at"]    = _parse_date(item.get("raised_at"))
    item["expected_by"]  = _parse_date(item.get("expected_by"))
    item["overdue_days"] = _parse_overdue_days(item.get("overdue_days"))
    item["confidence"]   = _parse_confidence(item.get("confidence"))
    # v7: required_period 保留字符串原样（如 "2023年度/2024年度/2025年度/2026年一季度"）
    rp = item.get("required_period")
    item["required_period"] = str(rp).strip() if rp is not None and str(rp).strip() else ""

    # 衍生字段
    status_raw = item.get("status_raw")
    item["status_raw"]      = status_raw if status_raw is not None else ""
    item["status_normalized"] = normalize_status(status_raw)

    # v7.7: 实时计算 overdue_days（每天自动更新）
    # 如果有 expected_by 日期，且状态不是已提供/不适用，实时算超期天数
    eb = item.get("expected_by")
    if eb:
        from datetime import date
        try:
            if isinstance(eb, str):
                from datetime import datetime
                eb_date = datetime.fromisoformat(eb).date() if 'T' in eb or '-' in eb else None
            elif isinstance(eb, (int, float)):
                # Excel 日期序列号
                from datetime import datetime, timedelta
                base = datetime(1899, 12, 30)
                eb_date = (base + timedelta(days=int(eb))).date()
            else:
                eb_date = None
            if eb_date:
                today = date.today()
                delta = (today - eb_date).days
                # 只在未提供状态下才算超期
                sn = item["status_normalized"]
                if sn not in ("已提供", "不适用") and delta > 0:
                    item["overdue_days"] = delta
        except Exception:
            pass
    item["risk_level"]       = compute_risk_level(item["overdue_days"])
    return item


def _find_item_row(ws, item_id: str) -> Optional[int]:
    """在第2列（二级分类/item_id）里找所在行号（1-based）；找不到返回 None。"""
    target = str(item_id).strip()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v is not None and str(v).strip() == target:
            return r
    return None


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def _resolve_xlsx_path(project_id: Optional[str] = None, xlsx_path: Optional[str] = None) -> str:
    """解析 PBC 清单路径：
    1. 显式传 xlsx_path：直接用（最高优先级，兼容旧测试）
    2. 传 project_id：从 projects 表取该项目 PBC 清单路径
    3. 都不传：用全局 config.pbc_list_path（兼容旧调用）
    """
    if xlsx_path:
        return xlsx_path
    if project_id:
        proj = get_project(project_id)
        if proj is None:
            raise FileNotFoundError(f"项目不存在: {project_id}")
        return proj.get("pbc_list_path") or ""
    return str(get_config().pbc_list_path)


def read_pbc_list(
    xlsx_path: Optional[str] = None,
    project_id: Optional[str] = None,
) -> list[dict[str, Any]]:
    """读取 PBC 清单 Excel，返回 list[PBCItem]。

    传 project_id：从 projects 表取该项目 PBC 清单路径。
    传 xlsx_path：直接用指定路径（兼容旧调用，最高优先级）。
    都不传：用全局 config.pbc_list_path。

    v7.5: 对损坏的 Excel 文件做容错——返回空列表 + 日志，不抛 500。
    """
    path_str = _resolve_xlsx_path(project_id=project_id, xlsx_path=xlsx_path)
    p = Path(path_str)
    if not p.exists():
        raise FileNotFoundError(f"PBC 清单不存在: {path_str}")

    try:
        wb = load_workbook(str(p), data_only=True)
    except Exception as e:
        logger.error("PBC 清单读取失败（文件可能损坏）: %s → %r", path_str, e)
        # 返回空列表而不是抛异常，避免 500
        return []

    ws = wb.active
    items: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        # 跳过完全空行（item_id/二级分类 为空）
        if ws.cell(r, 2).value in (None, ""):
            continue
        items.append(_row_to_item(r, ws))
    wb.close()
    return items


def get_item_by_id(
    item_id: str,
    xlsx_path: Optional[str] = None,
    project_id: Optional[str] = None,
) -> Optional[dict[str, Any]]:
    """单条查询；不存在返回 None。

    v7.5: 对损坏文件做容错。
    """
    path_str = _resolve_xlsx_path(project_id=project_id, xlsx_path=xlsx_path)
    p = Path(path_str)
    if not p.exists():
        raise FileNotFoundError(f"PBC 清单不存在: {path_str}")
    try:
        wb = load_workbook(str(p), data_only=True)
    except Exception as e:
        logger.error("PBC 清单读取失败（文件可能损坏）: %s → %r", path_str, e)
        return None
    ws = wb.active
    row = _find_item_row(ws, item_id)
    if row is None:
        wb.close()
        return None
    item = _row_to_item(row, ws)
    wb.close()
    return item


def write_pbc_list(
    rows: list[dict[str, Any]],
    xlsx_path: Optional[str] = None,
    project_id: Optional[str] = None,
) -> None:
    """增量更新 PBC 清单（保留条件格式/数据验证）。

    - 只更新 status_raw / confidence / file_path / remark 字段
    - 按 item_id 匹配行
    - 不重写整张表，不动其他列
    """
    path_str = _resolve_xlsx_path(project_id=project_id, xlsx_path=xlsx_path)
    p = Path(path_str)
    if not p.exists():
        raise FileNotFoundError(f"PBC 清单不存在: {path_str}")

    # 用 load_workbook（不 write_only），保留条件格式
    wb = load_workbook(str(p))
    ws = wb.active

    updated = 0
    for row_data in rows:
        item_id = row_data.get("item_id")
        if not item_id:
            continue
        row_idx = _find_item_row(ws, str(item_id))
        if row_idx is None:
            logger.warning("write_pbc_list: 未找到 item_id=%s，跳过", item_id)
            continue
        for field_name, col_idx in _UPDATABLE_FIELDS.items():
            if field_name in row_data:
                ws.cell(row_idx, col_idx).value = row_data[field_name]
        updated += 1

    wb.save(str(p))
    wb.close()
    logger.info("write_pbc_list 完成，更新 %d 行", updated)


def update_item_status(
    item_id: str,
    new_status: str,
    changed_by: str,
    note: str = "",
    xlsx_path: Optional[str] = None,
    project_id: Optional[str] = None,
) -> tuple[bool, str, Optional[dict[str, Any]]]:
    """单条状态更新（P0-4: 失败时 rollback 已写 Excel + 不写 SQLite）。

    流程：
      1. 取当前 item
      2. 状态机校验失败 → 直接返回 False，不写 Excel，不写 SQLite
      3. 备份旧 status_raw → 写 Excel
      4. 写 SQLite state_changes
         - SQLite 失败 → 把 Excel 改回旧 status_raw（rollback）→ 返回 False
      5. 重新读取返回最新 item

    Returns:
        (success, message, updated_item)
        success=True 时 updated_item 为最新 item；False 时为 None
    """
    path_str = _resolve_xlsx_path(project_id=project_id, xlsx_path=xlsx_path)

    # 1. 取当前 item
    current = get_item_by_id(item_id, xlsx_path=path_str)
    if current is None:
        return False, f"未找到资料项: {item_id}", None

    old_raw = current.get("status_raw", "") or ""
    old_norm = normalize_status(old_raw)
    new_norm = normalize_status(new_status)

    # 2. 状态机校验（在归一化后的标准状态上判断）
    # P0-4: 失败直接返回，不写 Excel，不写 SQLite
    if not is_valid_transition(old_norm, new_norm):
        return (
            False,
            f"非法状态流转: {old_norm!r} → {new_norm!r}（旧原始状态: {old_raw!r}）",
            None,
        )

    # 3. 写 Excel（仅 status_raw；用户传入的 new_status 视为新的原始状态）
    # P0-4: 先记录 old_raw 用于 rollback
    write_pbc_list([{
        "item_id": item_id,
        "status_raw": new_status,
    }], xlsx_path=path_str)

    # 4. 写 SQLite state_changes（带 project_id）
    # P0-4: SQLite 写入失败 → rollback Excel
    # P0-9: 写入包 retry（execute_with_retry）
    sqlite_ok = False
    sqlite_err = ""
    try:
        from app.core.db import execute_with_retry
        execute_with_retry(
            """INSERT INTO state_changes
               (item_id, old_status, new_status, changed_by, note, project_id)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (item_id, old_raw, new_status, changed_by, note or "", project_id or ""),
        )
        sqlite_ok = True
    except Exception as e:
        sqlite_err = f"{type(e).__name__}: {e}"
        logger.error("SQLite state_changes 写入失败（开始 rollback Excel）: %s", sqlite_err)
        # rollback：把 Excel 中的 status_raw 改回 old_raw
        try:
            write_pbc_list([{
                "item_id": item_id,
                "status_raw": old_raw,
            }], xlsx_path=path_str)
            logger.info("Excel rollback 成功: item_id=%s 恢复为 %r", item_id, old_raw)
        except Exception as rb_err:
            logger.error(
                "Excel rollback 失败！item_id=%s 当前 Excel 状态可能为 %r（应为 %r）: %r",
                item_id, new_status, old_raw, rb_err,
            )
        return (
            False,
            f"SQLite 写入失败已 rollback Excel: {sqlite_err}",
            None,
        )

    # 5. 重新读取最新 item 返回
    updated = get_item_by_id(item_id, xlsx_path=path_str)
    if updated is None:
        # 不应该发生，但保险一下
        return False, "更新后重新读取失败", None

    msg = f"状态更新成功: {old_norm} → {new_norm}"

    # v7.6: 写变更日志（状态变成"已提供"时记 approved）
    if new_norm == "已提供":
        try:
            from app.core.db import insert_change_log
            insert_change_log(
                project_id=project_id,
                file_name=item_id,
                change_type="approved",
                item_id=item_id,
                changed_by="manual",
                detail=f"复核通过: {old_norm} -> 已提供" + (f" ({note})" if note else ""),
            )
        except Exception:
            pass  # 日志写入失败不阻断

    return True, msg, updated
