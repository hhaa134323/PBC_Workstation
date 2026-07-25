"""PDF/Excel 智能解析（M3 实现）。

统一入口 parse_file 按扩展名分发：
- .pdf  → parse_pdf（pdfplumber），文字 < 50 字符标记 needs_ocr=True
- .xlsx/.xls/.csv → parse_excel（openpyxl 读 sheet 表头 + 前 20 行 → markdown）
- 其他文本类型（.txt/.md/.csv）直接读 UTF-8
- 其他二进制 → 仅取 metadata

约束：
- 路径全部 safe_path 包装
- 读前调 file_stable_size 等待大小稳定
- 50MB 以上直接拒绝 {"error": "file_too_large"}
- PDF 解析失败回退 extract_text_with_vision（Qwen3-VL-Plus）
"""
from __future__ import annotations

import base64
import io
import logging
import time
from pathlib import Path
from typing import Any, Optional

from app.config import get_config
from app.utils.path_utils import file_stable_size, safe_path

logger = logging.getLogger("pbc.file_parser")


def _get_ocr_model() -> str:
    """OCR 用模型名（v7.7: 从配置读，跟分类用同一个）。"""
    try:
        from app.core.ai_client import _get_model
        return _get_model()
    except Exception:
        return "qwen-plus"

# 文件大小上限（字节）
MAX_FILE_BYTES = 50 * 1024 * 1024  # 50MB

# 稳定等待秒数
STABLE_SECONDS = 2

# 文本截断长度（喂给 AI 的上限）
TEXT_TRUNCATE = 3000


# ----------------------------------------------------------------------
# 统一入口
# ----------------------------------------------------------------------
def parse_file(file_path: str | Path) -> dict[str, Any]:
    """统一入口，按扩展名分发。

    Returns:
        {
          "ok": True,
          "text": "...",          # 提取的文本（已截断到 3000 字以内供 AI 使用）
          "full_text": "...",     # 完整文本（供前端展示）
          "metadata": {...},
          "needs_ocr": False,     # 是否需要 OCR
        }
        或 {"ok": False, "error": "..."}
    """
    p = safe_path(file_path)
    if not p.exists():
        return {"ok": False, "error": "file_not_found", "path": str(p)}
    if not p.is_file():
        return {"ok": False, "error": "not_a_file", "path": str(p)}

    # 大小检查
    try:
        size = p.stat().st_size
    except OSError as e:
        return {"ok": False, "error": f"stat_failed: {e}"}
    if size > MAX_FILE_BYTES:
        return {
            "ok": False,
            "error": "file_too_large",
            "size": size,
            "limit": MAX_FILE_BYTES,
        }

    # 等待大小稳定
    if not file_stable_size(p, stable_seconds=STABLE_SECONDS, timeout=30):
        logger.warning("file_stable_size 返回 False: %s", p)
        # 不阻断，继续尝试解析

    ext = p.suffix.lower()
    metadata = {
        "path": str(p),
        "name": p.name,
        "ext": ext,
        "size": size,
    }
    try:
        if ext == ".pdf":
            return _parse_pdf(p, metadata)
        if ext in (".xlsx", ".xlsm"):
            return _parse_excel(p, metadata)
        if ext == ".xls":
            # openpyxl 不支持旧版 .xls，提示用户
            return {
                "ok": False,
                "error": "xls_legacy_not_supported",
                "metadata": metadata,
                "hint": "请另存为 .xlsx 后再上传",
            }
        if ext in (".csv", ".txt", ".md", ".json", ".xml"):
            return _parse_text(p, metadata)
        # 其他类型：只返回 metadata
        return {
            "ok": True,
            "text": "",
            "full_text": "",
            "metadata": metadata,
            "needs_ocr": False,
            "note": f"暂不支持解析 {ext} 文件，仅返回元数据",
        }
    except Exception as e:
        logger.exception("parse_file failed: %s", p)
        return {"ok": False, "error": f"{type(e).__name__}: {e}", "metadata": metadata}


# ----------------------------------------------------------------------
# PDF
# ----------------------------------------------------------------------
def parse_pdf(file_path: str | Path) -> dict[str, Any]:
    """解析 PDF，返回结构化字段。"""
    p = safe_path(file_path)
    return _parse_pdf(p, {"path": str(p), "name": p.name, "ext": ".pdf"})


def _parse_pdf(p: Path, metadata: dict[str, Any]) -> dict[str, Any]:
    """内部 PDF 解析实现。"""
    try:
        import pdfplumber
    except ImportError:
        return {
            "ok": True,
            "text": "",
            "full_text": "",
            "metadata": metadata,
            "needs_ocr": True,
            "error": "pdfplumber_not_installed",
        }

    pages_text: list[str] = []
    n_pages = 0
    try:
        with pdfplumber.open(str(p)) as pdf:
            n_pages = len(pdf.pages)
            for page in pdf.pages:
                t = page.extract_text() or ""
                pages_text.append(t)
    except Exception as e:
        logger.warning("pdfplumber 解析失败 %s: %r，将标记 needs_ocr", p, e)
        return {
            "ok": True,
            "text": "",
            "full_text": "",
            "metadata": {**metadata, "pages": n_pages},
            "needs_ocr": True,
            "error": "pdf_unreadable",
            "pdf_error": f"{type(e).__name__}: {e}",
        }

    full = "\n".join(pages_text).strip()
    needs_ocr = len(full) < 50  # 文字过少，疑似扫描件
    return {
        "ok": True,
        "text": full[:TEXT_TRUNCATE],
        "full_text": full,
        "metadata": {**metadata, "pages": n_pages, "char_count": len(full)},
        "needs_ocr": needs_ocr,
    }


# ----------------------------------------------------------------------
# Excel
# ----------------------------------------------------------------------
def parse_excel(file_path: str | Path) -> dict[str, Any]:
    """解析 Excel，转成 markdown 表格字符串。"""
    p = safe_path(file_path)
    return _parse_excel(p, {"path": str(p), "name": p.name, "ext": p.suffix.lower()})


def _parse_excel(p: Path, metadata: dict[str, Any]) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "ok": False,
            "error": "openpyxl not installed",
            "metadata": metadata,
        }

    chunks: list[str] = []
    sheet_info: list[dict[str, Any]] = []
    try:
        wb = load_workbook(str(p), data_only=True, read_only=True)
    except Exception as e:
        return {
            "ok": False,
            "error": f"openpyxl load failed: {e}",
            "metadata": metadata,
        }

    try:
        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            data_rows: list[tuple] = []
            for _ in range(20):
                r = next(rows_iter, None)
                if r is None:
                    break
                data_rows.append(r)

            sheet_info.append({
                "name": ws.title,
                "header": list(header) if header else [],
                "rows_preview": [list(r) for r in data_rows],
            })

            # 转 markdown 表格
            chunks.append(f"## Sheet: {ws.title}\n")
            if not header:
                chunks.append("(空表)\n")
                continue
            # 表头
            chunks.append("| " + " | ".join(_cell_str(c) for c in header) + " |")
            chunks.append("|" + "|".join(["---"] * len(header)) + "|")
            for r in data_rows:
                chunks.append("| " + " | ".join(_cell_str(c) for c in r) + " |")
            chunks.append("")
    finally:
        wb.close()

    full = "\n".join(chunks)
    return {
        "ok": True,
        "text": full[:TEXT_TRUNCATE],
        "full_text": full,
        "metadata": {**metadata, "sheets": [s["name"] for s in sheet_info]},
        "needs_ocr": False,
    }


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("\n", " ").replace("|", "\\|").strip()
    if len(s) > 60:
        s = s[:60] + "..."
    return s


# ----------------------------------------------------------------------
# 文本类
# ----------------------------------------------------------------------
def _parse_text(p: Path, metadata: dict[str, Any]) -> dict[str, Any]:
    try:
        text = p.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        return {
            "ok": False,
            "error": f"read_text failed: {e}",
            "metadata": metadata,
        }
    return {
        "ok": True,
        "text": text[:TEXT_TRUNCATE],
        "full_text": text,
        "metadata": metadata,
        "needs_ocr": False,
    }


# ----------------------------------------------------------------------
# 视觉模型兜底（扫描件）
# ----------------------------------------------------------------------
def extract_text_with_vision(file_path: str | Path, ai_client: Any) -> dict[str, Any]:
    """扫描件兜底：把 PDF 每页转图片，调 Qwen3-VL-Plus 提取文字。

    Returns:
        {"ok": True, "text": "...", "full_text": "...", "pages": N}
        {"ok": False, "error": "..."}
    """
    p = safe_path(file_path)
    if not p.exists() or not p.is_file():
        return {"ok": False, "error": "file_not_found"}

    if p.suffix.lower() != ".pdf":
        # 非暂 PDF（图片）也支持：直接把图片编码 base64
        return _vision_image(p, ai_client)

    # 1. PDF 每页 → PNG（用 pdfplumber 的 .to_image() 或 pdf2image）
    images_b64: list[str] = []
    try:
        import pdfplumber
    except ImportError:
        return {"ok": False, "error": "pdfplumber not installed"}

    try:
        with pdfplumber.open(str(p)) as pdf:
            for page in pdf.pages:
                # to_image 返回 PageImage，可 .original（PIL.Image）
                pi = page.to_image(resolution=150)
                img = pi.original
                buf = io.BytesIO()
                img.save(buf, format="PNG")
                b64 = base64.b64encode(buf.getvalue()).decode("ascii")
                images_b64.append(b64)
                if len(images_b64) >= 5:  # 限制 5 页，控成本
                    break
    except Exception as e:
        return {"ok": False, "error": f"pdf_to_image failed: {e}"}

    if not images_b64:
        return {"ok": False, "error": "no_pages_rendered"}

    # 2. 逐页调 vision 模型
    all_text: list[str] = []
    for idx, b64 in enumerate(images_b64):
        msg = {
            "role": "user",
            "content": [
                {
                    "type": "text",
                    "text": "请提取这张扫描件图片中的全部文字内容，按原文输出，不要总结。",
                },
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{b64}"},
                },
            ],
        }
        result = ai_client.chat(
            messages=[msg],
            model=_get_ocr_model(),
            temperature=0.1,
            item_id=f"ocr-{p.stem[:20]}-{idx}",
            action="vision_ocr",
        )
        if result.get("ok"):
            all_text.append(result.get("content", ""))
        else:
            all_text.append(f"[OCR 第{idx+1}页失败: {result.get('error')}]")

    full = "\n".join(all_text).strip()
    return {
        "ok": True,
        "text": full[:TEXT_TRUNCATE],
        "full_text": full,
        "pages": len(images_b64),
        "model": "qwen3-vl-plus",
    }


def _vision_image(p: Path, ai_client: Any) -> dict[str, Any]:
    """直接对图片文件调用 vision 模型。"""
    try:
        from PIL import Image
    except ImportError:
        return {"ok": False, "error": "pillow not installed"}

    try:
        img = Image.open(str(p))
        # 缩放到长边 ≤ 2000，避免太大
        img.thumbnail((2000, 2000))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception as e:
        return {"ok": False, "error": f"image_open failed: {e}"}

    msg = {
        "role": "user",
        "content": [
            {"type": "text", "text": "请提取这张图片中的全部文字内容，按原文输出。"},
            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
        ],
    }
    result = ai_client.chat(
        messages=[msg],
        model=_get_ocr_model(),
        temperature=0.1,
        item_id=f"ocr-{p.stem[:20]}",
        action="vision_ocr",
    )
    if not result.get("ok"):
        return result
    content = result.get("content", "")
    return {
        "ok": True,
        "text": content[:TEXT_TRUNCATE],
        "full_text": content,
        "model": "qwen3-vl-plus",
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        r = parse_file(sys.argv[1])
        print({k: v for k, v in r.items() if k != "full_text"})
