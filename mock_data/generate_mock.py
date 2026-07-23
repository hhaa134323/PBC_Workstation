# -*- coding: utf-8 -*-
"""
IPO 审计 PBC 智能管理工作站 · 模拟数据生成器
================================================
基于真实 PBC 需求清单（103 项），生成：
  1. _real_items.json        中间产物（解析的原始数据 + 扩展字段）
  2. 01_PBC_List.xlsx        统一清单（14 列，含条件格式 + 数据验证）
  3. 客户共享文件夹/{实体}/  精选 20 个"已提供"类模拟文件（PDF + Excel）

可重复执行：每次运行都会清空 mock_data/ 重新生成。
"""
import os
import re
import json
import shutil
import random
from datetime import datetime, date
from collections import Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)

# ============================================================
# 路径与常量
# ============================================================
ROOT = os.path.dirname(os.path.abspath(__file__))
REAL_XLSX = r"D:\AgentProjects\IpoPBC\测试资料\副本审计基本资料需求_2025.xlsx"
REAL_SHEET = "审计所需资料清单_2025"
HEADER_ROW = 5  # 表头所在行（数据从第 6 行开始）
DATA_START_ROW = 6

OUT_DIR = ROOT
OUT_LIST_XLSX = os.path.join(OUT_DIR, "01_PBC_List.xlsx")
OUT_JSON = os.path.join(OUT_DIR, "_real_items.json")
OUT_FILES_DIR = os.path.join(OUT_DIR, "客户共享文件夹")

# 实体归属目录名（"ABC集团" 的目录名为 "集团合并"）
ENTITY_DIR = {
    "ABC集团": "集团合并",
    "ABC科技": "ABC科技",
    "ABC制造": "ABC制造",
    "ABC商贸": "ABC商贸",
}

# 一级分类 → 实体归属（按思路 7.2 + 实体均衡）
# ABC集团：历史沿革、业务及财务概览、关联方、期后（集团级聚合）
# ABC科技：货币资金、薪酬、税务相关
# ABC制造：存货、长期资产、租赁类、费用(部分)
# ABC商贸：往来科目、短期借款、收入成本、成本、营业外收支、政府补助、其他、费用(部分)
CATEGORY_TO_ENTITY = {
    "历史沿革": "ABC集团",
    "业务及财务概览": "ABC集团",
    "关联方": "ABC集团",
    "期后": "ABC集团",
    "营业外收支": "ABC商贸",
    "政府补助": "ABC商贸",
    "货币资金": "ABC科技",
    "薪酬": "ABC科技",
    "税务相关": "ABC科技",
    "存货": "ABC制造",
    "长期资产": "ABC制造",
    "租赁类": "ABC制造",
    "往来科目": "ABC商贸",
    "短期借款": "ABC商贸",
    "收入成本": "ABC商贸",
    "成本": "ABC商贸",
    "其他流动资产、其他非流动资产、其他流动负债、其他非流动负债": "ABC商贸",
    "费用": "ABC商贸",  # 默认全部费用归 ABC商贸；下方按编号微调
}

# 14 列字段（严格顺序）
COLUMNS = [
    "资料编号", "一级分类", "相关科目", "问题/需求描述", "格式", "优先级",
    "提出时间", "期望提供日期", "逾期天数", "资料提供情况", "备注",
    "实体归属", "置信度", "文件路径",
]

# 7.1 14 列字段对应的"原始字段名 → 内部字段名"映射
# 用于从真实清单原始字段读取数据
RAW_TO_FIELD = {
    "资料编号": "资料编号",
    "相关科目": "相关科目",
    "问题/需求描述": "问题/需求描述",
    "格式": "格式",
    "优先级": "优先级",
    "提出时间": "提出时间",
    "期望提供日期": "期望提供日期",
    "逾期天数": "逾期天数",
    "资料提供情况": "资料提供情况",
    "备注": "备注",
}

# 真实清单的"资料编号"前缀 → 一级分类映射
# （从原始 Excel 的分类标题行推断，已和旧版 _real_items.json 比对一致）
PREFIX_TO_CATEGORY = {
    "历": "历史沿革",
    "概览": "业务及财务概览",
    "收": "收入成本",
    "成": "成本",
    "营业外": "营业外收支",
    "补助": "政府补助",
    "费": "费用",
    "银": "货币资金",
    "存": "存货",
    "往来": "往来科目",
    "长期资产": "长期资产",
    "薪": "薪酬",
    "关": "关联方",
    "税": "税务相关",
    "租": "租赁类",
    "借": "短期借款",
    "其他": "其他流动资产、其他非流动资产、其他流动负债、其他非流动负债",
    "期后": "期后",
}

# 实体均衡微调：将部分费用项从 ABC商贸 移到 ABC制造，使两实体都进入 [22, 34] 区间
# 经过测算：默认映射下，集团/制造各 20 项（略低于 22），商贸 34 项。
# 微调方案：营业外-1 + 补助-1 → 集团（这两类天然偏集团级）；费-2 + 费-4 → 制造（制造业自有运营费用）
ENTITY_OVERRIDE = {
    "营业外-1": "ABC集团",
    "补助-1": "ABC集团",
    "费-2": "ABC制造",
    "费-4": "ABC制造",
}

# 数据验证允许的状态值
ALLOWED_STATUSES = ["已提供", "已提供，审核中", "未提供", "不适用", "待定"]

# 精选生成模拟文件的 20 项（覆盖 4 实体 + PDF/Excel 多类型）
# 每项：(资料编号, 资料名称, 文件类型[pdf|xlsx], 内容生成器 key)
MOCK_FILE_PLAN = [
    # ===== ABC集团 / 集团合并目录（5 个）=====
    ("历-1", "股权架构图", "pdf", "charter_structure"),
    ("历-3", "营业执照及工商登记文件", "pdf", "business_license"),
    ("历-4", "公司章程", "pdf", "company_charter"),
    ("概览-3", "集团合并财务报表", "xlsx", "consolidated_fs"),
    ("概览-4", "集团合并调整明细", "xlsx", "consolidation_adj"),
    # ===== ABC科技（5 个）=====
    ("银-1", "银行开户销户证明", "pdf", "bank_account_cert"),
    ("银-2", "银行对账单", "xlsx", "bank_statement"),
    ("薪-1", "工资明细表", "xlsx", "salary_detail"),
    ("薪-6", "股权激励计划", "pdf", "equity_incentive"),
    ("税-9", "递延所得税计算表", "xlsx", "deferred_tax"),
    # ===== ABC制造（6 个）=====
    ("存-1", "存货明细表", "xlsx", "inventory_detail"),
    ("存-4", "盘点计划", "pdf", "stocktake_plan"),
    ("存-5", "原材料保质期台账", "xlsx", "material_shelf_life"),
    ("长期资产-2", "固定资产产权证明", "pdf", "property_cert"),
    ("长期资产-4", "固定资产明细表", "xlsx", "fixed_assets"),
    ("租-2", "使用权资产和租赁负债计算表", "xlsx", "lease_calc"),
    # ===== ABC商贸（4 个）=====
    ("往来-1", "应收账款明细表及账龄", "xlsx", "ar_aging"),
    ("往来-4", "应付账款明细表及账龄", "xlsx", "ap_aging"),
    ("借-1", "银行借款合同", "pdf", "loan_contract"),
    ("借-2", "借款台账及利息计算表", "xlsx", "loan_ledger"),
]

# 可重复随机数
RNG = random.Random(42)

# 默认报告期间
PERIOD = "2025年12月31日"


# ============================================================
# 1. 解析真实清单
# ============================================================
def parse_real_list():
    """读取真实 PBC 清单 Excel，提取 103 条目，返回 list of dict。"""
    wb = openpyxl.load_workbook(REAL_XLSX, data_only=True)
    ws = wb[REAL_SHEET]

    items = []
    data_row_pattern = re.compile(r"^[\u4e00-\u9fa5A-Za-z]+-\d+$")

    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        rid = str(row[0]).strip()
        # 只保留"前缀-数字"格式的资料编号行，跳过分类标题行
        if not data_row_pattern.match(rid):
            continue

        # 推断一级分类（前缀匹配；多字符前缀优先匹配，如"长期资产"优先于"长"）
        prefix = None
        for p in sorted(PREFIX_TO_CATEGORY.keys(), key=len, reverse=True):
            if rid.startswith(p + "-") or rid.startswith(p):
                # 严格匹配 prefix-数字 的前缀部分
                if rid.split("-")[0] == p:
                    prefix = p
                    break
        category = PREFIX_TO_CATEGORY.get(prefix, "其他")

        item = {
            "资料编号": rid,
            "一级分类": category,
            "相关科目": _clean(row[1]),
            "问题/需求描述": _clean(row[2]),
            "格式": _clean(row[3]),
            "优先级": _clean(row[4]),
            "提出时间": _serialize_dt(row[5]),
            "期望提供日期": _serialize_dt(row[6]),
            "逾期天数": _clean(row[7]),
            "资料提供情况": _clean(row[8]),
            "备注": _clean(row[9]),
        }
        items.append(item)

    wb.close()
    return items


def _clean(val):
    """清理单元格值：None → ""，其他转 str 去首尾空白。"""
    if val is None:
        return ""
    s = str(val).strip()
    return s


def _serialize_dt(val):
    """datetime/date → 'YYYY-MM-DD' 字符串；其他 → str。"""
    if val is None or val == "":
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


# ============================================================
# 2. 整合 14 列统一格式
# ============================================================
def extend_items(items):
    """为每条目补齐：一级分类、实体归属、置信度、文件路径。"""
    for item in items:
        rid = item["资料编号"]

        # 实体归属（先按分类，再按 override 微调）
        if rid in ENTITY_OVERRIDE:
            entity = ENTITY_OVERRIDE[rid]
        else:
            entity = CATEGORY_TO_ENTITY.get(item["一级分类"], "ABC商贸")
        item["实体归属"] = entity

        # 置信度：已提供/审核中类 0.85-0.95，其他 0.70-0.85
        status = item["资料提供情况"]
        if "已提供" in status:
            conf = round(RNG.uniform(0.85, 0.95), 2)
        else:
            conf = round(RNG.uniform(0.70, 0.85), 2)
        item["置信度"] = conf

        # 文件路径初始留空（后续回填）
        item["文件路径"] = ""

    # 按资料编号排序（让"历-1, 历-2..."保持自然顺序）
    items.sort(key=lambda x: _sort_key(x["资料编号"]))
    return items


def _sort_key(rid):
    """生成可排序的 tuple：('历', 1)。多字符前缀也能正确排序。"""
    parts = rid.split("-", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return (parts[0], int(parts[1]))
    return (rid, 0)


# ============================================================
# 3. 生成 01_PBC_List.xlsx
# ============================================================
def write_pbc_list(items, file_path=None, with_paths=True):
    """生成统一清单 Excel。14 列、表头加粗、列宽自适应、条件格式 + 数据验证。"""
    if file_path is None:
        file_path = OUT_LIST_XLSX

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PBC清单"

    # --- 表头 ---
    header_font = Font(bold=True, size=11, color="000000")
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # --- 数据 ---
    cell_align = Alignment(vertical="top", wrap_text=True)
    for r_idx, item in enumerate(items, start=2):
        for c_idx, col_name in enumerate(COLUMNS, start=1):
            val = item.get(col_name, "")
            # 文件路径仅在 with_paths=True 时回填
            if col_name == "文件路径" and not with_paths:
                val = ""
            # 置信度转 float
            if col_name == "置信度" and val != "":
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    pass
            # 逾期天数转 int
            if col_name == "逾期天数" and val != "":
                try:
                    val = int(val)
                except (TypeError, ValueError):
                    pass
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = cell_align
            cell.border = border

    # --- 列宽自适应 ---
    col_widths = {
        "资料编号": 10, "一级分类": 18, "相关科目": 14, "问题/需求描述": 60,
        "格式": 8, "优先级": 8, "提出时间": 12, "期望提供日期": 14,
        "逾期天数": 10, "资料提供情况": 18, "备注": 24,
        "实体归属": 10, "置信度": 10, "文件路径": 38,
    }
    for c_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = col_widths.get(col_name, 14)

    # 冻结首行 + 资料编号列
    ws.freeze_panes = "B2"

    # --- 条件格式（应用在"资料提供情况"列，思路 6.1 四状态机）---
    status_col_letter = get_column_letter(COLUMNS.index("资料提供情况") + 1)
    n_rows = len(items)
    status_range = f"{status_col_letter}2:{status_col_letter}{n_rows + 1}"

    # 未提供 → 红色填充（FFC7CE 字体 9C0006）
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("未提供",{status_col_letter}2))'],
            fill=red_fill, font=red_font, stopIfTrue=False
        )
    )

    # 审核中 → 深绿填充（C6EFCE 字体 006100）
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("审核中",{status_col_letter}2))'],
            fill=green_fill, font=green_font, stopIfTrue=False
        )
    )

    # 已提供（精确等于）→ 浅绿填充（E2EFDA 字体 375623）
    light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    light_green_font = Font(color="375623")
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'EXACT({status_col_letter}2,"已提供")'],
            fill=light_green_fill, font=light_green_font, stopIfTrue=False
        )
    )

    # 不适用 → 灰色填充（D9D9D9 字体 808080）
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    gray_font = Font(color="808080")
    ws.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=[f'ISNUMBER(SEARCH("不适用",{status_col_letter}2))'],
            fill=gray_fill, font=gray_font, stopIfTrue=False
        )
    )

    # --- 数据验证：状态列允许的值（放宽验证以容纳真实数据）---
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(ALLOWED_STATUSES) + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="warning",  # 警告但不阻止输入（容纳"已提供纳税申报表"等历史值）
        errorTitle="状态值不在标准列表",
        error="建议使用标准状态值：已提供 / 已提供，审核中 / 未提供 / 不适用 / 待定。是否仍要继续？",
        promptTitle="状态选择",
        prompt="从下拉列表选择状态：已提供 / 已提供，审核中 / 未提供 / 不适用 / 待定",
    )
    dv.add(status_range)
    ws.add_data_validation(dv)

    # 行高
    ws.row_dimensions[1].height = 30

    wb.save(file_path)
    wb.close()
    return file_path


# ============================================================
# 4. 生成模拟文件
# ============================================================
def generate_mock_files(items):
    """为 MOCK_FILE_PLAN 中的项生成模拟文件，回填 items 中的"文件路径"。"""
    # 注册中文字体
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))

    # 建立资料编号 → item 的索引
    item_map = {it["资料编号"]: it for it in items}

    generated = []
    for rid, name, ext, gen_key in MOCK_FILE_PLAN:
        if rid not in item_map:
            print(f"  [WARN] 资料编号 {rid} 未在清单中找到，跳过")
            continue

        item = item_map[rid]
        entity = item["实体归属"]
        entity_dir = ENTITY_DIR.get(entity, entity)

        # 安全文件名（替换非法字符）
        safe_name = re.sub(r'[\\/:*?"<>|]', "_", name)
        filename = f"{rid}_{safe_name}.{ext}"
        rel_path = f"客户共享文件夹/{entity_dir}/{filename}"
        abs_path = os.path.join(OUT_FILES_DIR, entity_dir, filename)

        # 生成文件
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
        if ext == "pdf":
            _gen_pdf(abs_path, item, name, gen_key)
        elif ext == "xlsx":
            _gen_xlsx(abs_path, item, name, gen_key)

        # 回填文件路径（相对路径，从"客户共享文件夹/"开始）
        item["文件路径"] = rel_path
        generated.append((rid, name, ext, rel_path))
        print(f"  [OK] {rid:12s} {entity:6s} → {rel_path}")

    return generated


# ---------- PDF 生成（reportlab + STSong-Light）----------
def _gen_pdf(path, item, doc_name, gen_key):
    """生成 PDF 模拟扫描件。内容包含资料编号、公司名、期间、关键词等。"""
    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=f"{item['资料编号']}_{doc_name}",
        author="ABC集团",
    )

    styles = getSampleStyleSheet()
    # 自定义中文样式（用 STSong-Light CID 字体）
    title_style = ParagraphStyle(
        "CN_Title", parent=styles["Title"],
        fontName="STSong-Light", fontSize=20, leading=28,
        alignment=TA_CENTER, spaceAfter=14,
    )
    h2_style = ParagraphStyle(
        "CN_H2", parent=styles["Heading2"],
        fontName="STSong-Light", fontSize=14, leading=20,
        spaceBefore=10, spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "CN_Body", parent=styles["BodyText"],
        fontName="STSong-Light", fontSize=11, leading=18,
        alignment=TA_LEFT, spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        "CN_Meta", parent=body_style,
        fontSize=10, leading=16, textColor=colors.HexColor("#555555"),
    )

    story = []

    # ---- 标题与元信息 ----
    story.append(Paragraph(f"{doc_name}", title_style))
    story.append(Spacer(1, 0.5 * cm))
    story.append(Paragraph(
        f"<b>资料编号：</b>{item['资料编号']}　　"
        f"<b>实体归属：</b>{item['实体归属']}<br/>"
        f"<b>报告期间：</b>{PERIOD}　　"
        f"<b>相关科目：</b>{item['相关科目'] or '—'}<br/>"
        f"<b>格式要求：</b>{item['格式'] or '—'}　　"
        f"<b>资料状态：</b>{item['资料提供情况']}",
        meta_style
    ))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(
        f"<b>需求描述：</b>{item['问题/需求描述']}", body_style
    ))
    story.append(Spacer(1, 0.4 * cm))

    # ---- 调用各内容生成器 ----
    content_fn = PDF_CONTENT_GENERATORS.get(gen_key, _pdf_default_content)
    content_fn(story, item, doc_name, h2_style, body_style)

    # ---- 页脚说明（防伪标记，AI 可识别）----
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph(
        "—— 本文件为 IPO 审计 PBC 智能管理工作站模拟数据，"
        "所有公司名、账户号、人名均为虚构，仅供产品测试演示使用，"
        "不涉及任何真实客户数据。 ——",
        ParagraphStyle(
            "CN_Foot", parent=body_style,
            fontSize=8, textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER, leading=12,
        )
    ))

    doc.build(story)


def _pdf_table(story, headers, rows, h2_style, title):
    """通用 PDF 表格渲染。"""
    story.append(Paragraph(title, h2_style))
    data = [headers] + rows
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl)


def _pdf_default_content(story, item, doc_name, h2_style, body_style):
    """默认 PDF 内容（无特定生成器时使用）。"""
    story.append(Paragraph("一、文件说明", h2_style))
    story.append(Paragraph(
        f"本文件为{item['实体归属']}就审计资料需求 "
        f"\"{item['资料编号']} - {doc_name}\" 提供的扫描件。"
        f"文件覆盖期间为 {PERIOD}。", body_style
    ))
    story.append(Paragraph("二、文件内容", h2_style))
    story.append(Paragraph(
        "（此处为模拟的文件正文内容。实际审计过程中，本文件应包含对应资料的全部正文、"
        "签字盖章页及附件清单。）", body_style
    ))


def _pdf_charter_structure(story, item, doc_name, h2_style, body_style):
    """历-1 股权架构图 PDF 内容。"""
    story.append(Paragraph("一、ABC集团股权架构概览", h2_style))
    story.append(Paragraph(
        "ABC集团（虚拟模拟主体）截至 2025年12月31日 的股权架构如下。"
        "集团控股 3 家全资/控股子公司：ABC科技、ABC制造、ABC商贸。", body_style
    ))
    _pdf_table(story,
        ["股东层级", "股东名称", "出资比例", "出资额(万元)", "出资时间"],
        [
            ["第一层", "ABC集团控股有限公司", "100.00%", "50000.00", "2010-06-15"],
            ["第二层", "ABC科技有限公司", "100.00%", "20000.00", "2012-03-20"],
            ["第二层", "ABC制造有限公司", "100.00%", "20000.00", "2012-08-10"],
            ["第二层", "ABC商贸有限公司", "100.00%", "10000.00", "2014-05-18"],
        ],
        h2_style, "二、股东出资明细")
    story.append(Paragraph("三、历次股权变更记录", h2_style))
    story.append(Paragraph(
        "自设立以来，ABC集团股权结构共经历 2 次变更："
        "（1）2014 年 5 月新增设立 ABC商贸有限公司作为全资子公司；"
        "（2）2020 年 8 月 ABC科技完成 B 轮增资 5000 万元（由集团独认缴）。"
        "其他股权结构自设立以来未发生变更。", body_style
    ))
    story.append(Paragraph("四、投资协议", h2_style))
    story.append(Paragraph(
        "相关投资协议（含历次增资协议、股东协议）原件已归档备查，"
        "本扫描件目录页仅作索引，详见附件清单。", body_style
    ))


def _pdf_business_license(story, item, doc_name, h2_style, body_style):
    """历-3 营业执照及工商登记文件 PDF 内容。"""
    story.append(Paragraph("一、营业执照清单", h2_style))
    _pdf_table(story,
        ["公司名称", "统一社会信用代码", "成立日期", "注册资本(万元)", "法定代表人"],
        [
            ["ABC集团控股有限公司", "91440100MA01ABC001", "2010-06-15", "50000", "张某甲"],
            ["ABC科技有限公司", "91440100MA01ABC002", "2012-03-20", "20000", "李某乙"],
            ["ABC制造有限公司", "91440100MA01ABC003", "2012-08-10", "20000", "王某丙"],
            ["ABC商贸有限公司", "91440100MA01ABC004", "2014-05-18", "10000", "赵某丁"],
        ],
        h2_style, "二、各公司营业执照信息")
    story.append(Paragraph("三、工商登记变更记录", h2_style))
    story.append(Paragraph(
        "自设立以来，各公司工商登记变更情况如下："
        "（1）2018 年 6 月，ABC科技经营范围新增\"软件开发\"；"
        "（2）2020 年 8 月，ABC集团注册资本由 30000 万元增至 50000 万元；"
        "（3）2022 年 10 月，ABC制造法定代表人由陈某变更为王某丙。"
        "上述变更均已取得换发营业执照，原件备查。", body_style
    ))
    story.append(Paragraph("四、附件清单", h2_style))
    story.append(Paragraph(
        "1. ABC集团营业执照扫描件<br/>"
        "2. ABC科技营业执照扫描件<br/>"
        "3. ABC制造营业执照扫描件<br/>"
        "4. ABC商贸营业执照扫描件<br/>"
        "5. 工商变更登记申请书（共 3 份）", body_style
    ))


def _pdf_company_charter(story, item, doc_name, h2_style, body_style):
    """历-4 公司章程 PDF 内容。"""
    story.append(Paragraph("ABC集团公司章程（节选）", h2_style))
    story.append(Paragraph("第一章 总则", h2_style))
    story.append(Paragraph(
        "第一条 为规范 ABC集团控股有限公司（以下简称\"公司\"）的组织和行为，"
        "保护公司、股东和债权人的合法权益，根据《中华人民共和国公司法》"
        "（以下简称\"《公司法》\"）及其他有关法律、行政法规的规定，制定本章程。", body_style
    ))
    story.append(Paragraph(
        "第二条 公司名称：ABC集团控股有限公司。公司住所：广州市天河区××路 × 号。"
        "统一社会信用代码：91440100MA01ABC001。", body_style
    ))
    story.append(Paragraph("第二章 经营范围与注册资本", h2_style))
    story.append(Paragraph(
        "第三条 公司经营范围：以自有资金从事投资活动；企业管理咨询服务；"
        "信息技术咨询服务。（依法须经批准的项目，经相关部门批准后方可开展经营活动）", body_style
    ))
    story.append(Paragraph(
        "第四条 公司注册资本为人民币 50000 万元，全部为货币出资。", body_style
    ))
    story.append(Paragraph("第三章 股东与股东会", h2_style))
    story.append(Paragraph(
        "第五条 公司股东为发起人张某甲，出资 50000 万元，占注册资本 100%。", body_style
    ))
    story.append(Paragraph("第四章 董事会与经理", h2_style))
    story.append(Paragraph(
        "第六条 公司设董事会，成员 5 人，由股东会选举产生。"
        "董事任期 3 年，任期届满，可连选连任。", body_style
    ))
    story.append(Paragraph("第五章 附则", h2_style))
    story.append(Paragraph(
        "本章程自股东签字之日起生效。本章程修正案另附。", body_style
    ))
    story.append(Paragraph(
        "（模拟签字盖章页）股东签字：张某甲　　日期：2010-06-15", body_style
    ))


def _pdf_bank_account_cert(story, item, doc_name, h2_style, body_style):
    """银-1 银行开户销户证明 PDF 内容。"""
    story.append(Paragraph("一、开户销户证明说明", h2_style))
    story.append(Paragraph(
        f"截至 {PERIOD}，{item['实体归属']} 在各银行开立的银行账户清单如下。"
        "本证明由各开户银行出具并盖章确认。", body_style
    ))
    _pdf_table(story,
        ["开户银行", "账户类型", "银行账号", "开户日期", "状态"],
        [
            ["中国工商银行广州天河支行", "基本户", "4400123401001234567", "2012-04-10", "正常"],
            ["中国建设银行广州珠江支行", "一般户", "4400156701002345678", "2013-06-20", "正常"],
            ["招商银行广州分行营业部", "一般户", "1209087654321098", "2015-09-15", "正常"],
            ["中国农业银行广州分行", "一般户", "4400345601003456789", "2016-03-08", "正常"],
            ["中国工商银行广州天河支行", "保证金户", "4400123401004567890", "2019-11-25", "正常"],
            ["广发银行广州分行营业部", "一般户", "9550880123456789", "2018-07-12", "已销户"],
            ["中信银行广州分行", "一般户", "8112001012345678", "2020-05-30", "已销户"],
        ],
        h2_style, "二、银行账户清单")
    story.append(Paragraph("三、销户说明", h2_style))
    story.append(Paragraph(
        "广发银行账户于 2021-08-10 销户，原因为业务调整；"
        "中信银行账户于 2022-04-20 销户，原因为账户整合优化。"
        "其余账户均处于正常使用状态。", body_style
    ))


def _pdf_equity_incentive(story, item, doc_name, h2_style, body_style):
    """薪-6 股权激励计划 PDF 内容。"""
    story.append(Paragraph("ABC科技股权激励计划（节选）", h2_style))
    story.append(Paragraph("第一章 总则", h2_style))
    story.append(Paragraph(
        "第一条 为充分调动公司中高层管理人员及核心骨干的积极性，"
        "完善公司激励机制，根据《中华人民共和国公司法》"
        "及公司章程的有关规定，制定本股权激励计划。", body_style
    ))
    story.append(Paragraph("第二章 激励对象与方式", h2_style))
    story.append(Paragraph(
        "第二条 本计划激励对象共 30 人，包括公司高级管理人员 5 人、"
        "核心技术人员 15 人、中层管理人员 10 人。具体名单见附件。", body_style
    ))
    story.append(Paragraph(
        "第三条 激励方式：限制性股票。计划授予股票总数 1000 万股，"
        "占公司总股本的 5%。授予价格 8.50 元/股。", body_style
    ))
    story.append(Paragraph("第三章 解锁条件", h2_style))
    story.append(Paragraph(
        "第四条 限制性股票分三期解锁，每期解锁 1/3，解锁条件如下："
        "（1）第一个解锁期：2023 年度扣非净利润不低于 6000 万元；"
        "（2）第二个解锁期：2024 年度扣非净利润不低于 7500 万元；"
        "（3）第三个解锁期：2025 年度扣非净利润不低于 9000 万元。", body_style
    ))
    story.append(Paragraph("第四章 实施情况", h2_style))
    story.append(Paragraph(
        f"截至 {PERIOD}，第一期、第二期均已满足解锁条件并完成解锁。"
        "第三期（2025 年度）业绩条件预计可达成，待年度审计完成后统一办理解锁手续。", body_style
    ))


def _pdf_stocktake_plan(story, item, doc_name, h2_style, body_style):
    """存-4 盘点计划 PDF 内容。"""
    story.append(Paragraph(f"{item['实体归属']}存货盘点计划", h2_style))
    story.append(Paragraph("一、盘点基准日", h2_style))
    story.append(Paragraph(
        f"本次存货盘点基准日为 {PERIOD}。盘点实施期间：2025年12月28日至2025年12月31日。", body_style
    ))
    story.append(Paragraph("二、盘点范围", h2_style))
    story.append(Paragraph(
        "盘点范围包括 ABC制造全部自有仓库（共 3 个）及在途存货。"
        "具体仓库：原材料仓（A 仓）、半成品仓（B 仓）、库存商品仓（C 仓）。", body_style
    ))
    story.append(Paragraph("三、盘点人员及分工", h2_style))
    _pdf_table(story,
        ["姓名", "所属部门", "盘点区域", "角色"],
        [
            ["陈某 A", "财务部", "A 仓", "盘点员"],
            ["林某 B", "财务部", "B 仓", "盘点员"],
            ["吴某 C", "仓储部", "A 仓", "复盘员"],
            ["黄某 D", "仓储部", "B 仓", "复盘员"],
            ["许某 E", "生产部", "C 仓", "盘点员"],
            ["郑某 F", "内审部", "全部", "监盘员"],
        ],
        h2_style, "四、盘点人员明细")
    story.append(Paragraph("五、盘点程序", h2_style))
    story.append(Paragraph(
        "1. 盘点前：停止所有存货收发业务，整理仓库使存货分类摆放整齐；<br/>"
        "2. 盘点中：采用\"双人双盲\"方式盘点，盘点员与复盘员独立清点并记录；<br/>"
        "3. 盘点后：汇总差异，分析差异原因（计量误差/呆滞/盘盈盘亏），编制盘点小结；<br/>"
        "4. 监盘：内审部全程监盘，会计师事务所审计员独立抽盘。", body_style
    ))
    story.append(Paragraph("六、盘点小结", h2_style))
    story.append(Paragraph(
        "本次盘点共盘点 SKU 1247 项，账面存货金额合计 12348.65 万元。"
        "盘点差异 8 项（盘亏 5 项、盘盈 3 项），差异金额合计 -2.34 万元，"
        "差异率 0.019%，在可接受范围内。差异原因均为计量误差，已按规定调整。", body_style
    ))


def _pdf_property_cert(story, item, doc_name, h2_style, body_style):
    """长期资产-2 固定资产产权证明 PDF 内容。"""
    story.append(Paragraph(f"{item['实体归属']}固定资产产权证明清单", h2_style))
    story.append(Paragraph(
        f"截至 {PERIOD}，{item['实体归属']}持有的房屋、土地、车辆等产权证明扫描件如下：",
        body_style
    ))
    _pdf_table(story,
        ["资产类别", "证书名称", "证书编号", "登记日期", "所有权人"],
        [
            ["房屋", "不动产权证书", "粤(2018)广州市不动产权第0012345号", "2018-05-10", "ABC制造有限公司"],
            ["房屋", "不动产权证书", "粤(2020)广州市不动产权第0056789号", "2020-09-22", "ABC制造有限公司"],
            ["土地", "国有土地使用证", "穗国用(2013)字第A12345号", "2013-04-15", "ABC制造有限公司"],
            ["车辆", "机动车行驶证", "粤A·12345", "2019-06-18", "ABC制造有限公司"],
            ["车辆", "机动车行驶证", "粤A·67890", "2021-11-30", "ABC制造有限公司"],
            ["设备", "进口设备报关单", "海关单号 123456789012345678", "2022-02-14", "ABC制造有限公司"],
        ],
        h2_style, "一、产权证明清单")
    story.append(Paragraph("二、说明", h2_style))
    story.append(Paragraph(
        "上述产权证明均已通过相关主管部门核发，权属清晰，无抵押、查封等权利限制（抵押情况详见借款合同清单）。"
        "原件存放于公司行政部档案室，本扫描件仅供审计查阅。", body_style
    ))


def _pdf_loan_contract(story, item, doc_name, h2_style, body_style):
    """借-1 银行借款合同 PDF 内容。"""
    story.append(Paragraph(f"{item['实体归属']}银行借款合同汇总", h2_style))
    story.append(Paragraph(
        f"截至 {PERIOD}，{item['实体归属']}尚未结清的银行借款合同清单如下：",
        body_style
    ))
    _pdf_table(story,
        ["合同编号", "贷款银行", "借款金额(万元)", "借款起止日", "年利率"],
        [
            ["RL-2023-001", "中国工商银行广州天河支行", "3000.00", "2023-06-30 至 2026-06-29", "4.35%"],
            ["RL-2024-005", "中国建设银行广州珠江支行", "5000.00", "2024-03-15 至 2027-03-14", "4.20%"],
            ["RL-2024-008", "招商银行广州分行营业部", "2000.00", "2024-09-10 至 2025-09-09", "4.10%"],
            ["RL-2025-002", "中国农业银行广州分行", "4500.00", "2025-02-20 至 2028-02-19", "3.85%"],
        ],
        h2_style, "一、借款合同清单")
    story.append(Paragraph("二、合同主要条款（以 RL-2024-005 为例）", h2_style))
    story.append(Paragraph(
        "<b>第一条 借款金额：</b>人民币 5000 万元整。<br/>"
        "<b>第二条 借款用途：</b>用于补充公司流动资金，不得挪作他用。<br/>"
        "<b>第三条 借款期限：</b>3 年，自 2024-03-15 起至 2027-03-14 止。<br/>"
        "<b>第四条 借款利率：</b>固定年利率 4.20%，按季结息，结息日为每季末月 20 日。<br/>"
        "<b>第五条 担保方式：</b>由 ABC集团控股有限公司提供连带责任保证担保。<br/>"
        "<b>第六条 还款方式：</b>按季付息，到期一次还本。<br/>"
        "<b>第七条 提前还款：</b>借款人可提前 30 日书面通知贷款人后提前还款，无违约金。", body_style
    ))
    story.append(Paragraph("三、抵押/质押/担保情况", h2_style))
    story.append(Paragraph(
        "上述 4 笔银行借款均由 ABC集团控股有限公司提供连带责任保证担保，无抵押/质押。"
        "担保合同编号及原件详见附件。", body_style
    ))


# PDF 内容生成器注册表
PDF_CONTENT_GENERATORS = {
    "charter_structure": _pdf_charter_structure,
    "business_license": _pdf_business_license,
    "company_charter": _pdf_company_charter,
    "bank_account_cert": _pdf_bank_account_cert,
    "equity_incentive": _pdf_equity_incentive,
    "stocktake_plan": _pdf_stocktake_plan,
    "property_cert": _pdf_property_cert,
    "loan_contract": _pdf_loan_contract,
}


# ---------- Excel 生成（openpyxl）----------
def _gen_xlsx(path, item, doc_name, gen_key):
    """生成 Excel 模拟数据文件。"""
    wb = openpyxl.Workbook()
    # 默认 Sheet 用作封面/说明
    cover = wb.active
    cover.title = "封面"
    _write_cover(cover, item, doc_name)

    # 调用各内容生成器
    content_fn = XLSX_CONTENT_GENERATORS.get(gen_key, _xlsx_default_content)
    content_fn(wb, item, doc_name)

    wb.save(path)
    wb.close()


def _write_cover(ws, item, doc_name):
    """写封面 Sheet：元信息。"""
    ws["A1"] = doc_name
    ws["A1"].font = Font(bold=True, size=16, name="等线")
    ws["A2"] = f"资料编号：{item['资料编号']}"
    ws["A3"] = f"实体归属：{item['实体归属']}"
    ws["A4"] = f"相关科目：{item['相关科目'] or '—'}"
    ws["A5"] = f"报告期间：{PERIOD}"
    ws["A6"] = f"资料状态：{item['资料提供情况']}"
    ws["A7"] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A9"] = "需求描述："
    ws["A10"] = item["问题/需求描述"]
    ws["A10"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A11"] = "—— 本文件为 IPO 审计 PBC 智能管理工作站模拟数据，所有公司名/账户号/人名均为虚构，仅供产品测试演示使用。"
    ws["A11"].font = Font(italic=True, size=9, color="888888")
    ws.column_dimensions["A"].width = 100


def _xlsx_default_content(wb, item, doc_name):
    """默认 Excel 内容：1 个 Sheet 模拟数据。"""
    ws = wb.create_sheet("明细")
    headers = ["序号", "项目", "金额(元)", "备注"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="D9E1F2")
    for i in range(1, 11):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=f"项目 {i}")
        ws.cell(row=i + 1, column=3, value=round(RNG.uniform(1000, 100000), 2))
        ws.cell(row=i + 1, column=4, value="")
    for col in "ABCD":
        ws.column_dimensions[col].width = 20


def _style_header_row(ws, row=1, n_cols=None):
    """美化表头行。"""
    if n_cols is None:
        n_cols = ws.max_column
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="4472C4")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 24


def _autofit(ws, widths):
    """按 widths dict 设列宽。"""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _xlsx_consolidated_fs(wb, item, doc_name):
    """概览-3 集团合并财务报表：3 个 Sheet（资产负债表/利润表/现金流量表）。"""
    # 资产负债表
    ws = wb.create_sheet("合并资产负债表")
    ws["A1"] = "ABC集团合并资产负债表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：ABC集团控股有限公司及子公司　　截止日：{PERIOD}　　单位：元"
    headers = ["报表项目", "期末余额", "期初余额", "变动(%)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=4)
    bs_rows = [
        ("流动资产：", "", "", ""),
        ("  货币资金", 385426731.22, 302184532.10, 27.52),
        ("  应收账款", 156832408.55, 132580310.20, 18.29),
        ("  存货", 123486510.00, 108902400.00, 13.39),
        ("  其他流动资产", 45632180.00, 38210500.00, 19.42),
        ("流动资产合计", 711377829.77, 581877742.30, 22.26),
        ("非流动资产：", "", "", ""),
        ("  固定资产", 268540200.00, 285310000.00, -5.88),
        ("  在建工程", 32180500.00, 18520000.00, 73.71),
        ("  无形资产", 45200000.00, 48000000.00, -5.83),
        ("  递延所得税资产", 8250000.00, 6180000.00, 33.50),
        ("非流动资产合计", 354170700.00, 358010000.00, -1.07),
        ("资产总计", 1065548529.77, 939887742.30, 13.37),
        ("流动负债：", "", "", ""),
        ("  短期借款", 145000000.00, 100000000.00, 45.00),
        ("  应付账款", 92840500.00, 85300000.00, 8.84),
        ("  应付职工薪酬", 18540000.00, 16200000.00, 14.44),
        ("  应交税费", 12380000.00, 9840000.00, 25.81),
        ("流动负债合计", 268760500.00, 211340000.00, 27.17),
        ("非流动负债：", "", "", ""),
        ("  长期借款", 95000000.00, 50000000.00, 90.00),
        ("  递延收益", 6250000.00, 5000000.00, 25.00),
        ("非流动负债合计", 101250000.00, 55000000.00, 84.09),
        ("负债合计", 370010500.00, 266340000.00, 38.93),
        ("所有者权益：", "", "", ""),
        ("  实收资本", 500000000.00, 500000000.00, 0.00),
        ("  资本公积", 85000000.00, 85000000.00, 0.00),
        ("  盈余公积", 42180000.00, 38540000.00, 9.44),
        ("  未分配利润", 68358029.77, 50007742.30, 36.70),
        ("所有者权益合计", 695538029.77, 673547742.30, 3.27),
        ("负债和所有者权益总计", 1065548529.77, 939887742.30, 13.37),
    ]
    for r, row in enumerate(bs_rows, start=5):
        for c, val in enumerate(row, start=1):
            if isinstance(val, (int, float)) and c in (2, 3):
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            elif isinstance(val, (int, float)) and c == 4:
                ws.cell(row=r, column=c, value=val / 100).number_format = '0.00%'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 32, "B": 22, "C": 22, "D": 12})

    # 利润表
    ws2 = wb.create_sheet("合并利润表")
    ws2["A1"] = "ABC集团合并利润表"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = f"编制单位：ABC集团控股有限公司及子公司　　2025年度　　单位：元"
    for c, h in enumerate(["报表项目", "本期金额", "上期金额", "变动(%)"], start=1):
        ws2.cell(row=4, column=c, value=h)
    _style_header_row(ws2, row=4, n_cols=4)
    is_rows = [
        ("一、营业收入", 985420000.00, 876580000.00, 12.41),
        ("  减：营业成本", 712300000.00, 638420000.00, 11.57),
        ("    税金及附加", 8520000.00, 7840000.00, 8.67),
        ("    销售费用", 45180000.00, 42310000.00, 6.78),
        ("    管理费用", 38650000.00, 35280000.00, 9.55),
        ("    研发费用", 28540000.00, 23180000.00, 23.13),
        ("    财务费用", 12840000.00, 8520000.00, 50.70),
        ("二、营业利润", 139390000.00, 121030000.00, 15.17),
        ("  加：营业外收入", 3250000.00, 2180000.00, 49.08),
        ("  减：营业外支出", 1580000.00, 1420000.00, 11.27),
        ("三、利润总额", 141060000.00, 121790000.00, 15.83),
        ("  减：所得税费用", 21159000.00, 18268500.00, 15.82),
        ("四、净利润", 119901000.00, 103521500.00, 15.82),
        ("  归属于母公司股东的净利润", 119901000.00, 103521500.00, 15.82),
        ("五、其他综合收益的税后净额", 0.00, 0.00, 0.00),
        ("六、综合收益总额", 119901000.00, 103521500.00, 15.82),
    ]
    for r, row in enumerate(is_rows, start=5):
        for c, val in enumerate(row, start=1):
            if isinstance(val, (int, float)) and c in (2, 3):
                ws2.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            elif isinstance(val, (int, float)) and c == 4:
                ws2.cell(row=r, column=c, value=val / 100).number_format = '0.00%'
            else:
                ws2.cell(row=r, column=c, value=val)
    _autofit(ws2, {"A": 38, "B": 22, "C": 22, "D": 12})

    # 现金流量表
    ws3 = wb.create_sheet("合并现金流量表")
    ws3["A1"] = "ABC集团合并现金流量表"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A2"] = f"编制单位：ABC集团控股有限公司及子公司　　2025年度　　单位：元"
    for c, h in enumerate(["报表项目", "本期金额", "上期金额"], start=1):
        ws3.cell(row=4, column=c, value=h)
    _style_header_row(ws3, row=4, n_cols=3)
    cf_rows = [
        ("一、经营活动产生的现金流量", "", ""),
        ("  销售商品、提供劳务收到的现金", 1135230000.00, 1010150000.00),
        ("  收到的税费返还", 8250000.00, 6840000.00),
        ("  收到其他与经营活动有关的现金", 12840000.00, 9580000.00),
        ("  经营活动现金流入小计", 1156320000.00, 1026570000.00),
        ("  购买商品、接受劳务支付的现金", 820150000.00, 738290000.00),
        ("  支付给职工以及为职工支付的现金", 152840000.00, 138520000.00),
        ("  支付的各项税费", 95180000.00, 84310000.00),
        ("  支付其他与经营活动有关的现金", 28540000.00, 25180000.00),
        ("  经营活动现金流出小计", 1096710000.00, 986300000.00),
        ("  经营活动产生的现金流量净额", 59610000.00, 40270000.00),
        ("二、投资活动产生的现金流量", "", ""),
        ("  收回投资收到的现金", 5000000.00, 3200000.00),
        ("  购建固定资产支付的现金", 48540000.00, 32180000.00),
        ("  投资活动现金流出小计", 48540000.00, 32180000.00),
        ("  投资活动产生的现金流量净额", -43540000.00, -28980000.00),
        ("三、筹资活动产生的现金流量", "", ""),
        ("  取得借款收到的现金", 195000000.00, 100000000.00),
        ("  偿还债务支付的现金", 105000000.00, 80000000.00),
        ("  分配股利支付的现金", 30000000.00, 25000000.00),
        ("  筹资活动现金流出小计", 135000000.00, 105000000.00),
        ("  筹资活动产生的现金流量净额", 60000000.00, -5000000.00),
        ("四、现金及现金等价物净增加额", 76070000.00, 6290000.00),
        ("  期初现金及现金等价物余额", 302184532.10, 295894532.10),
        ("  期末现金及现金等价物余额", 378254532.10, 302184532.10),
    ]
    for r, row in enumerate(cf_rows, start=5):
        for c, val in enumerate(row, start=1):
            if isinstance(val, (int, float)) and c in (2, 3):
                ws3.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws3.cell(row=r, column=c, value=val)
    _autofit(ws3, {"A": 38, "B": 22, "C": 22})


def _xlsx_consolidation_adj(wb, item, doc_name):
    """概览-4 集团合并调整明细。"""
    ws = wb.create_sheet("合并调整明细")
    ws["A1"] = "ABC集团合并调整明细表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：ABC集团控股有限公司　　{PERIOD}　　单位：元"
    headers = ["调整编号", "调整类型", "借/贷", "报表项目", "涉及实体", "调整金额", "调整原因"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))
    rows = [
        ("A001", "内部交易抵销", "借", "营业收入", "ABC科技→ABC商贸", 12500000.00, "集团内 ABC科技向 ABC商贸销售货物"),
        ("A001", "内部交易抵销", "贷", "营业成本", "ABC科技→ABC商贸", 10800000.00, "对应结转内部销售成本"),
        ("A002", "内部交易抵销", "借", "营业收入", "ABC制造→ABC科技", 8200000.00, "集团内 ABC制造向 ABC科技销售产品"),
        ("A002", "内部交易抵销", "贷", "营业成本", "ABC制造→ABC科技", 6900000.00, "对应结转内部销售成本"),
        ("A003", "内部往来抵销", "借", "应付账款", "ABC商贸", 12500000.00, "抵销 ABC商贸 对 ABC科技的应付"),
        ("A003", "内部往来抵销", "贷", "应收账款", "ABC科技", 12500000.00, "抵销 ABC科技 对 ABC商贸的应收"),
        ("A004", "内部往来抵销", "借", "其他应付款", "ABC商贸", 8200000.00, "抵销 ABC商贸 对 ABC制造的应付"),
        ("A004", "内部往来抵销", "贷", "其他应收款", "ABC制造", 8200000.00, "抵销 ABC制造 对 ABC商贸的应收"),
        ("A005", "存货未实现利润抵销", "借", "营业成本", "ABC商贸", 540000.00, "ABC商贸期末存货中含 ABC科技未实现利润"),
        ("A005", "存货未实现利润抵销", "贷", "存货", "ABC商贸", 540000.00, "对应存货跌价准备"),
        ("A006", "内部固定资产交易抵销", "借", "资产处置收益", "ABC制造", 800000.00, "ABC制造 向 ABC科技销售固定资产未实现利润"),
        ("A006", "内部固定资产交易抵销", "贷", "固定资产-原值", "ABC科技", 800000.00, "对应固定资产原值抵销"),
        ("A007", "内部固定资产折旧抵销", "借", "固定资产-累计折旧", "ABC科技", 80000.00, "对应本期已抵销固定资产多提折旧"),
        ("A007", "内部固定资产折旧抵销", "贷", "管理费用-折旧", "ABC科技", 80000.00, "对应本期折旧费用抵销"),
        ("A008", "内部利息抵销", "借", "财务费用-利息收入", "ABC集团", 1250000.00, "集团内 ABC集团 向 ABC商贸提供借款的利息"),
        ("A008", "内部利息抵销", "贷", "财务费用-利息支出", "ABC商贸", 1250000.00, "对应 ABC商贸支付的内部借款利息"),
        ("A009", "递延所得税调整", "借", "所得税费用-递延所得税", "合并", 165000.00, "对应存货未实现利润产生的递延所得税资产"),
        ("A009", "递延所得税调整", "贷", "递延所得税资产", "合并", 165000.00, "对应递延所得税资产确认"),
        ("A010", "少数股东损益", "借", "少数股东损益", "ABC商贸", 0.00, "ABC商贸为全资子公司，无少数股东"),
        ("A011", "投资收益与净资产抵销", "借", "投资收益", "ABC集团", 28540000.00, "ABC集团 对子公司投资收益抵销"),
        ("A011", "投资收益与净资产抵销", "贷", "长期股权投资", "ABC集团", 28540000.00, "对应长期股权投资权益法调整"),
    ]
    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c == 6 and isinstance(val, (int, float)):
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 12, "B": 22, "C": 8, "D": 24, "E": 22, "F": 18, "G": 50})


def _xlsx_bank_statement(wb, item, doc_name):
    """银-2 银行对账单。"""
    ws = wb.create_sheet("银行对账单")
    ws["A1"] = "银行对账单"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"账户名：{item['实体归属']}　　开户行：中国工商银行广州天河支行　　账号：4400123401001234567　　期间：2025年"
    ws["A2"].font = Font(size=10)
    headers = ["交易日期", "摘要", "借方(支出)", "贷方(收入)", "余额", "对方账户", "对方户名"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    rows = []
    balance = 5840230.50
    banks = ["工行", "建行", "招行", "农行", "中行", "广发", "中信"]
    counterparties = ["ABC集团控股", "ABC科技", "ABC制造", "ABC商贸", "某供应商A", "某客户B", "某客户C", "税务"]
    for i in range(1, 41):
        day = RNG.randint(1, 28)
        month = (i - 1) // 4 + 1
        if month > 12:
            month = 12
        date = f"2025-{month:02d}-{day:02d}"
        direction = RNG.choice(["in", "out"])
        amount = round(RNG.uniform(50000, 800000), 2)
        if direction == "in":
            balance += amount
            debit, credit = "", amount
        else:
            balance -= amount
            debit, credit = amount, ""
        cp_bank = RNG.choice(banks)
        cp_acct = f"622848{RNG.randint(10000000, 99999999)}"
        cp_name = RNG.choice(counterparties)
        summary = RNG.choice([
            "销售回款", "采购付款", "工资发放", "税款缴纳", "费用报销",
            "借款到账", "还款支出", "利息收入", "服务费支出", "押金收取"
        ])
        rows.append((date, summary, debit, credit, round(balance, 2), cp_acct, cp_name))

    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c in (3, 4, 5):
                if val != "":
                    ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
                else:
                    ws.cell(row=r, column=c, value="")
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 14, "B": 16, "C": 14, "D": 14, "E": 18, "F": 22, "G": 16})

    # Sheet2: 银行存款余额调节表
    ws2 = wb.create_sheet("余额调节表")
    ws2["A1"] = "银行存款余额调节表"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = f"账户：{item['实体归属']} 4400123401001234567　　调节日：{PERIOD}"
    headers2 = ["调节项", "金额(元)", "说明"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=4, column=c, value=h)
    _style_header_row(ws2, row=4, n_cols=3)
    recon_rows = [
        ("企业账面余额", 12485620.30, "ABC科技 银行存款日记账余额"),
        ("加：银行已收企业未收", 235800.00, "客户 B 12月31日到账，企业1月入账"),
        ("减：银行已付企业未付", 18520.00, "银行扣收 12月服务费"),
        ("调节后存款余额", 12702900.30, "应等于银行对账单余额"),
        ("", "", ""),
        ("银行对账单余额", 12702900.30, "银行出具的对账单余额"),
        ("加：企业已收银行未收", 0.00, "无"),
        ("减：企业已付银行未付", 0.00, "无"),
        ("调节后存款余额", 12702900.30, "应等于企业账面调节后余额"),
    ]
    for r, row in enumerate(recon_rows, start=5):
        for c, val in enumerate(row, start=1):
            if c == 2 and isinstance(val, (int, float)):
                ws2.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws2.cell(row=r, column=c, value=val)
    _autofit(ws2, {"A": 28, "B": 20, "C": 40})


def _xlsx_salary_detail(wb, item, doc_name):
    """薪-1 工资明细表（员工编号维度，已脱敏）。"""
    ws = wb.create_sheet("工资明细表")
    ws["A1"] = "工资明细表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　期间：2025年12月　　单位：元（已脱敏，仅员工编号）"
    headers = ["员工编号", "部门", "岗位类别", "基本工资", "绩效奖金", "社保个人", "公积金个人", "个税", "实发工资"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    depts = ["研发部", "销售部", "生产部", "财务部", "行政部", "采购部", "人力资源部"]
    posts = [("高管", 30000, 80000), ("经理", 15000, 35000),
             ("主管", 10000, 22000), ("专员", 6000, 15000), ("助理", 5000, 10000)]
    rows = []
    for i in range(1, 51):
        emp_id = f"E2025{i:04d}"
        dept = RNG.choice(depts)
        post_name, base_min, base_max = RNG.choice(posts)
        base = round(RNG.uniform(base_min, base_max), 2)
        bonus = round(base * RNG.uniform(0.1, 0.5), 2)
        social = round(base * 0.105, 2)
        housing = round(base * 0.07, 2)
        taxable = base + bonus - social - housing - 5000
        if taxable > 90000:
            tax = round(taxable * 0.45 - 181920, 2)
        elif taxable > 30000:
            tax = round(taxable * 0.25 - 2660, 2)
        elif taxable > 12000:
            tax = round(taxable * 0.20 - 1410, 2)
        elif taxable > 3000:
            tax = round(taxable * 0.10 - 210, 2)
        else:
            tax = max(0, round(taxable * 0.03, 2))
        actual = round(base + bonus - social - housing - tax, 2)
        rows.append((emp_id, dept, post_name, base, bonus, social, housing, tax, actual))

    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c >= 4:
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 12, "B": 14, "C": 12, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14})

    # 合计行
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    for c in range(4, 10):
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(bold=True)


def _xlsx_deferred_tax(wb, item, doc_name):
    """税-9 递延所得税计算表。"""
    ws = wb.create_sheet("递延所得税计算表")
    ws["A1"] = "递延所得税计算表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　截止日：{PERIOD}　　单位：元"
    headers = ["项目", "账面价值", "计税基础", "可抵扣暂时性差异", "应纳税暂时性差异", "税率", "递延所得税资产", "递延所得税负债"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    rows = [
        ("资产减值准备-应收账款", 5250000.00, 0.00, 5250000.00, 0.00, 0.25, 1312500.00, 0.00),
        ("资产减值准备-存货", 1840000.00, 0.00, 1840000.00, 0.00, 0.25, 460000.00, 0.00),
        ("固定资产加速折旧", 268540200.00, 265200000.00, 3340200.00, 0.00, 0.25, 835050.00, 0.00),
        ("无形资产摊销差异", 45200000.00, 44500000.00, 700000.00, 0.00, 0.25, 175000.00, 0.00),
        ("内部交易未实现利润", 540000.00, 0.00, 0.00, 540000.00, 0.25, 0.00, 135000.00),
        ("可弥补亏损", 0.00, 0.00, 0.00, 0.00, 0.25, 0.00, 0.00),
        ("递延收益(政府补助)", 6250000.00, 0.00, 6250000.00, 0.00, 0.25, 1562500.00, 0.00),
        ("股权激励费用", 8500000.00, 0.00, 8500000.00, 0.00, 0.25, 2125000.00, 0.00),
        ("公允价值变动-交易性金融资产", 12000000.00, 12000000.00, 0.00, 0.00, 0.25, 0.00, 0.00),
        ("预计负债-未决诉讼", 2500000.00, 0.00, 2500000.00, 0.00, 0.25, 625000.00, 0.00),
    ]
    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c == 1:
                ws.cell(row=r, column=c, value=val)
            elif c == 6:
                ws.cell(row=r, column=c, value=val).number_format = '0.00%'
            else:
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'

    # 合计行
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    for c in [2, 3, 4, 5, 7, 8]:
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(bold=True)
    _autofit(ws, {"A": 32, "B": 16, "C": 16, "D": 18, "E": 18, "F": 8, "G": 18, "H": 18})

    # Sheet2: 调节表
    ws2 = wb.create_sheet("递延所得税调节表")
    ws2["A1"] = "递延所得税期初余额调节表"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = f"编制单位：{item['实体归属']}　　2025年度　　单位：元"
    headers2 = ["项目", "递延所得税资产", "递延所得税负债"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=4, column=c, value=h)
    _style_header_row(ws2, row=4, n_cols=3)
    recon = [
        ("一、期初余额", 4520000.00, 540000.00),
        ("二、本期增加", 3574550.00, 135000.00),
        ("三、本期减少（转回）", 1250000.00, 540000.00),
        ("四、期末余额", 6844550.00, 135000.00),
    ]
    for r, row in enumerate(recon, start=5):
        for c, val in enumerate(row, start=1):
            if c >= 2:
                ws2.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws2.cell(row=r, column=c, value=val)
    _autofit(ws2, {"A": 28, "B": 22, "C": 22})


def _xlsx_inventory_detail(wb, item, doc_name):
    """存-1 存货明细表。"""
    ws = wb.create_sheet("存货明细表")
    ws["A1"] = "存货明细表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　截止日：{PERIOD}　　单位：元"
    headers = ["存货编号", "存货类别", "存货名称", "规格型号", "计量单位", "数量", "单价", "金额", "仓库", "存放位置"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    raw_materials = [("钢材-A3", "吨"), ("铝材-6061", "吨"), ("塑料-ABS", "千克"),
                     ("电子元件-电容", "个"), ("包装材料-纸箱", "个")]
    wips = [("半成品-主板A", "块"), ("半成品-机箱B", "台")]
    finished_goods = [("产成品-智能终端X1", "台"), ("产成品-终端Y2", "台"), ("产成品-模块Z3", "块")]
    warehouses = [("原材料仓", "A区"), ("半成品仓", "B区"), ("库存商品仓", "C区")]

    rows = []
    for i in range(1, 31):
        cat_idx = (i - 1) // 10
        if cat_idx == 0:
            cat, name, unit = "原材料", *raw_materials[i % len(raw_materials)]
            wh, pos = warehouses[0]
            qty = RNG.randint(100, 5000)
            price = round(RNG.uniform(5, 500), 2)
        elif cat_idx == 1:
            cat, name, unit = "半成品", *wips[i % len(wips)]
            wh, pos = warehouses[1]
            qty = RNG.randint(50, 500)
            price = round(RNG.uniform(200, 2000), 2)
        else:
            cat, name, unit = "库存商品", *finished_goods[i % len(finished_goods)]
            wh, pos = warehouses[2]
            qty = RNG.randint(20, 300)
            price = round(RNG.uniform(800, 5000), 2)
        sn = f"INV-{cat_idx}{i:04d}"
        amt = round(qty * price, 2)
        rows.append((sn, cat, name, "标准型", unit, qty, price, amt, wh, pos))

    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c in (7, 8):
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 14, "B": 10, "C": 22, "D": 12, "E": 10, "F": 10, "G": 12, "H": 16, "I": 14, "J": 10})

    # 合计行
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=f"=SUM(H5:H{total_row-1})")
    ws.cell(row=total_row, column=8).number_format = '#,##0.00'
    ws.cell(row=total_row, column=8).font = Font(bold=True)


def _xlsx_material_shelf_life(wb, item, doc_name):
    """存-5 原材料保质期台账。"""
    ws = wb.create_sheet("保质期台账")
    ws["A1"] = "原材料保质期台账"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　{PERIOD}　　单位：天/元"
    headers = ["原材料编号", "原材料名称", "批次", "入库日期", "保质期(天)", "到期日", "库存数量", "库存金额", "状态", "处置建议"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    materials = [("M001", "钢材-A3", "无保质期限制"), ("M002", "铝材-6061", "无保质期限制"),
                 ("M003", "塑料-ABS", "保质期 365 天"), ("M004", "电子元件-电容", "保质期 730 天"),
                 ("M005", "包装材料-纸箱", "保质期 180 天"), ("M006", "胶水-工业胶", "保质期 90 天"),
                 ("M007", "油漆-喷涂漆", "保质期 365 天"), ("M008", "润滑油", "保质期 1095 天")]

    rows = []
    for i in range(1, 21):
        mid, mname, shelf_text = materials[i % len(materials)]
        batch = f"B2025{RNG.randint(1, 12):02d}{RNG.randint(1, 28):02d}-{RNG.randint(100, 999)}"
        in_date_month = RNG.randint(1, 12)
        in_date_day = RNG.randint(1, 28)
        in_date = f"2025-{in_date_month:02d}-{in_date_day:02d}"
        # 计算保质期天数
        if "无保质期" in shelf_text:
            shelf = 0
            expire = "无"
        else:
            shelf = int(shelf_text.replace("保质期", "").replace("天", "").strip())
            expire_date = datetime(2025, in_date_month, in_date_day)
            expire_dt = expire_date.fromordinal(expire_date.toordinal() + shelf)
            expire = expire_dt.strftime("%Y-%m-%d")

        qty = RNG.randint(50, 2000)
        price = round(RNG.uniform(5, 500), 2)
        amt = round(qty * price, 2)
        # 状态判断
        if shelf == 0:
            status = "正常"
            suggest = "无需特别处理"
        else:
            days_to_expire = (datetime(2025, 12, 31) - datetime.strptime(in_date, "%Y-%m-%d")).days - 0
            # 用到期日比较
            try:
                expire_dt = datetime.strptime(expire, "%Y-%m-%d")
                days_left = (expire_dt - datetime(2025, 12, 31)).days
                if days_left < 0:
                    status = "已过期"
                    suggest = "立即报废处理"
                elif days_left < 30:
                    status = "临期"
                    suggest = "优先使用或退货"
                elif days_left < 90:
                    status = "短期"
                    suggest = "加强监控"
                else:
                    status = "正常"
                    suggest = "正常使用"
            except Exception:
                status = "正常"
                suggest = "无"

        rows.append((mid, mname, batch, in_date, shelf, expire, qty, amt, status, suggest))

    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c == 8:
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws.cell(row=r, column=c, value=val)

    # 状态列条件格式
    status_col = "I"
    n = len(rows)
    rng = f"{status_col}5:{status_col}{n + 4}"
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{status_col}5="已过期"'],
                    fill=PatternFill("solid", start_color="FFC7CE"),
                    font=Font(color="9C0006")))
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{status_col}5="临期"'],
                    fill=PatternFill("solid", start_color="FFEB9C"),
                    font=Font(color="9C6500")))
    ws.conditional_formatting.add(rng,
        FormulaRule(formula=[f'{status_col}5="短期"'],
                    fill=PatternFill("solid", start_color="DDEBF7"),
                    font=Font(color="1F4E78")))

    _autofit(ws, {"A": 14, "B": 18, "C": 22, "D": 12, "E": 12, "F": 12, "G": 10, "H": 14, "I": 10, "J": 22})


def _xlsx_fixed_assets(wb, item, doc_name):
    """长期资产-4 固定资产明细表。"""
    ws = wb.create_sheet("固定资产明细表")
    ws["A1"] = "固定资产明细表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　截止日：{PERIOD}　　单位：元"
    headers = ["资产编号", "资产名称", "资产类别", "规格型号", "计量单位", "购入日期", "原值", "累计折旧", "净值", "折旧年限", "残值率"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    cats = [
        ("房屋及建筑物", "厂房A", "栋", 5200000.00, 20, 0.05),
        ("房屋及建筑物", "办公楼", "栋", 3800000.00, 20, 0.05),
        ("机器设备", "数控车床", "台", 850000.00, 10, 0.05),
        ("机器设备", "注塑机", "台", 1250000.00, 10, 0.05),
        ("机器设备", " CNC加工中心", "台", 2100000.00, 10, 0.05),
        ("机器设备", "激光切割机", "台", 680000.00, 10, 0.05),
        ("机器设备", "检测设备A", "套", 320000.00, 10, 0.05),
        ("运输设备", "叉车", "台", 180000.00, 5, 0.05),
        ("运输设备", "货车", "台", 280000.00, 5, 0.05),
        ("电子设备", "服务器机柜", "套", 450000.00, 5, 0.05),
        ("电子设备", "电脑", "台", 8500.00, 3, 0.05),
        ("办公设备", "打印机", "台", 5200.00, 3, 0.05),
        ("办公设备", "空调", "台", 12000.00, 5, 0.05),
    ]

    rows = []
    for i, (cat, name, unit, origin, life, residual) in enumerate(cats, start=1):
        aid = f"FA-{i:04d}"
        # 购入日期在 2018-2024 之间
        buy_year = RNG.randint(2018, 2024)
        buy_month = RNG.randint(1, 12)
        buy_day = RNG.randint(1, 28)
        buy_date = f"{buy_year}-{buy_month:02d}-{buy_day:02d}"
        # 累计折旧（简化：按月数 × 月折旧额）
        months = (2025 - buy_year) * 12 + (12 - buy_month)
        monthly_dep = origin * (1 - residual) / (life * 12)
        acc_dep = round(min(monthly_dep * months, origin * (1 - residual)), 2)
        net_value = round(origin - acc_dep, 2)
        rows.append((aid, name, cat, "标准型", unit, buy_date, origin, acc_dep, net_value, life, residual))

    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c in (7, 8, 9):
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            elif c == 11:
                ws.cell(row=r, column=c, value=val).number_format = '0.00%'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 12, "B": 18, "C": 16, "D": 12, "E": 10, "F": 14, "G": 16, "H": 16, "I": 16, "J": 10, "K": 10})

    # 合计行
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    for c in [7, 8, 9]:
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(bold=True)


def _xlsx_lease_calc(wb, item, doc_name):
    """租-2 使用权资产和租赁负债计算表。"""
    ws = wb.create_sheet("租赁负债计算表")
    ws["A1"] = "使用权资产与租赁负债计算表"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　截止日：{PERIOD}　　单位：元"
    headers = ["租赁合同编号", "租赁资产", "租赁期(月)", "月租金", "折现率(月)", "租赁付款额现值", "使用权资产原值", "累计摊销", "使用权资产净值", "租赁负债余额"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    leases = [
        ("L2023-001", "厂房A租赁", 60, 250000.00, 0.004),
        ("L2023-002", "办公楼租赁", 36, 80000.00, 0.004),
        ("L2024-005", "仓库租赁", 24, 35000.00, 0.004),
        ("L2024-008", "店面租赁", 36, 60000.00, 0.004),
        ("L2025-002", "设备租赁", 12, 18000.00, 0.004),
    ]

    rows = []
    for aid, asset, months, monthly, rate in leases:
        # 计算现值（年金现值公式）
        pv = monthly * (1 - (1 + rate) ** (-months)) / rate
        pv = round(pv, 2)
        # 假设到 2025-12-31 已使用 1/3 期限
        elapsed = months // 3
        # 累计摊销（简化：直线法按月）
        monthly_amort = pv / months
        acc_amort = round(monthly_amort * elapsed, 2)
        net_rou = round(pv - acc_amort, 2)
        # 租赁负债余额 = 剩余租金的现值
        remaining = months - elapsed
        lease_liab = round(monthly * (1 - (1 + rate) ** (-remaining)) / rate, 2)
        rows.append((aid, asset, months, monthly, rate, pv, pv, acc_amort, net_rou, lease_liab))

    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c in (4, 6, 7, 8, 9, 10):
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            elif c == 5:
                ws.cell(row=r, column=c, value=val).number_format = '0.0000'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 14, "B": 18, "C": 12, "D": 14, "E": 12, "F": 18, "G": 18, "H": 14, "I": 18, "J": 18})

    # 合计行
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    for c in [6, 7, 8, 9, 10]:
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(bold=True)

    # Sheet2: 租赁明细
    ws2 = wb.create_sheet("租赁合同明细")
    ws2["A1"] = "租赁合同明细"
    headers2 = ["合同编号", "出租方", "租赁资产", "租赁起始日", "租赁终止日", "月租金", "免租期", "押金"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=3, column=c, value=h)
    _style_header_row(ws2, row=3, n_cols=len(headers2))
    lease_detail = [
        ("L2023-001", "广州市黄埔工业区管委会", "厂房A", "2023-01-01", "2027-12-31", 250000, "无", 500000),
        ("L2023-002", "广州珠江物业管理有限公司", "办公楼 5F", "2023-06-01", "2026-05-31", 80000, "1个月", 160000),
        ("L2024-005", "广州仓储服务有限公司", "仓库 B-12", "2024-03-01", "2026-02-28", 35000, "无", 70000),
        ("L2024-008", "广州天河商业管理公司", "店面 102", "2024-09-01", "2027-08-31", 60000, "无", 180000),
        ("L2025-002", "三一重工融资租赁公司", "数控机床", "2025-02-01", "2026-01-31", 18000, "无", 0),
    ]
    for r, row in enumerate(lease_detail, start=4):
        for c, val in enumerate(row, start=1):
            if c in (6, 8):
                ws2.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws2.cell(row=r, column=c, value=val)
    _autofit(ws2, {"A": 14, "B": 30, "C": 18, "D": 14, "E": 14, "F": 14, "G": 10, "H": 14})


def _xlsx_ar_aging(wb, item, doc_name):
    """往来-1 应收账款明细表及账龄。"""
    ws = wb.create_sheet("应收账款明细表")
    ws["A1"] = "应收账款明细表及账龄"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　截止日：{PERIOD}　　单位：元"
    headers = ["客户代码", "客户名称", "客户性质", "所属集团", "期初余额", "本期借方", "本期贷方", "期末余额", "1年以内", "1-2年", "2-3年", "3年以上", "坏账准备"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    customers = [
        ("C001", "某客户甲有限公司", "民营企业", "无集团", 1250000.00, 3580000.00, 2830000.00),
        ("C002", "某客户乙有限公司", "国有企业", "某集团A", 2800000.00, 5200000.00, 3100000.00),
        ("C003", "某客户丙股份公司", "上市公司", "某集团B", 980000.00, 2450000.00, 1890000.00),
        ("C004", "某客户丁有限公司", "外资企业", "无集团", 5250000.00, 8200000.00, 6800000.00),
        ("C005", "某客户戊有限公司", "民营企业", "无集团", 0.00, 1820000.00, 1520000.00),
        ("C006", "某客户己股份公司", "民营企业", "某集团C", 1580000.00, 3250000.00, 1580000.00),
        ("C007", "某客户庚有限公司", "国有企业", "某集团D", 2100000.00, 0.00, 0.00),
        ("C008", "某客户辛有限公司", "民营企业", "无集团", 320000.00, 850000.00, 720000.00),
        ("C009", "某客户壬有限公司", "外资企业", "无集团", 1120000.00, 2380000.00, 1680000.00),
        ("C010", "某客户癸有限公司", "民营企业", "无集团", 0.00, 580000.00, 380000.00),
    ]

    rows = []
    for cid, cname, cnature, group, begin, add, reduce in customers:
        end = round(begin + add - reduce, 2)
        # 账龄分布（简化）
        if end == 0:
            under1, between1_2, between2_3, over3 = 0, 0, 0, 0
        elif cid in ("C007",):
            # 长账龄
            under1 = round(end * 0.2, 2)
            between1_2 = round(end * 0.3, 2)
            between2_3 = round(end * 0.3, 2)
            over3 = round(end - under1 - between1_2 - between2_3, 2)
        elif cid in ("C003",):
            under1 = round(end * 0.7, 2)
            between1_2 = round(end * 0.3, 2)
            between2_3 = 0
            over3 = 0
        else:
            under1 = round(end * 0.85, 2)
            between1_2 = round(end * 0.15, 2)
            between2_3 = 0
            over3 = 0
        # 坏账准备（按账龄计提：1年内5%、1-2年10%、2-3年30%、3年以上100%）
        bad = round(under1 * 0.05 + between1_2 * 0.10 + between2_3 * 0.30 + over3 * 1.00, 2)
        rows.append((cid, cname, cnature, group, begin, add, reduce, end,
                    under1, between1_2, between2_3, over3, bad))

    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c >= 5 and isinstance(val, (int, float)):
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 10, "B": 22, "C": 12, "D": 12, "E": 14, "F": 14, "G": 14, "H": 14, "I": 12, "J": 12, "K": 12, "L": 12, "M": 14})

    # 合计行
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    for c in range(5, 14):
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(bold=True)


def _xlsx_ap_aging(wb, item, doc_name):
    """往来-4 应付账款明细表及账龄。"""
    ws = wb.create_sheet("应付账款明细表")
    ws["A1"] = "应付账款明细表及账龄"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　截止日：{PERIOD}　　单位：元"
    headers = ["供应商代码", "供应商名称", "供应商性质", "所属集团", "期初余额", "本期借方", "本期贷方", "期末余额", "1年以内", "1-2年", "2-3年", "3年以上"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    suppliers = [
        ("S001", "某供应商甲有限公司", "民营企业", "无集团", 1580000.00, 2830000.00, 3580000.00),
        ("S002", "某供应商乙股份公司", "国有企业", "某集团A", 2250000.00, 5100000.00, 4200000.00),
        ("S003", "某供应商丙有限公司", "外资企业", "无集团", 980000.00, 1450000.00, 1890000.00),
        ("S004", "某供应商丁有限公司", "民营企业", "无集团", 3250000.00, 4200000.00, 2800000.00),
        ("S005", "某供应商戊有限公司", "民营企业", "无集团", 0.00, 1820000.00, 1520000.00),
        ("S006", "某供应商己股份公司", "民营企业", "某集团C", 1580000.00, 2250000.00, 1580000.00),
        ("S007", "某供应商庚有限公司", "国有企业", "某集团D", 0.00, 0.00, 0.00),
        ("S008", "某供应商辛有限公司", "民营企业", "无集团", 320000.00, 650000.00, 720000.00),
        ("S009", "某供应商壬有限公司", "外资企业", "无集团", 1120000.00, 2380000.00, 1680000.00),
        ("S010", "某供应商癸有限公司", "民营企业", "无集团", 0.00, 580000.00, 380000.00),
    ]

    rows = []
    for sid, sname, snature, group, begin, debit, credit in suppliers:
        end = round(begin - debit + credit, 2)
        if end == 0:
            under1, between1_2, between2_3, over3 = 0, 0, 0, 0
        else:
            under1 = round(end * 0.78, 2)
            between1_2 = round(end * 0.18, 2)
            between2_3 = round(end * 0.04, 2)
            over3 = round(end - under1 - between1_2 - between2_3, 2)
        rows.append((sid, sname, snature, group, begin, debit, credit, end,
                    under1, between1_2, between2_3, over3))

    for r, row in enumerate(rows, start=5):
        for c, val in enumerate(row, start=1):
            if c >= 5 and isinstance(val, (int, float)):
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 10, "B": 22, "C": 12, "D": 12, "E": 14, "F": 14, "G": 14, "H": 14, "I": 12, "J": 12, "K": 12, "L": 12})

    # 合计行
    total_row = len(rows) + 5
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    for c in range(5, 13):
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(bold=True)


def _xlsx_loan_ledger(wb, item, doc_name):
    """借-2 借款台账及利息计算表。"""
    ws = wb.create_sheet("借款台账")
    ws["A1"] = "银行借款台账"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"编制单位：{item['实体归属']}　　截止日：{PERIOD}　　单位：元"
    headers = ["借款编号", "合同编号", "贷款银行", "借款类别", "借款金额", "借款起日", "借款止日", "期限(月)", "年利率", "已还本金", "未还本金", "担保方式"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header_row(ws, row=4, n_cols=len(headers))

    loans = [
        ("L2023-001", "RL-2023-001", "工行广州天河支行", "流动资金贷款", 30000000.00, "2023-06-30", "2026-06-29", 36, 0.0435, 0.00, 30000000.00, "集团担保"),
        ("L2024-005", "RL-2024-005", "建行广州珠江支行", "流动资金贷款", 50000000.00, "2024-03-15", "2027-03-14", 36, 0.0420, 0.00, 50000000.00, "集团担保"),
        ("L2024-008", "RL-2024-008", "招行广州分行营业部", "流动资金贷款", 20000000.00, "2024-09-10", "2025-09-09", 12, 0.0410, 20000000.00, 0.00, "信用借款"),
        ("L2025-002", "RL-2025-002", "农行广州分行", "项目贷款", 45000000.00, "2025-02-20", "2028-02-19", 36, 0.0385, 0.00, 45000000.00, "集团担保"),
        ("L2025-005", "RL-2025-005", "中行广州分行", "流动资金贷款", 15000000.00, "2025-07-15", "2026-07-14", 12, 0.0395, 0.00, 15000000.00, "信用借款"),
    ]

    for r, row in enumerate(loans, start=5):
        for c, val in enumerate(row, start=1):
            if c in (5, 10, 11):
                ws.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            elif c == 9:
                ws.cell(row=r, column=c, value=val).number_format = '0.00%'
            else:
                ws.cell(row=r, column=c, value=val)
    _autofit(ws, {"A": 14, "B": 14, "C": 22, "D": 16, "E": 16, "F": 12, "G": 12, "H": 10, "I": 10, "J": 14, "K": 14, "L": 12})

    # 合计行
    total_row = len(loans) + 5
    ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    for c in [5, 10, 11]:
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}5:{col_letter}{total_row-1})")
        ws.cell(row=total_row, column=c).number_format = '#,##0.00'
        ws.cell(row=total_row, column=c).font = Font(bold=True)

    # Sheet2: 利息计算表
    ws2 = wb.create_sheet("利息计算表")
    ws2["A1"] = "借款利息计算表"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = f"编制单位：{item['实体归属']}　　2025年度　　单位：元"
    headers2 = ["借款编号", "期次", "计息起始日", "计息终止日", "计息天数", "计息本金", "年利率", "利息金额"]
    for c, h in enumerate(headers2, start=1):
        ws2.cell(row=4, column=c, value=h)
    _style_header_row(ws2, row=4, n_cols=len(headers2))

    # 为每笔借款生成 4 个季度的利息计算
    rows = []
    loan_data = [
        ("L2023-001", "2025-01-01", "2025-03-31", 30_000_000, 0.0435),
        ("L2023-001", "2025-04-01", "2025-06-30", 30_000_000, 0.0435),
        ("L2023-001", "2025-07-01", "2025-09-30", 30_000_000, 0.0435),
        ("L2023-001", "2025-10-01", "2025-12-31", 30_000_000, 0.0435),
        ("L2024-005", "2025-01-01", "2025-03-31", 50_000_000, 0.0420),
        ("L2024-005", "2025-04-01", "2025-06-30", 50_000_000, 0.0420),
        ("L2024-005", "2025-07-01", "2025-09-30", 50_000_000, 0.0420),
        ("L2024-005", "2025-10-01", "2025-12-31", 50_000_000, 0.0420),
        ("L2024-008", "2025-01-01", "2025-03-31", 20_000_000, 0.0410),
        ("L2024-008", "2025-04-01", "2025-06-30", 20_000_000, 0.0410),
        ("L2024-008", "2025-07-01", "2025-09-09", 20_000_000, 0.0410),
        ("L2025-002", "2025-02-20", "2025-03-31", 45_000_000, 0.0385),
        ("L2025-002", "2025-04-01", "2025-06-30", 45_000_000, 0.0385),
        ("L2025-002", "2025-07-01", "2025-09-30", 45_000_000, 0.0385),
        ("L2025-002", "2025-10-01", "2025-12-31", 45_000_000, 0.0385),
        ("L2025-005", "2025-07-15", "2025-09-30", 15_000_000, 0.0395),
        ("L2025-005", "2025-10-01", "2025-12-31", 15_000_000, 0.0395),
    ]
    for r, (lid, start, end, principal, rate) in enumerate(loan_data, start=5):
        s_dt = datetime.strptime(start, "%Y-%m-%d")
        e_dt = datetime.strptime(end, "%Y-%m-%d")
        days = (e_dt - s_dt).days + 1
        interest = round(principal * rate * days / 360, 2)
        for c, val in enumerate([lid, f"2025-Q{(r-5)//4+1}", start, end, days, principal, rate, interest], start=1):
            if c in (6, 8):
                ws2.cell(row=r, column=c, value=val).number_format = '#,##0.00'
            elif c == 7:
                ws2.cell(row=r, column=c, value=val).number_format = '0.00%'
            else:
                ws2.cell(row=r, column=c, value=val)
    _autofit(ws2, {"A": 14, "B": 14, "C": 14, "D": 14, "E": 12, "F": 18, "G": 10, "H": 16})

    # 合计行
    total_row = len(loan_data) + 5
    ws2.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=total_row, column=8, value=f"=SUM(H5:H{total_row-1})")
    ws2.cell(row=total_row, column=8).number_format = '#,##0.00'
    ws2.cell(row=total_row, column=8).font = Font(bold=True)


# Excel 内容生成器注册表
XLSX_CONTENT_GENERATORS = {
    "consolidated_fs": _xlsx_consolidated_fs,
    "consolidation_adj": _xlsx_consolidation_adj,
    "bank_statement": _xlsx_bank_statement,
    "salary_detail": _xlsx_salary_detail,
    "deferred_tax": _xlsx_deferred_tax,
    "inventory_detail": _xlsx_inventory_detail,
    "material_shelf_life": _xlsx_material_shelf_life,
    "fixed_assets": _xlsx_fixed_assets,
    "lease_calc": _xlsx_lease_calc,
    "ar_aging": _xlsx_ar_aging,
    "ap_aging": _xlsx_ap_aging,
    "loan_ledger": _xlsx_loan_ledger,
}


# ============================================================
# 5. 主流程
# ============================================================
def clean_output_dir():
    """清理输出目录（保留目录骨架）。
    注意：沙箱环境的 shutil.rmtree 受限，改用 os.walk + os.remove。
    """
    # 清空 客户共享文件夹/ 下所有实体目录的内容（保留目录本身）
    for entity_dir in ENTITY_DIR.values():
        d = os.path.join(OUT_FILES_DIR, entity_dir)
        os.makedirs(d, exist_ok=True)
        for root, dirs, files in os.walk(d, topdown=False):
            for f in files:
                try:
                    os.remove(os.path.join(root, f))
                except OSError:
                    pass
            for sub in dirs:
                try:
                    os.rmdir(os.path.join(root, sub))
                except OSError:
                    pass
    # 删除旧的清单文件和 JSON
    for f in [OUT_LIST_XLSX, OUT_JSON]:
        if os.path.exists(f):
            try:
                os.remove(f)
            except OSError:
                pass


def main():
    print("=" * 70)
    print("IPO 审计 PBC 智能管理工作站 · 模拟数据生成器")
    print("=" * 70)
    print(f"真实清单: {REAL_XLSX}")
    print(f"输出目录: {OUT_DIR}")
    print()

    # 清理输出
    print("[1/5] 清理输出目录...")
    clean_output_dir()
    for entity_dir in ENTITY_DIR.values():
        os.makedirs(os.path.join(OUT_FILES_DIR, entity_dir), exist_ok=True)
    print("      已清理")
    print()

    # 解析真实清单
    print("[2/5] 解析真实清单 Excel...")
    items = parse_real_list()
    print(f"      解析得到 {len(items)} 条目")
    if len(items) != 103:
        print(f"      [WARN] 预期 103 条，实际 {len(items)} 条")
    print()

    # 整合 14 列
    print("[3/5] 整合 14 列统一格式（含实体归属/置信度）...")
    items = extend_items(items)
    # 状态分布
    status_dist = Counter(it["资料提供情况"] for it in items)
    print(f"      状态分布:")
    for s, n in status_dist.most_common():
        print(f"        {s}: {n}")
    entity_dist = Counter(it["实体归属"] for it in items)
    print(f"      实体分布:")
    for e, n in entity_dist.most_common():
        print(f"        {e}: {n}")
    cat_dist = Counter(it["一级分类"] for it in items)
    print(f"      一级分类分布:")
    for c, n in cat_dist.most_common():
        print(f"        {c}: {n}")
    print()

    # 保存中间 JSON
    print("[4/5] 保存中间产物 _real_items.json...")
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    print(f"      已保存: {OUT_JSON}")
    print()

    # 生成模拟文件 + 回填路径
    print("[5/5] 生成模拟文件 + 回填文件路径...")
    generated = generate_mock_files(items)
    print(f"      共生成 {len(generated)} 个文件")
    print()

    # 生成 01_PBC_List.xlsx
    print("生成 01_PBC_List.xlsx...")
    write_pbc_list(items, OUT_LIST_XLSX, with_paths=True)
    print(f"      已保存: {OUT_LIST_XLSX}")
    print()

    # 抽样检查
    print("=" * 70)
    print("生成完成 · 抽样检查")
    print("=" * 70)
    # 检查文件大小
    total_size = 0
    print("\n生成的模拟文件清单:")
    for rid, name, ext, rel_path in generated:
        abs_path = os.path.join(OUT_DIR, rel_path)
        size = os.path.getsize(abs_path)
        total_size += size
        print(f"  {rid:14s} {ext:4s} {size:>8,d} bytes  {rel_path}")
    print(f"\n文件总大小: {total_size:,d} bytes ({total_size/1024:.1f} KB)")

    # 重新加载清单检查
    print("\n重新加载 01_PBC_List.xlsx 检查:")
    wb = openpyxl.load_workbook(OUT_LIST_XLSX)
    ws = wb.active
    print(f"  Sheet: {ws.title}")
    print(f"  数据行数: {ws.max_row - 1} (期望 103)")
    print(f"  数据列数: {ws.max_column} (期望 14)")
    print(f"  表头: {[ws.cell(1, c).value for c in range(1, ws.max_column + 1)]}")
    # 抽样第 1 行和最后一行
    print(f"  第 1 行数据: {[ws.cell(2, c).value for c in range(1, ws.max_column + 1)]}")
    print(f"  最后 1 行数据: {[ws.cell(ws.max_row, c).value for c in range(1, ws.max_column + 1)]}")
    # 文件路径回填数
    fp_col = COLUMNS.index("文件路径") + 1
    filled = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, fp_col).value)
    print(f"  文件路径已回填: {filled} / {ws.max_row - 1}")
    wb.close()


if __name__ == "__main__":
    main()
